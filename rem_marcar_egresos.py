#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Limpieza y marcado de egresos de Salud Mental (export RAYEN) para REM.

  1. Recorta filas de cabecera (banner + filtros).
  2. Convierte 'AÑO APLICACIÓN FORMULARIO' a número.
  3. Por cada egreso (Alta / Traslado / Otras Causas) agrega columnas con la
     patología y el subtipo. Reconoce el diagnóstico aunque el subtipo esté a
     la izq o der del ESTADO.
  4. Quita el "inmovilizar paneles" heredado del export.

USO:
  - Arrastra el .xlsx ENCIMA de este .py.
  - Doble-clic y pega la ruta.
  - Terminal:  python rem_marcar_egresos.py entrada.xlsx [salida.xlsx]
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("\nERROR: falta 'openpyxl'. Instálalo con:\n    pip install openpyxl\n")
    try: input("Presiona Enter para salir...")
    except EOFError: pass
    sys.exit(1)

# ╔═══════════════════════════════════════════════════════════════════╗
# ║  ZONA DE CONFIGURACIÓN CLÍNICA  — editar aquí si cambia el form     ║
# ╚═══════════════════════════════════════════════════════════════════╝

# Egresos a marcar. Alta y Traslado van a estadística; Otras Causas se flaggea
# para revisión MANUAL (decidir abandono vs clínica caso a caso).
BUSQUEDAS = {
    "Alta":        ["EGRESO", "ALTA"],
    "Traslado":    ["EGRESO", "TRASLADO"],
    "OtrasCausas": ["EGRESO", "OTRAS", "CAUSAS"],
}

# Diagnósticos QUE TIENEN SUBTIPO.
#   clave = N.- de la pregunta del diagnóstico
#   valor = N.- de la pregunta de su columna de subtipo
DIAGNOSTICOS_CON_SUBTIPO = {
    4:  6,    # Violencia   -> 6.- TIPO DE VIOLENCIA
    11: 12,   # Suicidio    -> 12.- TIPO DE SUICIDIO
    18: 20,   # Depresión   -> 20.- TIPO DE DEPRESIÓN
    41: 43,   # Ansiedad    -> 43.- TIPO DE TRASTORNO DE ANSIEDAD
    44: 45,   # Alzheimer   -> 45.- ETAPA
}

# Remapeo de bloques deprecated -> (patología, subtipo) forzados.
REMAP_DIAGNOSTICO = {
    9: ("Violencia", "Sexual"),   # Abuso Sexual (deprecated) -> Violencia > Víctima > Sexual
}

# Nombre de patología: False = header crudo (feo pero funciona).
# True (día de ocio) = usa limpiar_patologia() + OVERRIDE_PATOLOGIA.
LIMPIAR_NOMBRE_PATOLOGIA = False
OVERRIDE_PATOLOGIA = {
    # 65: "Otros trastornos del comportamiento",
}

# ╔═══════════════════════════════════════════════════════════════════╗
# ║  CONFIG TÉCNICA (rara vez se toca)                                  ║
# ╚═══════════════════════════════════════════════════════════════════╝
HOJA = None
ANCLA_ENCABEZADO = ["AÑO", "APLICACION", "FORMULARIO"]
USAR_BLANCO_EN_A = True
N_HARDCODE = 16
ANIO_TOKENS = ["AÑO", "APLICACION", "FORMULARIO"]
ANIO_COL_FALLBACK = 11
SEP_MULTI = " | "
MAX_FILAS_BUSQUEDA_HEADER = 60

SUBTIPO_NUMS = set(DIAGNOSTICOS_CON_SUBTIPO.values())
# ───────────────────────────────────────────────────────────────────


def norm(v):
    if v is None: return ""
    s = str(v).upper().strip()
    for a, b in zip("ÁÉÍÓÚÜÑ", "AEIOUUN"): s = s.replace(a, b)
    return re.sub(r"\s+", " ", s)


def to_year(v):
    if v is None: return None
    d = re.sub(r"\D", "", str(v))
    return int(d[:4]) if d else None


def num_pregunta(header):
    m = re.match(r"\s*(\d+)\s*\.\-", str(header))
    return int(m.group(1)) if m else None


def es_estado(h):
    return norm(h).endswith("ESTADO")


# ── LIMPIEZA DE NOMBRE DE PATOLOGÍA (pendiente, día de ocio) ──────────
def limpiar_patologia(header):
    """'18.- ¿ TIENE  DEPRESIÓN ?' -> 'Depresión'. Solo si LIMPIAR_NOMBRE_PATOLOGIA."""
    n = num_pregunta(header)
    if n in OVERRIDE_PATOLOGIA:
        return OVERRIDE_PATOLOGIA[n]
    s = re.sub(r"^\s*\d+\s*\.\-\s*", "", str(header))
    s = s.replace("¿", "").replace("?", "")
    s = re.sub(r"^\s*(TIENE|ES|SUFRE DE|PRESENTA|PADECE)\s+", "", s, flags=re.I)
    s = re.sub(r"^\s*PACIENTE\s+PRESENTA\s+", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:1].upper() + s[1:].lower() if s else s
# ─────────────────────────────────────────────────────────────────────


def limpiar_subtipo(valor, subtipo_header):
    """Recorta el sustantivo del header: valor 'Depresión Moderada' con header
    '20.- TIPO DE DEPRESIÓN' -> 'Moderada'. Si no hay 'TIPO DE', deja el valor."""
    if valor in (None, ""): return ""
    s = str(valor).strip()
    m = re.search(r"TIPO DE\s+(.+)$", norm(subtipo_header))
    if m:
        noun = m.group(1).strip()
        if norm(s).startswith(noun):
            rec = s[len(noun):].strip(" -,:")
            if rec: s = rec
    return s


def encontrar_fila_encabezado(ws):
    tope = min(ws.max_row, MAX_FILAS_BUSQUEDA_HEADER)
    ancla = [norm(t) for t in ANCLA_ENCABEZADO]
    for r in range(1, tope + 1):
        vals = [norm(c.value) for c in ws[r]]
        if all(any(tok in v for v in vals) for tok in ancla):
            return r, "ancla"
    if USAR_BLANCO_EN_A:
        for r in range(1, tope + 1):
            if norm(ws.cell(row=r, column=1).value) == "":
                return r + 1, "blanco_en_A"
    return N_HARDCODE + 1, "hardcode"


def encontrar_diagnostico(headers, c0):
    """Camina a la izquierda del ESTADO saltando subtipos/estados hasta dar
    con la pregunta del diagnóstico. Devuelve índice 0-based o c0 si no hay."""
    d = c0 - 1
    while d >= 0:
        h = headers[d]
        n = num_pregunta(h)
        if es_estado(h) or (n in SUBTIPO_NUMS) or (n is None):
            d -= 1
            continue
        return d
    return c0


def procesar(entrada: Path, salida: Path):
    wb = openpyxl.load_workbook(entrada)
    ws = wb[HOJA] if HOJA else wb.active

    header_idx, modo = encontrar_fila_encabezado(ws)
    if header_idx > 1:
        ws.delete_rows(1, header_idx - 1)
    ws.freeze_panes = None
    print(f"[corte] modo={modo} | encabezado fila original {header_idx} | borradas {header_idx-1} filas")

    headers = [c.value for c in ws[1]]
    ncols = ws.max_column
    num2col = {num_pregunta(h): i for i, h in enumerate(headers) if num_pregunta(h) is not None}

    anio_tok = [norm(t) for t in ANIO_TOKENS]
    anio_col = next((i for i, h in enumerate([norm(x) for x in headers], 1)
                     if all(t in h for t in anio_tok)), None)
    if anio_col is None or anio_col > ncols:
        anio_col = ANIO_COL_FALLBACK if ANIO_COL_FALLBACK <= ncols else None
    if anio_col:
        print(f"[año] columna {anio_col} -> '{headers[anio_col-1]}'")

    nuevas = []
    for k in BUSQUEDAS:
        nuevas += [f"Egreso_{k}_Patologia", f"Egreso_{k}_Subtipo"]
    for j, nombre in enumerate(nuevas):
        ws.cell(row=1, column=ncols + 1 + j, value=nombre)

    busq = {k: [norm(t) for t in v] for k, v in BUSQUEDAS.items()}
    n_match = {k: 0 for k in BUSQUEDAS}

    for r in range(2, ws.max_row + 1):
        if anio_col:
            ca = ws.cell(row=r, column=anio_col)
            y = to_year(ca.value)
            if y is not None:
                ca.value = y; ca.number_format = "0"

        fila = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        fila_n = [norm(v) for v in fila]

        for j, (k, toks) in enumerate(busq.items()):
            pats, subs = [], []
            for c0 in range(ncols):
                if not es_estado(headers[c0]):
                    continue
                if not all(t in fila_n[c0] for t in toks):
                    continue
                d = encontrar_diagnostico(headers, c0)
                diag_num = num_pregunta(headers[d])

                if diag_num in REMAP_DIAGNOSTICO:
                    pat, sub = REMAP_DIAGNOSTICO[diag_num]
                else:
                    pat = (limpiar_patologia(headers[d]) if LIMPIAR_NOMBRE_PATOLOGIA
                           else str(headers[d]))
                    sub_num = DIAGNOSTICOS_CON_SUBTIPO.get(diag_num)
                    sub_col0 = num2col.get(sub_num) if sub_num else None
                    sub_val = fila[sub_col0] if sub_col0 is not None else ""
                    sub = limpiar_subtipo(sub_val, headers[sub_col0] if sub_col0 is not None else "")
                pats.append(pat)
                subs.append(sub)

            if pats:
                ws.cell(row=r, column=ncols + 1 + 2*j,     value=SEP_MULTI.join(pats))
                ws.cell(row=r, column=ncols + 1 + 2*j + 1, value=SEP_MULTI.join(subs))
                n_match[k] += 1

    wb.save(salida)
    print(f"[ok] guardado: {salida}")
    print(f"[resumen] filas de datos: {ws.max_row - 1}")
    for k, v in n_match.items():
        print(f"          Egreso {k}: {v} filas")


def pedir_ruta():
    print(__doc__)
    return input("Pega la ruta del .xlsx (o Enter para salir): ").strip().strip('"').strip("'")


def main():
    if len(sys.argv) >= 2:
        entrada = Path(sys.argv[1].strip().strip('"').strip("'"))
        salida = (Path(sys.argv[2]) if len(sys.argv) > 2
                  else entrada.with_name(entrada.stem + "_procesado.xlsx"))
    else:
        ruta = pedir_ruta()
        if not ruta: return
        entrada = Path(ruta)
        salida = entrada.with_name(entrada.stem + "_procesado.xlsx")
    if not entrada.exists():
        print(f"\nERROR: no encuentro el archivo:\n  {entrada}"); return
    procesar(entrada, salida)


if __name__ == "__main__":
    sin_terminal = len(sys.argv) < 2
    try:
        main()
    except Exception as e:
        import traceback; traceback.print_exc(); print(f"\nERROR: {e}")
        sin_terminal = True
    finally:
        if sin_terminal:
            try: input("\nPresiona Enter para salir...")
            except EOFError: pass
