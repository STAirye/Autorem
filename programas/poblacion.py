#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 5 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# Version: 1.8.2
#
# This program is free software: you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by the
# Free Software Foundation, either version 3 of the License, or (at your
# option) any later version. Distributed WITHOUT ANY WARRANTY. See the GNU
# General Public License for more details: <https://www.gnu.org/licenses/>.
# ==========================================================================
"""
poblacion.py — port del DAX de la tabla «Ferrada» del PowerBI (Fase 1, SP·P6 A.1).

Construye, desde 3 exports crudos, la tabla intermedia POR-PACIENTE que hoy se
obtiene abriendo el PowerBI y exportando a mano el visual «Población PSM». Es
INFRAESTRUCTURA TRANSVERSAL (la tabla Ferrada la usan 8 páginas del PowerBI, no
solo Salud Mental) — vive en `programas/`, no en `modulos/`. Ver
`docs/SP_P6_poblacion_plan.md` (plan) y `docs/SP_P6_config_por_dx.md` (config,
tablas A/B/C, decisiones D1-D5).

Alcance SLIM (§3.2 del plan): de las ~150 columnas de Ferrada se portan ~55 —
solo las columnas `(form)` (28 diagnósticos/factores de riesgo), la actividad
(¿Ingresado?/¿Activo 12m?/rescate 6m-13m/¿Embarazada?), demografía básica y los
campos de traza del Inscritos (Estado/Situación/Motivo y Fecha Pasivación/Sector).
Fuera: fármacos, columnas `(fecha)`/`(mixto)`/`(dg)`, y todo dato de CONTACTO
(nombre, dirección, teléfono, mail — §8 CLAUDE.md, no solo economía de columnas).

Los NOMBRES de columna de salida son los del export PowerBI (para poder diffear
celda a celda contra un mes ya conocido). El corte reemplaza a TODAY() (§4.1):
todo se calcula hacia atrás desde el último día del MES REPORTADO, nunca desde
"hoy", para poder recalcular un mes pasado.

Divergencias esperadas contra el PowerBI (documentadas, NO son bugs del port):
  - D1 (fila 56/TGD no especificado): usa la pregunta 91 con fallback a la 63.
  - D2 (factores de riesgo): Violencia y Suicidio NO filtran por instrumento
    médico (el DAX sí filtraba — eso era el bug; el correcto es Abuso Sexual,
    que nunca filtró).
  - D5: si el diagnóstico está Activo pero su subtipo viene vacío, no tributa.
  - §4.3: egreso por DIAGNÓSTICO (el DAX original egresaba TODOS los dx del RUN
    con cualquier egreso del mes). Columna de auditoría + aviso en el log.
  - §4.7: gestante = control prenatal/formulario con matrona en ventana de 3
    meses (vs. los 2 meses off-by-one del DAX).
"""

from pathlib import Path

import numpy as np
import pandas as pd

from programas.rem_utils import (
    norm, fecha_col, cargar_atenciones, cargar_canonico,
    resolver_columnas, contiene_alguno, gestante_runs, PUEBLO_VACIO,
    OPENPYXL_OK, OPENPYXL_ERR, openpyxl, ArchivoInvalido, verificar_hoja_unica,
    buscar_col, num_pregunta, encontrar_fila_encabezado, _rango_mes,
)
from programas import formatos
from programas.rem_saludmental import DIAGNOSTICOS_CON_SUBTIPO

# ══════════════════════════════════════════════════════════════════════
# Tabla A/B — config por diagnóstico (docs/SP_P6_config_por_dx.md, APROBADA)
# ══════════════════════════════════════════════════════════════════════
# El DAX es UNIFORME: TODAS las 28 fórmulas (form) siguen el mismo patrón
# (§4.2 del plan) -> UNA función parametrizada (_estado_dx) + esta tabla de
# datos, no 28 casos especiales. `subtipo` se REUSA de rem_saludmental
# (fuente única con el A05) en vez de repetir los números acá.

def _spec(col, dx, estado, *, instrumento=True, subtipo_col=None,
          subtipo2=None, subtipo2_col=None):
    return dict(col=col, dx=dx, estado=estado, instrumento=instrumento,
                subtipo=DIAGNOSTICOS_CON_SUBTIPO.get(dx), subtipo_col=subtipo_col,
                subtipo2=subtipo2, subtipo2_col=subtipo2_col)


# Diagnósticos (filas 25-58 del P6): TODOS filtran INSTRUMENTO ⊃ MEDIC (D3).
TABLA_DX = [
    _spec("Depresión (form)", 18, 19, subtipo_col="Depresión gravedad (form)"),
    _spec("Depresión Postparto (form)", 21, 22),
    _spec("Bipolaridad (form)", 23, 24),
    _spec("OH Perjudicial (form)", 31, 32),
    _spec("OH Dependiente (form)", 33, 34),
    _spec("Drogas Perjudicial (form)", 35, 36),
    _spec("Drogas Dependiente (form)", 37, 38),
    _spec("OH y Drogas (form)", 39, 40),
    _spec("TDAH (form)", 57, 58),
    _spec("Oposicionista desafiante (form)", 69, 70),
    _spec("Ansiedad separación (form)", 71, 72),
    _spec("Otras Infancia/Adolescencia (form)", 73, 74),
    _spec("Ansiedad (form)", 41, 42, subtipo_col="Ansiedad (tipo)"),
    _spec("Demencia (form)", 44, 46, subtipo_col="Demencia gravedad (form)"),
    _spec("Esquizofrenia (form)", 51, 52),
    _spec("Adaptativo (form)", 49, 50),
    _spec("Conducta Alimentaria (form)", 55, 56),
    _spec("Retraso Mental (form)", 59, 60),
    _spec("Personalidad (form)", 61, 62),
    _spec("Autismo (form)", 83, 84),
    _spec("Asperger (form)", 85, 86),
    _spec("Rett (form)", 87, 88),
    _spec("Desintegrativo niñez (form)", 89, 90),
    _spec("TGD (form)", 91, 92),   # D1: fallback a la 63/64, ver _aplicar_fallback_tgd
    _spec("Otros (form)", 65, 66),
]

# Factores de riesgo (filas 15-23): NINGUNO filtra por instrumento (D2) —
# ⚠ contraintuitivo: un factor de riesgo lo puede registrar cualquier
# estamento, no solo médico. Abuso Sexual ya estaba bien en el DAX (única
# excepción correcta); Violencia y Suicidio SÍ filtraban y eso era el bug.
TABLA_FR = [
    _spec("Violencia (form)", 4, 5, instrumento=False,
          subtipo_col="Violencia Tipo (form)",
          subtipo2=7, subtipo2_col="Violencia Victima o Agresor (form)"),
    _spec("Abuso Sexual (form)", 9, 10, instrumento=False),
    _spec("Suicidio (form)", 11, 13, instrumento=False, subtipo_col="Suicidio Tipo (form)"),
]

TODAS_LAS_SPECS = TABLA_FR + TABLA_DX

# D1 — TGD no especificado: la 91 no tiene columna en el PowerBI (nunca se
# llenó); se recupera vía la pregunta "padre" 63, pero SOLO si ninguna de las
# TGD específicas (83/85/87/89/91) está activa (evita doble conteo).
_TGD_ESPECIFICAS = ["Autismo (form)", "Asperger (form)", "Rett (form)",
                    "Desintegrativo niñez (form)"]
_Q_TGD_FALLBACK, _E_TGD_FALLBACK = 63, 64

# Preguntas que se leen del formulario histórico (unión de todo lo de arriba
# + la 1, "madre de hijo <5", que NO tiene subtipo/estado propio).
QUESTIONS = sorted({1, _Q_TGD_FALLBACK, _E_TGD_FALLBACK} | {
    n for spec in TODAS_LAS_SPECS
    for n in (spec["dx"], spec["estado"], spec["subtipo"], spec["subtipo2"]) if n
})
ESTADOS_TODOS = sorted({spec["estado"] for spec in TODAS_LAS_SPECS} | {_E_TGD_FALLBACK})

# Las 7 actividades SM validadas (rem_sm_actividades.ADA_TRIBUTAN habla de
# CASILLAS del REM; esta lista es la del DAX "SM Activo 12m", ver §8.1 del
# plan: rescate 6m/13m usan la MISMA lista, no el 'contains "salud mental"'
# laxo del DAX original). Fuente única para Activo12m / rescate 6m / rescate 13m.
ACTIVIDADES_SM_7 = [
    "control salud mental",
    "controles salud mental",
    "consulta de salud mental",
    "visita domiciliaria integral familia con integrante con patologia de salud mental",
    "visita domiciliaria integral a familia con adulto mayor con demencia",
    "visita domiciliaria integral a familia con niños/as de 5 a 9 años con problemas y/o trastorno",
    "visita integral de salud mental a domicilio",
]

# ── Columnas de salida (nombres = export PowerBI; §3.2 slim) ──
COL_IDENTIDAD = ["Número", "Tipo de identificación", "Sexo", "Género", "Edad", "Sector",
                 "Situación", "Estado", "Motivo Pasivación", "Fecha Pasivación",
                 "¿Originario o Migrante?", "Pueblo Originario", "PROTECCION NIÑEZ"]
COL_ACTIVIDAD = ["¿Ingresado?", "¿Activo 12m?", "¿Última atención hace 6m?",
                 "¿Última atención hace 13m?", "¿Embarazada?", "Madre <5 años"]


def _col_dx():
    cols = [s["col"] for s in TODAS_LAS_SPECS]
    for s in TODAS_LAS_SPECS:
        if s["subtipo_col"]:
            cols.append(s["subtipo_col"])
        if s["subtipo2_col"]:
            cols.append(s["subtipo2_col"])
    return cols


COLUMNAS_SALIDA = COL_IDENTIDAD + COL_ACTIVIDAD + _col_dx()


# ══════════════════════════════════════════════════════════════════════
# Input 1 — Informe Inscritos y Adscritos (snapshot, IRIS)
# ══════════════════════════════════════════════════════════════════════
MAPA_INSCRITOS = {
    "RUN":          ("exact", "NUMERO TIPO IDENTIFICACION"),
    "TIPOID":       ("exact", "TIPO IDENTIFICACION"),
    "SEXO":         ("exact", "SEXO"),
    "GENERO":       ("exact", "GENERO"),
    "FNAC":         ("exact", "FECHA DE NACIMIENTO"),
    "EDADANOS":     ("exact", "EDAD AÑOS"),
    "SITUACION":    ("exact", "SITUACION"),
    "ESTADO":       ("exact", "ESTADO"),
    "FPASIV":       ("exact", "FECHA PASIVACION"),
    "MPASIV":       ("exact", "MOTIVO PASIVACION"),
    "SECTOR":       ("exact", "SECTOR"),
    "ALERTAS":      ("subs", ["ALERTAS", "ADMINISTRATIVAS"]),
    "PUEBLO":       ("exact", "PUEBLO INDIG"),
    "NACIONALIDAD": ("exact", "NACIONALIDAD"),
}


def cargar_inscritos(entrada, log=print):
    """'Informe Inscritos y Adscritos' (IRIS, snapshot) -> DataFrame, 1 fila por
    RUN. Es la base de la tabla Ferrada: TODA la población, sin filtrar."""
    d, _ = cargar_canonico(entrada, None, lambda h: resolver_columnas(h, MAPA_INSCRITOS),
                          requeridas=("RUN", "SEXO", "ESTADO", "SITUACION"))
    d["RUN"] = d["RUN"].astype(str).str.strip()
    d = d[d["RUN"] != ""].drop_duplicates(subset="RUN", keep="last").reset_index(drop=True)
    log(f"[poblacion] Inscritos: {len(d)} personas (snapshot)")
    return d


# ══════════════════════════════════════════════════════════════════════
# Input 2 — Formulario 'Control de Salud Mental' histórico (IRIS, multi-archivo)
# ══════════════════════════════════════════════════════════════════════
def _leer_formulario_1(entrada, log):
    """Un archivo del histórico -> DataFrame (RUN, FECHA cruda, INSTR, q<N> por
    cada pregunta de QUESTIONS). Reusa la detección de encabezado/identidad de
    `formatos`/`rem_saludmental` (perfil IRIS): mismo export que el A05."""
    if not OPENPYXL_OK:
        raise ImportError(f"Falta 'openpyxl' (pip install openpyxl). Detalle: {OPENPYXL_ERR}")
    nombre = Path(str(entrada)).name
    verificar_hoja_unica(entrada)
    # NO read_only: detectar_eje/encontrar_fila_encabezado usan ws[r] / ws.cell(),
    # que el modo read-only de openpyxl no soporta (mismo patrón que
    # rem_saludmental.abrir_validado). El costo de memoria es aceptable: es el
    # mismo tipo de archivo que ya procesa el A05, un mes/año a la vez.
    wb = openpyxl.load_workbook(entrada, data_only=True)
    ws = wb.active
    if formatos.detectar_eje(ws) != "iris":
        wb.close()
        raise ArchivoInvalido(
            "no_iris",
            f"Archivo «{nombre}»:\n\nEl histórico de 'Control de Salud Mental' debe "
            "venir en formato IRIS (no el Reporte Administrativo): trae columnas que "
            "el Administrativo no tiene y que este módulo necesita.")
    header_idx, _modo = encontrar_fila_encabezado(
        ws, formatos.ANCLA_IRIS, True, 16, formatos.MAX_FILAS_HEADER)
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = list(filas[header_idx - 1])
    headers_n = [norm(h) for h in headers]
    rut_col, _edad_col, _sexo_col = formatos.resolver_identidad(headers_n)
    fecha_col_i = buscar_col(headers_n, tokens=["FECHA", "FORMULARIO"])
    instr_col_i = buscar_col(headers_n, exacto="INSTRUMENTO")
    if not rut_col or not fecha_col_i:
        raise ArchivoInvalido(
            "sin_columnas",
            f"Archivo «{nombre}»:\n\nNo encuentro RUT y/o FECHA FORMULARIO. "
            "¿Es el export IRIS de 'Control de Salud Mental', sin modificar?")
    num2col = {num_pregunta(h): i for i, h in enumerate(headers) if num_pregunta(h) is not None}

    out = {"RUN": [], "FECHA": [], "INSTR": []}
    for n in QUESTIONS:
        out[f"q{n}"] = []
    for row in filas[header_idx:]:
        rut = row[rut_col - 1] if rut_col - 1 < len(row) else None
        if not str(rut or "").strip():
            continue
        out["RUN"].append(str(rut).strip())
        out["FECHA"].append(row[fecha_col_i - 1] if fecha_col_i - 1 < len(row) else None)
        out["INSTR"].append(row[instr_col_i - 1] if instr_col_i and instr_col_i - 1 < len(row) else "")
        for n in QUESTIONS:
            c = num2col.get(n)
            out[f"q{n}"].append(row[c] if c is not None and c < len(row) else None)
    return pd.DataFrame(out)


def cargar_formulario_sm(entrada, log=print):
    """Histórico COMPLETO del formulario 'Control de Salud Mental' (IRIS) -> uno
    o varios archivos (acepta LISTA, como el ADA del A23). Devuelve DataFrame
    largo: 1 fila por formulario aplicado, con q<N>/q<N>_n por cada pregunta que
    consume la config de arriba."""
    archivos = entrada if isinstance(entrada, (list, tuple)) else [entrada]
    partes = []
    for e in archivos:
        try:
            partes.append(_leer_formulario_1(e, log))
        except ArchivoInvalido:
            raise
        except Exception as ex:
            raise ArchivoInvalido(
                "no_legible",
                f"No pude leer «{Path(str(e)).name}»:\n\n{ex}\n\n"
                "¿Es un .xlsx válido del export IRIS, sin modificar?") from ex
    d = pd.concat(partes, ignore_index=True) if len(partes) > 1 else partes[0]
    d["FECHA"] = fecha_col(d["FECHA"], log, "FECHA FORMULARIO (histórico SM)")
    d["INSTR_n"] = d["INSTR"].map(norm)
    for n in QUESTIONS:
        d[f"q{n}_n"] = d[f"q{n}"].map(norm)
    span = (f"{d['FECHA'].min():%Y-%m}..{d['FECHA'].max():%Y-%m}"
            if d["FECHA"].notna().any() else "sin fechas")
    log(f"[poblacion] Formulario SM histórico: {len(d)} formularios, "
        f"{d['RUN'].nunique()} RUN distintos ({span})")
    return d


# ══════════════════════════════════════════════════════════════════════
# Motor: último formulario por RUN×diagnóstico (parametrizado, D2/D3/D5/§4.3)
# ══════════════════════════════════════════════════════════════════════
def _mes_offset(corte, n_meses):
    """(inicio, fin) del mes calendario que está `n_meses` ANTES del mes de
    `corte` (Timestamp = último día del mes reportado)."""
    import calendar
    y, m = corte.year, corte.month - n_meses
    while m <= 0:
        m += 12
        y -= 1
    return pd.Timestamp(y, m, 1), pd.Timestamp(y, m, calendar.monthrange(y, m)[1])


def _estado_dx(df, dx, estado, corte, mes_ini, mes_fin, *, instrumento=True,
              subtipo=None, subtipo2=None):
    """Motor ÚNICO (parametrizado por dx/estado/instrumento) que reemplaza las
    28 fórmulas (form) del DAX (§4.2 del plan): busca, por RUN, el último
    formulario con `dx`⊃SI & `estado`⊃ingreso|seguimiento (filtrando INSTRUMENTO
    ⊃ MEDIC solo si `instrumento`=True — D2/D3) hasta `corte`; separa el egreso
    POR DIAGNÓSTICO (§4.3, no el bug del PowerBI): Egresado si ESE mismo bloque
    ESTADO trae 'egreso' dentro del mes reportado [mes_ini, mes_fin].

    Devuelve DataFrame indexado por RUN: 'activo_base' (bool, antes del egreso —
    para la auditoría §4.3), 'estado' ('Activo'/'Egresado'/''), y 'subtipo'/
    'subtipo2' (valor CRUDO del último formulario activo; D5: no se filtra acá
    si viene vacío — eso lo decide quien consume la tabla)."""
    qd, qe = f"q{dx}_n", f"q{estado}_n"
    cond_si = df[qd] == "SI"
    cond_ing = cond_si & (df[qe].str.contains("INGRES", na=False) |
                          df[qe].str.contains("SEGUIMIEN", na=False))
    cond_egr = cond_si & df[qe].str.contains("EGRES", na=False)
    if instrumento:
        instr_ok = df["INSTR_n"].str.contains("MEDIC", na=False)
        cond_ing &= instr_ok
        cond_egr &= instr_ok

    activos = (df[cond_ing & (df["FECHA"] <= corte)]
              .sort_values("FECHA").groupby("RUN").last())
    egresados_mes = set(df.loc[cond_egr & df["FECHA"].between(mes_ini, mes_fin), "RUN"])

    out = pd.DataFrame(index=sorted(set(activos.index) | egresados_mes))
    out["activo_base"] = out.index.isin(activos.index)
    out["estado"] = np.where(out.index.isin(egresados_mes), "Egresado",
                             np.where(out["activo_base"], "Activo", ""))
    if subtipo:
        col = f"q{subtipo}"
        out["subtipo"] = activos[col].reindex(out.index) if col in activos.columns else ""
    if subtipo2:
        col2 = f"q{subtipo2}"
        out["subtipo2"] = activos[col2].reindex(out.index) if col2 in activos.columns else ""
    return out


def _aplicar_fallback_tgd(P, form, corte, mes_ini, mes_fin, log):
    """D1: la fila 56 (TGD no especificado) usa la pregunta 91 — que nunca tiene
    dato en el PowerBI (0 referencias en el spec) — con fallback a la pregunta
    "padre" 63, SOLO para quien no tenga ninguna TGD específica activa."""
    est63 = _estado_dx(form, _Q_TGD_FALLBACK, _E_TGD_FALLBACK, corte, mes_ini, mes_fin)
    ya_tiene_tgd = P[_TGD_ESPECIFICAS].apply(lambda c: c.isin(["Activo", "Egresado"])).any(axis=1)
    fallback_estado = P["Número"].map(est63["estado"]).fillna("")
    usa_fallback = (P["TGD (form)"] == "") & ~ya_tiene_tgd & (fallback_estado != "")
    n = int(usa_fallback.sum())
    if n:
        log(f"[poblacion] D1: {n} persona(s) recuperada(s) para «TGD (form)» vía la "
            f"pregunta 63 (fallback; la 91 no tiene columna en el PowerBI actual).")
    P.loc[usa_fallback, "TGD (form)"] = fallback_estado[usa_fallback]
    return P


def _egreso_powerbi_bug(form, mes_ini, mes_fin):
    """RUN con CUALQUIER columna ESTADO (de las 28) marcando 'egreso' en el mes
    reportado — replica el bug histórico del PowerBI (§4.3): un egreso de UN
    diagnóstico egresaba TODOS los dx activos del RUN, sin mirar el bloque."""
    m = pd.Series(False, index=form.index)
    for n in ESTADOS_TODOS:
        col = f"q{n}_n"
        if col in form.columns:
            m |= form[col].str.contains("EGRES", na=False)
    m &= form["FECHA"].between(mes_ini, mes_fin)
    return set(form.loc[m, "RUN"])


# ══════════════════════════════════════════════════════════════════════
# Actividad SM (Ingresado / Activo 12m / rescate 6m-13m / Gestante)
# ══════════════════════════════════════════════════════════════════════
def _runs_actividad_sm(d_ada, ini, fin):
    w = d_ada[(d_ada["FECHA"] >= ini) & (d_ada["FECHA"] <= fin) &
             contiene_alguno(d_ada["ACT_n"], ACTIVIDADES_SM_7)]
    return set(w["RUN"])


def _flags_actividad(d_ada, corte):
    """(activos_12m, rescate_6m, rescate_13m): sets de RUN. Ver §8.1 del plan —
    las 3 usan la MISMA lista de actividades validada (ACTIVIDADES_SM_7)."""
    ini12, _ = _mes_offset(corte, 11)
    activos12 = _runs_actividad_sm(d_ada, ini12, corte)

    ini6, fin6 = _mes_offset(corte, 6)
    activos6 = _runs_actividad_sm(d_ada, ini6, fin6)
    posteriores6 = set(d_ada.loc[(d_ada["FECHA"] > fin6) & (d_ada["FECHA"] <= corte) &
                                 contiene_alguno(d_ada["ACT_n"], ACTIVIDADES_SM_7), "RUN"])
    rescate6 = activos6 - posteriores6

    ini13, fin13 = _mes_offset(corte, 13)
    activos13mes = _runs_actividad_sm(d_ada, ini13, fin13)
    rescate13 = activos13mes - activos12

    return activos12, rescate6, rescate13


def _ultima_respuesta(form, pregunta, corte):
    """Última respuesta NO vacía a `pregunta`, por RUN, hasta `corte` (para
    'Madre <5 años': se reporta con el dato más reciente que exista, sin
    exigir ingreso/seguimiento — así lo trae el DAX)."""
    q = f"q{pregunta}_n"
    w = form[(form[q] != "") & (form["FECHA"] <= corte)].sort_values("FECHA").groupby("RUN").last()
    return w[q] if q in w.columns else pd.Series(dtype=object)


# ══════════════════════════════════════════════════════════════════════
# Ensamblado: construir_poblacion()
# ══════════════════════════════════════════════════════════════════════
def construir_poblacion(inscritos, formulario_sm, ada, mes=None, log=print):
    """Arma la tabla «Ferrada» SLIM (§3.2): TODOS los inscritos (snapshot), 1
    fila por RUN, sin filtrar por Activo/Ingresado (eso lo aplica quien consuma
    la tabla — el P6 filtra, el rescate filtra distinto). `mes`=(año,mes) fija
    el corte (None = mes anterior, como el resto de los módulos pandas).
    `ada` puede ser un DataFrame ya cargado (para no releer el archivo si el
    caller ya lo tiene) o ruta(s) del export de Atenciones."""
    ini, fin = _rango_mes(mes)
    corte = fin
    mes_ini, mes_fin = ini, fin

    insc = cargar_inscritos(inscritos, log=log)
    form = cargar_formulario_sm(formulario_sm, log=log)
    d_ada = ada if isinstance(ada, pd.DataFrame) else cargar_atenciones(ada, log=log)

    P = insc.rename(columns={"RUN": "Número", "TIPOID": "Tipo de identificación",
                             "SITUACION": "Situación", "ESTADO": "Estado",
                             "FPASIV": "Fecha Pasivación", "MPASIV": "Motivo Pasivación",
                             "SECTOR": "Sector"}).copy()

    # ── Sexo/Género (Sexo vacío -> 'No informado', igual que el DAX) ──
    P["Sexo"] = P["SEXO"].where(P["SEXO"].notna() & (P["SEXO"].astype(str).str.strip() != ""),
                                "No informado")
    P["Género"] = P["GENERO"]

    # ── Edad al CORTE (no a hoy): fecha nacimiento; fallback EDAD AÑOS del snapshot ──
    fnac = fecha_col(P["FNAC"], log, "Fecha Nacimiento (Inscritos)")
    edad_calc = ((corte - fnac).dt.days / 365.25).apply(lambda x: int(x) if pd.notna(x) else np.nan)
    edad_fallback = pd.to_numeric(P["EDADANOS"], errors="coerce")
    P["Edad"] = edad_calc.where(edad_calc.notna(), edad_fallback)

    # ── Pueblo Originario / Migrante (norm-based; más robusto que el literal del DAX) ──
    P["Pueblo Originario"] = P["PUEBLO"]
    nac_n = P["NACIONALIDAD"].map(norm)
    pueblo_n = P["PUEBLO"].map(norm)
    es_migrante = ~nac_n.isin({"", "CHILENA", "DESCONOCIDO", "DESCONOCIDA"})
    es_originario = (nac_n == "CHILENA") & (~pueblo_n.isin(PUEBLO_VACIO))
    P["¿Originario o Migrante?"] = np.select([es_migrante, es_originario],
                                             ["Migrante", "Originario"], default="NO")

    # ── PROTECCION NIÑEZ (SENAME / Mejor Niñez, desde ALERTAS ADMINISTRATIVAS) ──
    alertas_n = P["ALERTAS"].map(norm)
    es_sename = alertas_n.str.contains("SENAME", na=False)
    es_mejorninez = alertas_n.str.contains("SPE", na=False) | alertas_n.str.contains("MEJOR NINEZ", na=False)
    P["PROTECCION NIÑEZ"] = np.select([es_sename, es_mejorninez], ["SENAME", "Mejor Niñez"], default="NO")

    # ── Actividad SM (§8.1: misma lista de 7 para las 3 definiciones) ──
    activos12, rescate6, rescate13 = _flags_actividad(d_ada, corte)
    P["¿Activo 12m?"] = np.where(P["Número"].isin(activos12), "SI", "NO")
    P["¿Última atención hace 6m?"] = np.where(P["Número"].isin(rescate6), "Si", "No")
    P["¿Última atención hace 13m?"] = np.where(P["Número"].isin(rescate13), "Si", "No")

    # ── Gestante (§4.7: 3 meses cerrados, matrona; NO el DAX de 2 meses) ──
    ini3 = ini - pd.DateOffset(months=2)
    if d_ada["FECHA"].notna().any() and d_ada["FECHA"].min() > ini3:
        log(f"[poblacion] ⚠ GESTANTES usa ventana de 3 meses (desde {ini3:%Y-%m}), pero "
            f"el ADA arranca en {d_ada['FECHA'].min():%Y-%m} -> puede SUBCONTAR.")
    gset = gestante_runs(d_ada, ini3, corte)
    P["¿Embarazada?"] = np.where(P["Número"].isin(gset), "SI", "NO")

    # ── Madre <5 años (pregunta 1, última respuesta con dato; SIN filtro de
    #     sexo acá -- eso es una regla del P6, no de Ferrada, §5.4.1 del plan) ──
    m5 = _ultima_respuesta(form, 1, corte)
    P["Madre <5 años"] = P["Número"].map(m5).fillna("")

    # ── Los 28 diagnósticos/factores de riesgo (motor único, ver _estado_dx) ──
    divergencias_detalle = []   # 1 fila por (Diagnóstico, RUN) que diverge — auditoría §4.3
    egreso_bug_runs = _egreso_powerbi_bug(form, mes_ini, mes_fin)
    for spec in TODAS_LAS_SPECS:
        est = _estado_dx(form, spec["dx"], spec["estado"], corte, mes_ini, mes_fin,
                         instrumento=spec["instrumento"], subtipo=spec["subtipo"],
                         subtipo2=spec["subtipo2"])
        P[spec["col"]] = P["Número"].map(est["estado"]).fillna("")
        if spec["subtipo_col"]:
            P[spec["subtipo_col"]] = P["Número"].map(est.get("subtipo", pd.Series(dtype=object))).fillna("")
        if spec["subtipo2_col"]:
            P[spec["subtipo2_col"]] = P["Número"].map(est.get("subtipo2", pd.Series(dtype=object))).fillna("")

        # Auditoría §4.3: valor que habría dado el PowerBI (bug cross-dx de egreso).
        # Se guarda el RUN (no solo el conteo) para que Egreso_Divergencias sea auditable.
        base = P["Número"].map(est["activo_base"]).fillna(False)
        pbi = np.where(base & P["Número"].isin(egreso_bug_runs), "Egresado",
                       np.where(base, "Activo", ""))
        diverge = ((P[spec["col"]] != pbi) & base).to_numpy()
        if diverge.any():
            for run, nuestro, pbi_v in zip(P.loc[diverge, "Número"], P.loc[diverge, spec["col"]],
                                           pbi[diverge]):
                divergencias_detalle.append({"Diagnostico": spec["col"], "RUN": run,
                                             "Valor_port": nuestro, "Valor_PowerBI": pbi_v})

    P = _aplicar_fallback_tgd(P, form, corte, mes_ini, mes_fin, log)

    # ── ¿Ingresado? (SI si CUALQUIER dx/FR quedó 'Activo') ──
    cols_dx = [s["col"] for s in TODAS_LAS_SPECS]
    P["¿Ingresado?"] = np.where((P[cols_dx] == "Activo").any(axis=1), "SI", "NO")

    div_df = pd.DataFrame(divergencias_detalle,
                          columns=["Diagnostico", "RUN", "Valor_port", "Valor_PowerBI"])
    if len(div_df):
        div_df = div_df.sort_values(["Diagnostico", "RUN"]).reset_index(drop=True)
        resumen = div_df.groupby("Diagnostico", sort=False).size()
        log(f"[poblacion] ⚠ §4.3 egreso por diagnóstico: {len(div_df)} persona(s)×diagnóstico "
            f"difieren del PowerBI (que egresaba TODOS los dx con cualquier egreso del RUN en "
            f"el mes). Detalle (con RUT) en la hoja 'Egreso_Divergencias'. Esperado, no es bug. "
            + " · ".join(f"{d}={n}" for d, n in resumen.items()))
    else:
        log("[poblacion] §4.3: sin divergencias de egreso este mes.")

    P.attrs["mes"] = (ini.year, ini.month)
    P.attrs["egreso_divergencias"] = div_df
    n_ingresados = int((P["¿Ingresado?"] == "SI").sum())
    log(f"[poblacion] Ferrada: {len(P)} personas en el snapshot | {n_ingresados} con "
        f"¿Ingresado?=SI (mes {ini:%Y-%m})")
    if not (1300 <= n_ingresados <= 1500):
        log(f"[poblacion] ⚠ {n_ingresados} ingresados está fuera del rango histórico "
            "~1300-1500 esperado para este centro (sanity check, no error automático).")

    # .attrs de pandas NO se propaga de forma confiable al recortar columnas
    # (es "experimental" en pandas) -> se reasigna explícito sobre la salida.
    P_out = P[COLUMNAS_SALIDA].copy()
    P_out.attrs["mes"] = P.attrs["mes"]
    P_out.attrs["egreso_divergencias"] = P.attrs["egreso_divergencias"]
    return P_out


def escribir_divergencias(wb, div, sheet_name="Egreso_Divergencias"):
    """Escribe `div` (columnas Diagnostico/RUN/Valor_port/Valor_PowerBI, YA ordenado
    por Diagnostico) como bloques COLAPSABLES por diagnóstico (agrupación de filas de
    Excel, +/- en el margen izquierdo, §4.3 del plan): 1 fila de cabecera con el
    diagnóstico y el total, y debajo el detalle por RUN. `wb` = Workbook openpyxl.
    Nada que escribir -> no crea la hoja."""
    if div is None or not len(div):
        return None
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.outlinePr.summaryBelow = False   # cabecera ARRIBA de su detalle -> el
                                                          # +/- de colapsar queda junto a ella
    ws.append(["Diagnóstico", "N personas", "RUN", "Valor_port", "Valor_PowerBI (con el bug §4.3)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    r = 2
    for diag, sub in div.groupby("Diagnostico", sort=False):
        ws.cell(row=r, column=1, value=diag).font = Font(bold=True)
        ws.cell(row=r, column=2, value=len(sub))
        r += 1
        for _, fila in sub.iterrows():
            ws.cell(row=r, column=3, value=fila["RUN"])
            ws.cell(row=r, column=4, value=fila["Valor_port"])
            ws.cell(row=r, column=5, value=fila["Valor_PowerBI"])
            ws.row_dimensions[r].outline_level = 1   # colapsable bajo la cabecera del diagnóstico
            r += 1
    ws.freeze_panes = "A2"
    for i, w in enumerate([38, 11, 14, 16, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def escribir(P, salida):
    """Escribe `PSM_Poblacion` (la tabla intermedia, snapshot mensual archivable)
    + `Egreso_Divergencias` (auditoría §4.3, con RUT y colapsable por diagnóstico —
    ver `escribir_divergencias`) en un solo .xlsx."""
    with pd.ExcelWriter(salida, engine="openpyxl") as xw:
        P.to_excel(xw, index=False, sheet_name="PSM_Poblacion")
        escribir_divergencias(xw.book, P.attrs.get("egreso_divergencias"))
    return str(salida)
