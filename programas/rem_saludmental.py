#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# Version: 1.7.13
#
# This program is free software: you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by the
# Free Software Foundation, either version 3 of the License, or (at your
# option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTABILITY
# or FITNESS FOR A PARTICULAR PURPOSE. See the GNU General Public License
# for more details: <https://www.gnu.org/licenses/>.
# ==========================================================================
"""
rem_saludmental.py — capa compartida del formulario 'Control de Salud Mental'.

Conoce la ESTRUCTURA del formulario (no una sección concreta del REM): dónde está
el encabezado, cómo se ubican RUT/edad/sexo, la caminata al diagnóstico, subtipos,
demografía y el motor de marcado de eventos por token de ESTADO.

El mismo formulario se puede descargar en DOS formatos de export distintos:
  - PERFIL_IRIS   ('Control de Salud Mental' de IRIS)   -> el que la herramienta
                  fue hecha para procesar; trae todas las columnas.
  - PERFIL_ADMIN  ('Reporte Administrativo' de RAYEN)   -> mismo formulario, otro
                  layout de encabezado y SIN varias columnas demográficas.

Un 'perfil' encapsula esas diferencias de formato; las tareas (egresos/ingresos)
son agnósticas al formato. Lo consumen rem_a05_o_egresos.py y rem_a05_n_ingresos.py.
La GUI/CLI (autorem.py) deja elegir el perfil al inicio.

╔══════════════════════════════════════════════════════════════════════╗
║  Aquí vive la CONFIGURACIÓN CLÍNICA del formulario.                    ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import re

from programas.rem_utils import (
    OPENPYXL_OK, OPENPYXL_ERR, openpyxl,
    ArchivoInvalido,
    norm, buscar_col, num_pregunta,
    encontrar_fila_encabezado, edad_anios, mes_de_celda,
)
from programas import formatos

# ── Diagnósticos con subtipo (clave = N.- diagnóstico, valor = N.- subtipo) ──
# La numeración de preguntas es IDÉNTICA en IRIS y en el Administrativo.
DIAGNOSTICOS_CON_SUBTIPO = {
    4:  6,    # Violencia   -> 6.- TIPO DE VIOLENCIA
    11: 12,   # Suicidio    -> 12.- TIPO DE SUICIDIO
    18: 20,   # Depresión   -> 20.- TIPO DE DEPRESIÓN
    41: 43,   # Ansiedad    -> 43.- TIPO DE TRASTORNO DE ANSIEDAD
    44: 45,   # Alzheimer   -> 45.- ETAPA
}
SUBTIPO_NUMS = set(DIAGNOSTICOS_CON_SUBTIPO.values())

# Remapeo de bloques deprecated -> (patología, subtipo) forzados.
REMAP_DIAGNOSTICO = {
    9: ("Violencia", "Sexual"),   # Abuso Sexual (deprecated) -> Violencia > Víctima > Sexual
}

# Diagnósticos que NO van al A05 Salud Mental (se saltan del output):
#   epilepsia -> REM adulto; programas de rehabilitación/acompañamiento -> no son
#   egresos/ingresos de diagnóstico SM.
EXCLUIR_PATOLOGIA = {75, 77, 79, 81}

# Nombre de patología: True = limpiar_patologia() + OVERRIDE (nombres canónicos,
# tomados de SP·P6 y revisados jul-2026). False = header crudo.
LIMPIAR_NOMBRE_PATOLOGIA = True
OVERRIDE_PATOLOGIA = {
    4:  "Violencia",
    11: "Suicidio",
    18: "Depresión",
    21: "Depresión post parto",
    23: "Trastorno bipolar",                 # RAYEN trae typo 'Transtorno'
    25: "Depresión refractaria",
    27: "Depresión grave con psicosis",
    29: "Depresión con alto riesgo suicida",
    31: "Consumo perjudicial de alcohol",
    33: "Consumo dependiente de alcohol",
    35: "Consumo perjudicial de drogas",
    37: "Consumo dependiente de drogas",
    39: "Consumo de drogas y alcohol",
    41: "Trastorno de ansiedad",
    44: "Demencia (incluye Alzheimer)",
    47: "Psicosis",
    49: "Trastorno adaptativo",
    51: "Esquizofrenia",
    53: "Primer episodio de esquizofrenia",
    55: "Trastorno de la conducta alimentaria",
    57: "TDAH (trastorno hipercinético)",
    59: "Retraso mental",
    61: "Trastorno de personalidad",
    63: "Trastorno generalizado del desarrollo",
    65: "Otros trastornos (no incluidos)",
    67: "Trastornos conductuales asociados a demencia",
    69: "Trastorno disocial desafiante y oposicionista",
    71: "Trastorno de ansiedad de separación en la infancia",
    73: "Otros trastornos del comportamiento y las emociones (infancia)",
    83: "Autismo",
    85: "Asperger",
    87: "Síndrome de Rett",
    89: "Trastorno desintegrativo de la infancia",
    91: "Trastorno generalizado del desarrollo no especificado",
    # 9 (Abuso Sexual) NO va aquí: lo maneja REMAP_DIAGNOSTICO -> Violencia/Sexual.
    # 75/77/79/81 excluidos (ver EXCLUIR_PATOLOGIA).
}

# Override de nombre de SUBTIPO, por Nº de pregunta del subtipo. Cada regla es
# (keywords, nombre): si el valor crudo normalizado contiene CUALQUIER keyword,
# sale ese nombre. Sirve para juntar variantes (el REM tiene 'Pánico' a secas ->
# los dos pánicos de RAYEN caen en uno) y acortar. Los subtipos SIN entrada acá
# (Depresión, Violencia, Suicidio, Alzheimer) salen por el recorte del header,
# que ya los deja limpios (Moderada / Física / Intento / Moderado).
OVERRIDE_SUBTIPO = {
    43: [  # 43.- TIPO DE TRASTORNO DE ANSIEDAD  (¡PANICO antes que FOBIA! 'AgoroFOBIA')
        (["PANICO"],                       "Pánico"),        # junta 'Pánico' y 'Pánico sin agorafobia'
        (["FOBIA"],                        "Fobia social"),
        (["GENERALIZ"],                    "Generalizada"),
        (["TEPT", "ESTRES", "TRAUMATICO"], "TEPT"),
        (["OTROS"],                        "Otros"),
    ],
}

# ── Detección de columnas de identificación (RUT/edad/sexo, ambos formatos) ──
# El eje IRIS/Admin y la resolución de identidad viven en `programas.formatos`
# (compartidos con A03/estamentos). El QUIRK de 'AÑO APLICACIÓN FORMULARIO' (que
# trae la EDAD, no el año) está documentado allá.

# Caracterización demográfica (devuelve "SI"/"" salvo Trans). Validado jul-2026
# contra valores reales de 'ALERTAS ADMINISTRATIVAS' (IRIS). En el Administrativo
# las fuentes ALERTAS/PUEBLO ORIGINARIO/GÉNERO NO existen -> esos flags salen vacíos.
DEMOGRAFIA = {
    "Madre_menor5":        (["USTED", "MADRE"],       ["SI"]),                                   # pregunta 1 (ambos formatos)
    "Pueblos_Originarios": (["PUEBLO", "ORIGINARIO"], "_no_vacio"),                              # solo IRIS
    "SENAME":              (["ALERTAS"],              ["SENAME", "REINSERCION"]),                # solo IRIS
    "Proteccion_Ninez":    (["ALERTAS"],              ["MEJOR NINEZ", "PROTECCION ESPECIAL"]),   # solo IRIS
    "Migrante":            (["ALERTAS"],              ["MIGRANTE"]),                             # solo IRIS
}
COL_GENERO_TOKENS = ["GENERO"]   # Trans: valor de GÉNERO si contiene "TRANS" (solo IRIS)
NEGATIVOS_DEMO = {"", "NO", "NINGUNO", "NINGUNA", "NO APLICA", "SIN INFORMACION", "N/A"}

# ── Config técnica compartida ──
HOJA = None
ANIO_COL_FALLBACK = 11                 # último recurso para la edad (layout IRIS)
MAX_FILAS_BUSQUEDA_HEADER = formatos.MAX_FILAS_HEADER

# ── Layout de salida (compartido por egresos/ingresos) ──
ANCHOS_BASE = [14, 8, 8, 13, 44, 26, 13]   # RUT, Edad, Sexo, Tipo, Patologia, Subtipo, Falta
ANCHO_DEMO = 11
ANCHOS_COLA = [16, 11]                      # Trans, Fila_Origen


# ── Helpers de estructura del formulario ──────────────────────────────
def es_estado(h):
    return norm(h).endswith("ESTADO")


def limpiar_patologia(header):
    """'18.- ¿ TIENE  DEPRESIÓN ?' -> 'Depresión'. Solo si LIMPIAR_NOMBRE_PATOLOGIA."""
    n = num_pregunta(header)
    if n in OVERRIDE_PATOLOGIA:
        return OVERRIDE_PATOLOGIA[n]
    s = re.sub(r"^\s*\d+\s*\.\-\s*", "", str(header))
    s = s.replace("¿", "").replace("?", "")
    s = re.sub(r"^\s*(TIENE|ES|SUFRE DE|PRESENTA|PADECE)\s+", "", s, flags=re.I)
    s = re.sub(r"^\s*PACIENTE\s+PRESENTA\s+", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:1].upper() + s[1:].lower() if s else s


def limpiar_subtipo(valor, subtipo_header):
    """Nombre de subtipo limpio. 1) aplica OVERRIDE_SUBTIPO (mapa por Nº de
    pregunta del subtipo; ej. Ansiedad Q43 junta los dos 'Pánico'). 2) si no hay
    regla, recorta el sustantivo del header ('Depresión Moderada' con
    '20.- TIPO DE DEPRESIÓN' -> 'Moderada'; sin 'TIPO DE' deja el valor)."""
    if valor in (None, ""): return ""
    s = str(valor).strip()
    reglas = OVERRIDE_SUBTIPO.get(num_pregunta(subtipo_header))
    if reglas:
        vn = norm(s)
        for keys, nombre in reglas:
            if any(norm(k) in vn for k in keys):
                return nombre
    m = re.search(r"TIPO DE\s+(.+)$", norm(subtipo_header))
    if m:
        noun = m.group(1).strip()
        if norm(s).startswith(noun):
            rec = s[len(noun):].strip(" -,:")
            if rec: s = rec
    return s


def flag_demo(cell, regla):
    n = norm(cell)
    if regla == "_no_vacio":
        return "SI" if n not in NEGATIVOS_DEMO else ""
    if regla == "_no_chileno":
        return "SI" if (n and "CHILE" not in n) else ""
    if isinstance(regla, (list, tuple)):
        return "SI" if any(norm(k) in n for k in regla) else ""
    return ""


def encontrar_diagnostico(headers, c0):
    """Camina a la izquierda del ESTADO saltando subtipos/estados hasta dar
    con la pregunta del diagnóstico. Devuelve índice 0-based o c0 si no hay."""
    d = c0 - 1
    while d >= 0:
        h = headers[d]
        n = num_pregunta(h)
        if es_estado(h) or (n in SUBTIPO_NUMS) or (n is None):
            d -= 1
            continue
        return d
    return c0


# ── DETECCIÓN Y VALIDACIÓN DE FORMATO ─────────────────────────────────
# La detección del eje IRIS/Admin vive en `formatos.detectar_eje` (compartida con
# A03/estamentos, con las firmas RAYEN estándar). Acá solo los validadores con los
# mensajes específicos del formulario 'Control de Salud Mental'.
def detectar_formato(ws):
    """'iris' | 'administrativo' | 'desconocido' (firmas RAYEN estándar)."""
    return formatos.detectar_eje(ws)


_MSG_DESCONOCIDO = (
    "No reconozco este archivo como un export de 'Control de Salud Mental'.\n\n"
    "No encontré las firmas ni del formato IRIS ni del Administrativo.\n\n"
    "Asegúrate de descargarlo desde IRIS o desde el Reporte Administrativo de "
    "RAYEN, y de no haberlo modificado (no borres filas ni columnas del export)."
)


def validar_iris(ws):
    """Acepta el export IRIS. Si es el Administrativo, avisa que cambie el
    selector de formato. Si no es ninguno, 'desconocido'."""
    fmt = detectar_formato(ws)
    if fmt == "iris":
        return
    if fmt == "administrativo":
        raise ArchivoInvalido(
            "administrativo",
            "Elegiste el formato IRIS, pero este archivo parece el REPORTE "
            "ADMINISTRATIVO de RAYEN.\n\n"
            "Cambia el selector de formato a «Administrativo», o descarga el "
            "export de IRIS (Formularios RAYEN → 'Control de Salud Mental').")
    raise ArchivoInvalido("desconocido", _MSG_DESCONOCIDO)


def validar_admin(ws):
    """Acepta el Reporte Administrativo. Si es el IRIS, avisa que cambie el
    selector. Si no es ninguno, 'desconocido'."""
    fmt = detectar_formato(ws)
    if fmt == "administrativo":
        return
    if fmt == "iris":
        raise ArchivoInvalido(
            "iris",
            "Elegiste el formato Administrativo, pero este archivo parece el "
            "export de IRIS.\n\n"
            "Cambia el selector de formato a «IRIS» (es el recomendado: trae "
            "todas las columnas).")
    raise ArchivoInvalido("desconocido", _MSG_DESCONOCIDO)


# ── PERFILES DE FORMATO ───────────────────────────────────────────────
_DISCLAIMER_ADMIN = (
    "⚠ Formato ADMINISTRATIVO: esta herramienta fue hecha para IRIS.\n"
    "Funciona, pero OJO:\n"
    "  • Las columnas Pueblos Originarios, SENAME, Proteccion Niñez, Migrante y\n"
    "    Trans NO son calculables desde este export → saldrán VACÍAS.\n"
    "  • La edad se toma de 'Edad de registro formulario' (formato en años).\n"
    "    y no de la edad real del Usuario que se informa por fecha de nacimiento.\n"
    "Revisa los resultados antes de tabular."
)

PERFIL_IRIS = {
    "id": "iris",
    "nombre": "IRIS · Formularios Clinicos Control de Salud Mental  (recomendado)",
    "ancla": formatos.ANCLA_IRIS,
    "usar_blanco_en_a": True,
    "n_hardcode": 16,
    "validar": validar_iris,
    "disclaimer": "",
}
PERFIL_ADMIN = {
    "id": "administrativo",
    "nombre": "RAYEN · Reporte Administrativo",
    "ancla": formatos.ANCLA_ADMIN,
    "usar_blanco_en_a": formatos.HEADER_ADMIN["usar_blanco_en_a"],   # col A con blancos -> no fiarse
    "n_hardcode": formatos.HEADER_ADMIN["n_hardcode"],               # encabezado en fila 9 (n_hardcode+1)
    "validar": validar_admin,
    "disclaimer": _DISCLAIMER_ADMIN,
}
PERFILES = [PERFIL_IRIS, PERFIL_ADMIN]


def perfil_por_id(perfil_id):
    for p in PERFILES:
        if p["id"] == perfil_id:
            return p
    return None


# ── Apertura + preparación de columnas ────────────────────────────────
def abrir_validado(entrada, perfil):
    """Carga el workbook, toma la hoja de trabajo y valida contra el `perfil`.
    Devuelve (wb, ws). Levanta ImportError si falta openpyxl, ArchivoInvalido si
    el formato no corresponde al perfil elegido."""
    if not OPENPYXL_OK:
        raise ImportError(
            "Falta la librería 'openpyxl'. Si corres el .py suelto, instálala con:\n"
            "    pip install openpyxl\n"
            f"(detalle: {OPENPYXL_ERR})"
        )
    wb = openpyxl.load_workbook(entrada)
    ws = wb[HOJA] if HOJA else wb.active
    perfil["validar"](ws)   # portón: rechaza archivos que no son el formato elegido
    return wb, ws


def _preparar(ws, perfil, log):
    """Ubica encabezado (según el perfil) y detecta las columnas compartidas
    (RUT/edad/sexo/género + demografía). Devuelve un dict de contexto."""
    header_idx, modo = encontrar_fila_encabezado(
        ws, perfil["ancla"], perfil["usar_blanco_en_a"], perfil["n_hardcode"],
        MAX_FILAS_BUSQUEDA_HEADER)
    log(f"[corte] perfil={perfil['id']} modo={modo} | encabezado en fila {header_idx} "
        f"(la hoja original NO se modifica)")

    ncols = ws.max_column
    headers = [ws.cell(row=header_idx, column=c).value for c in range(1, ncols + 1)]
    headers_norm = [norm(h) for h in headers]
    num2col = {num_pregunta(h): i for i, h in enumerate(headers) if num_pregunta(h) is not None}

    # RUT/edad/sexo: se aceptan los nombres de AMBOS formatos (robusto ante mala elección)
    rut_col, edad_col, sexo_col = formatos.resolver_identidad(headers_norm)
    if edad_col is None or edad_col > ncols:
        edad_col = ANIO_COL_FALLBACK if ANIO_COL_FALLBACK <= ncols else None
    log(f"[cols] RUT=col{rut_col} | Edad=col{edad_col} | Sexo=col{sexo_col}")

    demo_cols = {flag: (buscar_col(headers_norm, tokens=[norm(t) for t in src]), regla)
                 for flag, (src, regla) in DEMOGRAFIA.items()}
    genero_col = buscar_col(headers_norm, tokens=[norm(t) for t in COL_GENERO_TOKENS])
    fecha_col = buscar_col(headers_norm, tokens=["FECHA", "FORMULARIO"])   # mes: FECHA FORMULARIO (IRIS y Admin)
    faltantes = [f for f, (c, _) in demo_cols.items() if not c]
    log("[demo] " + " ".join(f"{f}=col{c}" for f, (c, _) in demo_cols.items()) +
        f" Trans(Genero)=col{genero_col}")
    if faltantes:
        log(f"[demo] columnas AUSENTES en este formato (saldrán vacías): {', '.join(faltantes)}")

    return {
        "header_idx": header_idx, "ncols": ncols, "headers": headers,
        "headers_norm": headers_norm, "num2col": num2col,
        "rut_col": rut_col, "edad_col": edad_col, "sexo_col": sexo_col,
        "genero_col": genero_col, "demo_cols": demo_cols, "fecha_col": fecha_col,
    }


# ── MOTOR DE MARCADO (compartido egresos/ingresos, agnóstico al formato) ──
def marcar_eventos(wb, ws, perfil, *, busquedas, tipo_label, orden_tipos, hoja_salida,
                   tipo_col_header, avisar_sin_subtipo=frozenset(), etiqueta="evento",
                   mes=None, log=print):
    """Recorre las filas, detecta eventos por token de ESTADO (`busquedas`), les
    agrega patología + subtipo + demografía y escribe la hoja `hoja_salida`
    (formato largo: 1 fila por evento). NO guarda el workbook.

    `perfil` fija la ubicación de encabezado/columnas (formato IRIS o admin).
    `mes` = (año, mes) filtra los formularios por FECHA FORMULARIO a ese mes; None
    procesa el archivo completo. Si se pide un mes sin ningún formulario en el
    archivo, levanta ArchivoInvalido (fail loud, no cuenta silenciosa en 0).
    Devuelve {'total','por_tipo','falta_subtipo','hoja'}.
    """
    ctx = _preparar(ws, perfil, log)
    headers = ctx["headers"]; ncols = ctx["ncols"]; num2col = ctx["num2col"]
    header_idx = ctx["header_idx"]
    rut_col = ctx["rut_col"]; edad_col = ctx["edad_col"]; sexo_col = ctx["sexo_col"]
    genero_col = ctx["genero_col"]; demo_cols = ctx["demo_cols"]; fecha_col = ctx["fecha_col"]

    mes_activo = mes is not None
    if mes_activo and not fecha_col:
        raise ArchivoInvalido(
            "sin_fecha",
            "Pediste filtrar por un mes, pero no encuentro la columna «FECHA "
            "FORMULARIO» en este export.\n\nCarga el archivo tal como lo descargas "
            "de RAYEN/IRIS (sin borrar columnas), o procesa el archivo completo.")
    if mes_activo:
        log(f"[mes] filtrando por FECHA FORMULARIO = {mes[1]:02d}/{mes[0]}")

    busq = {k: [norm(t) for t in v] for k, v in busquedas.items()}
    estado_idx = [c0 for c0 in range(ncols) if es_estado(headers[c0])]   # fijo: precomputado
    eventos = []
    filas_en_mes = 0        # formularios que caen en el mes pedido
    filas_fecha_mala = 0    # formularios con RUT pero fecha ilegible (se excluyen)

    for r in range(header_idx + 1, ws.max_row + 1):
        fila = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        fila_n = [norm(v) for v in fila]
        rut  = fila[rut_col - 1]  if rut_col  else ""
        if mes_activo:
            if not str(rut or "").strip():
                continue   # fila vacía de relleno: no cuenta
            ym = mes_de_celda(fila[fecha_col - 1])
            if ym is None:
                filas_fecha_mala += 1
                continue
            if ym != mes:
                continue
            filas_en_mes += 1
        edad = edad_anios(fila[edad_col - 1]) if edad_col else None
        sexo = fila[sexo_col - 1] if sexo_col else ""
        demo = {flag: (flag_demo(fila[c - 1], regla) if c else "")
                for flag, (c, regla) in demo_cols.items()}
        trans = ""
        if genero_col:
            g = fila[genero_col - 1]
            if "TRANS" in norm(g):
                trans = str(g)

        for k, toks in busq.items():
            for c0 in estado_idx:
                if not all(t in fila_n[c0] for t in toks):
                    continue
                d = encontrar_diagnostico(headers, c0)
                diag_num = num_pregunta(headers[d])
                if diag_num in EXCLUIR_PATOLOGIA:
                    continue   # epilepsia / programas -> no van al A05 SM
                if diag_num in REMAP_DIAGNOSTICO:
                    pat, sub = REMAP_DIAGNOSTICO[diag_num]
                else:
                    pat = (limpiar_patologia(headers[d]) if LIMPIAR_NOMBRE_PATOLOGIA
                           else str(headers[d]))
                    sub_num = DIAGNOSTICOS_CON_SUBTIPO.get(diag_num)
                    sub_col0 = num2col.get(sub_num) if sub_num else None
                    sub_val = fila[sub_col0] if sub_col0 is not None else ""
                    sub = limpiar_subtipo(sub_val, headers[sub_col0] if sub_col0 is not None else "")
                falta_sub = ""
                if (k in avisar_sin_subtipo and diag_num in DIAGNOSTICOS_CON_SUBTIPO
                        and not str(sub).strip()):
                    falta_sub = "SI"
                ev = {"rut": rut, "edad": edad, "sexo": sexo, "tipo": k,
                      "pat": pat, "sub": sub, "falta_sub": falta_sub,
                      "trans": trans, "fila": r}
                ev.update(demo)
                eventos.append(ev)

    if mes_activo:
        if filas_fecha_mala:
            log(f"[mes] AVISO: {filas_fecha_mala} formulario(s) con fecha ilegible "
                f"quedaron FUERA del filtro (revisa la columna FECHA FORMULARIO).")
        if filas_en_mes == 0:
            raise ArchivoInvalido(
                "mes_vacio",
                f"No hay formularios de {mes[1]:02d}/{mes[0]} en este archivo.\n\n"
                "Revisa el mes/año elegido, o que el export cubra ese período. "
                "Si querías todo, elige «Archivo completo».")
        log(f"[mes] {filas_en_mes} formulario(s) en {mes[1]:02d}/{mes[0]}")

    # ── hoja nueva (la original intacta) ──
    if hoja_salida in wb.sheetnames:
        del wb[hoja_salida]
    ws2 = wb.create_sheet(hoja_salida)
    demo_keys = list(DEMOGRAFIA.keys())
    cols = (["RUT", "Edad_Formulario", "Sexo", tipo_col_header, "Patologia", "Subtipo",
             "Falta_Subtipo"] + demo_keys + ["Trans", "Fila_Origen"])
    ws2.append(cols)
    eventos.sort(key=lambda e: (orden_tipos.get(e["tipo"], 9), str(e["pat"]), str(e["sub"])))
    for e in eventos:
        fila_out = [e["rut"], e["edad"], e["sexo"], tipo_label.get(e["tipo"], e["tipo"]),
                    e["pat"], e["sub"], e["falta_sub"]]
        fila_out += [e.get(fk, "") for fk in demo_keys]
        fila_out += [e["trans"], e["fila"]]
        ws2.append(fila_out)

    # presentación amigable
    from openpyxl.styles import Font
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"
    if ws2.max_row >= 1:
        ws2.auto_filter.ref = ws2.dimensions
    from openpyxl.utils import get_column_letter
    anchos = ANCHOS_BASE + [ANCHO_DEMO] * len(demo_keys) + ANCHOS_COLA
    for i, w in enumerate(anchos, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    log(f"[ok] hoja '{hoja_salida}' escrita")

    n = {}
    for e in eventos:
        n[e["tipo"]] = n.get(e["tipo"], 0) + 1
    n_falta = sum(1 for e in eventos if e["falta_sub"] == "SI")
    log(f"[resumen] {etiqueta}s: {len(eventos)}")
    if avisar_sin_subtipo:
        log(f"          sin subtipo (debiendo tenerlo): {n_falta}")
    for k in busquedas:
        log(f"          {tipo_label.get(k, k)}: {n.get(k, 0)}")

    return {
        "total": len(eventos),
        "por_tipo": {tipo_label.get(k, k): n.get(k, 0) for k in busquedas},
        "falta_subtipo": n_falta,
        "hoja": hoja_salida,
    }
