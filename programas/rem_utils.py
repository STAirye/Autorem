#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# Version: 1.7.9
#
# This program is free software: you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by the
# Free Software Foundation, either version 3 of the License, or (at your
# option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTABILITY
# or FITNESS FOR A PARTICULAR PURPOSE. See the GNU General Public License
# for more details: <https://www.gnu.org/licenses/>.
# ==========================================================================
"""
rem_utils.py — utilidades comunes para los módulos del REM (RAYEN/IRIS).

Base compartida por todas las herramientas del proyecto autoREM. NO contiene
lógica de ninguna sección específica del REM; cada módulo de tarea
(modulos/rem_<pestaña>_<casilla>_<descriptor>.py) importa de aquí.

Contenido:
  - Guarda de dependencia (openpyxl) y excepción común (ArchivoInvalido).
  - Normalización y parsing de celdas: norm, to_year, solo_entero.
  - Búsqueda de columnas y numeración de preguntas RAYEN: buscar_col, num_pregunta.
  - Localización de la fila de encabezado (salta banner + filtros): encontrar_fila_encabezado.
  - Utilidad de SO: abrir_carpeta.
"""

import re
import sys
from pathlib import Path   # reexport de conveniencia para los módulos

# ── Versión del proyecto (fuente única de verdad) ──
# Convención X.Y.Z (ver CLAUDE.md §9):
#   X = programa · Y = módulos de programa acumulados · Z = corrección del módulo actual.
# Todos los .py comparten esta versión en su header; bumpear aquí al cambiarla.
VERSION = "1.7.9"

# openpyxl es la única dependencia externa real. En el .exe va empaquetado;
# corriendo como .py suelto puede faltar -> los módulos avisan con instrucciones.
try:
    import openpyxl
    OPENPYXL_OK = True
    OPENPYXL_ERR = ""
except ImportError as _e:
    openpyxl = None
    OPENPYXL_OK = False
    OPENPYXL_ERR = str(_e)


class ArchivoInvalido(Exception):
    """El archivo no es el export que el módulo esperaba.

    `categoria` la define cada módulo según sus propios formatos (por ejemplo,
    el A05 usa 'administrativo' | 'desconocido'). El mensaje es para el usuario.
    """
    def __init__(self, categoria, mensaje):
        self.categoria = categoria
        super().__init__(mensaje)


# ── Normalización y parsing de celdas ─────────────────────────────────
def norm(v):
    """Texto en MAYÚSCULA, sin tildes/ñ, con espacios colapsados. '' si es None/NaN."""
    if v is None or v != v: return ""   # v != v capta NaN (float) → '' (celda vacía)
    s = str(v).upper().strip()
    for a, b in zip("ÁÉÍÓÚÜÑ", "AEIOUUN"): s = s.replace(a, b)
    return re.sub(r"\s+", " ", s)


def to_year(v):
    """Primeros 4 dígitos de la celda como int (año). None si no hay dígitos."""
    if v is None: return None
    d = re.sub(r"\D", "", str(v))
    return int(d[:4]) if d else None


def solo_entero(v):
    """Primer entero que aparezca en la celda. None si no hay ninguno."""
    if v is None: return None
    m = re.search(r"\d+", str(v))
    return int(m.group()) if m else None


def edad_anios(v):
    """Edad en AÑOS (int) desde la celda. RAYEN a veces la trae pelada (IRIS:
    'AÑO APLICACIÓN FORMULARIO' = número) y a veces como texto (Administrativo:
    '99 años 12 meses 31 días'). Menor de 1 año (solo meses/días) -> 0. None si
    no hay dato. Genérica: sirve a cualquier export RAYEN."""
    if v is None:
        return None
    n = norm(v)                       # '99 ANOS 12 MESES 31 DIAS'
    m = re.search(r"(\d+)\s*ANOS?", n)
    if m:
        return int(m.group(1))
    if "MES" in n or "DIA" in n:      # menor de 1 año expresado en meses/días
        return 0
    return solo_entero(v)


# ── Búsqueda de columnas / numeración de preguntas RAYEN ──────────────
def buscar_col(headers_norm, tokens=None, exacto=None):
    """Índice 1-based de la 1ª columna cuyo header (ya normalizado) contiene
    TODOS los `tokens`, o coincide exactamente con `exacto`. None si no hay."""
    for i, h in enumerate(headers_norm, 1):
        if exacto is not None and h == norm(exacto): return i
        if tokens and all(t in h for t in tokens): return i
    return None


def num_pregunta(header):
    """Número 'N.-' con que RAYEN antepone cada pregunta ('18.- ¿...?' -> 18)."""
    m = re.match(r"\s*(\d+)\s*\.\-", str(header))
    return int(m.group(1)) if m else None


# ── Localización de la fila de encabezado (banner + filtros arriba) ────
def encontrar_fila_encabezado(ws, ancla, usar_blanco_en_a=True,
                              n_hardcode=16, max_filas=60):
    """Ubica la fila del encabezado real saltando el banner y los filtros que
    RAYEN pone arriba. Cascada de detección:
      1. Fila que contiene TODOS los tokens de `ancla` (match por substring).
      2. Si `usar_blanco_en_a`: la fila siguiente a la 1ª con la columna A vacía.
      3. Fallback: `n_hardcode` (+1).
    Devuelve (fila_encabezado_1based, modo)."""
    tope = min(ws.max_row, max_filas)
    ancla_n = [norm(t) for t in ancla]
    for r in range(1, tope + 1):
        vals = [norm(c.value) for c in ws[r]]
        if all(any(tok in v for v in vals) for tok in ancla_n):
            return r, "ancla"
    if usar_blanco_en_a:
        for r in range(1, tope + 1):
            if norm(ws.cell(row=r, column=1).value) == "":
                return r + 1, "blanco_en_A"
    return n_hardcode + 1, "hardcode"


# ── Lectura + clasificación de reportes (compartido; usado por módulos pandas) ──
def leer_xlsx(entrada, ancla=None, max_scan=40):
    """Lee un .xlsx con openpyxl y devuelve (headers, filas_de_datos). ROBUSTO a
    la 'dimension' rota de los exports copy-paste / de BD (con la que
    pandas.read_excel leería 0 filas). `ancla` = nombres de columna que deben
    estar TODOS en la fila de encabezado; si es None, toma la 1ª fila con >3
    celdas llenas."""
    ws = openpyxl.load_workbook(entrada, data_only=True, read_only=True).active
    filas = list(ws.iter_rows(values_only=True))
    ws.parent.close()
    if ancla:
        want = {norm(a) for a in ancla}
        hi = next((i for i, r in enumerate(filas[:max_scan]) if want <= {norm(v) for v in r}), 0)
    else:
        hi = next((i for i, r in enumerate(filas[:max_scan]) if sum(v not in (None, "") for v in r) > 3), 0)
    return list(filas[hi]), filas[hi + 1:]


def verificar_hoja_unica(entrada):
    """Guarda de integridad de los exports RAYEN/IRIS: SIEMPRE bajan UNA hoja con
    datos + 2 vacías. Si hay datos en más de una hoja, el archivo fue MODIFICADO
    (típicamente se le agregó una tabla dinámica) y sus resultados no son confiables
    → levanta ArchivoInvalido. Robusta a la 'dimension' rota igual que leer_xlsx:
    itera celdas (corta en la 1ª no vacía), no confía en max_row."""
    wb = openpyxl.load_workbook(entrada, data_only=True, read_only=True)
    con_datos = [ws.title for ws in wb.worksheets
                 if any(any(v not in (None, "") for v in row)
                        for row in ws.iter_rows(values_only=True))]
    wb.close()
    if len(con_datos) > 1:
        raise ArchivoInvalido(
            "modificado",
            f"El archivo trae datos en {len(con_datos)} hojas ({', '.join(con_datos)}). "
            "RAYEN/IRIS siempre baja UNA hoja con datos + 2 vacías, así que esto significa "
            "que el export fue MODIFICADO (¿le agregaste una tabla dinámica?). Vuelve a "
            "descargarlo SIN tocar desde RAYEN/IRIS y cárgalo tal cual.")


def resolver_columnas(headers, mapa):
    """Mapea nombres canónicos -> columnas reales por nombre SEMÁNTICO. `mapa` =
    {canonico: opcion | [opcion, ...]} con opcion = ('exact','NOMBRE') |
    ('subs',['TOK1','TOK2']). Una LISTA de opciones se prueba EN ORDEN (para
    formatos alternativos, p.ej. IRIS vs Admin). Match normalizado (tildes,
    mayúsculas, renumeración 'N.- '). Devuelve {canonico: nombre_columna | None}."""
    hn = [(c, norm(c)) for c in headers]

    def _match(modo, obj):
        if modo == "exact":
            t = norm(obj)
            return next((c for c, n in hn if n == t), None)
        toks = [norm(t) for t in obj]
        return next((c for c, n in hn if all(t in n for t in toks)), None)

    out = {}
    for canon, spec in mapa.items():
        opciones = spec if isinstance(spec, list) else [spec]
        out[canon] = next((c for modo, obj in opciones if (c := _match(modo, obj)) is not None), None)
    return out


# Mapa de columnas del reporte de ATENCIONES, compartido por los módulos que lo
# consumen (respiratorio A23, y a futuro cualquier otro programa). Alternativas
# IRIS | Admin ('Monitoreo de Actividades').
# ⚠ El Monitoreo admin queda INCOMPLETO (IRIS es la fuente plena):
#   (a) DEMOGRAFÍA: no trae nacionalidad/pueblo/fecha nac/nombres → origen-migrante
#       y nombre completo salen vacíos; edad desde 'AÑOS'.
#   (b) DIAGNÓSTICO EN TEXTO, sin código ICD → los indicadores que matchean por
#       código (Ira Alta 'j0', Bronquitis 'J20', EPOC Exacerbado 'J44.1') salen 0.
#       Los de texto (Neumonía/Influenza/Coqueluche) y los de ACTIVIDAD (KTR,
#       espirometría, controles…) sí cuadran (verificado vs IRIS, jul-2026).
#   (c) Estructura PADRE-HIJO -> forward-fill de cabecera en cargar_atenciones.
MAPA_ATENCIONES = {
    "RUN":     [("subs", ["NUMERO", "IDENTIFICACION"]), ("exact", "RUN")],
    "ATENID":  [("subs", ["ATEN", "ID"]), ("subs", ["ATENCION", "ID"])],  # solo IRIS
    "FECHA":   [("exact", "FECHA ATENCION"), ("exact", "FECHA CONSULTA")],
    "ACT":     [("exact", "ACTIVIDADES"), ("subs", ["ACTIVIDAD", "PROCEDIMIENTO"])],
    "DIAG":    [("exact", "DIAGNOSTICOS"), ("exact", "DIAGNOSTICO")],
    "INSTR":   ("exact", "INSTRUMENTO"),
    "PROF":    ("subs", ["PROFESIONAL", "ATENCION"]),   # NOMBRE del funcionario que atendió (IRIS)
    "TIPO":    ("subs", ["TIPO", "ATENCION"]),          # 'TIPO ATENCION' | 'TIPO DE ATENCION'
    "SEXO":    ("exact", "SEXO"),
    "SECTOR":  ("exact", "SECTOR"),
    "NACION":  ("exact", "NACIONALIDAD"),               # solo IRIS
    "EMIG":    [("subs", ["ES", "IMIGRANTE"]), ("subs", ["ES", "INMIGRANTE"])],  # flag migrante, IRIS
    "ALERTAS": ("subs", ["ALERTAS", "ADMINISTRATIVAS"]),  # SENAME/Mejor Niñez/Cuidador…, IRIS
    "FORMCLIN": ("subs", ["FORMULARIOS", "CLINICOS"]),  # para detectar gestante, IRIS
    "PUEBLO":  ("subs", ["PUEBLO", "ORIGINARIO"]),      # solo IRIS
    "FNAC":    ("subs", ["FECHA", "NACIMIENTO"]),       # solo IRIS
    "NOMBRES": ("exact", "NOMBRES"),                    # solo IRIS
    "APAT":    ("subs", ["APELLIDO", "PATERNO"]),       # solo IRIS
    "AMAT":    ("subs", ["APELLIDO", "MATERNO"]),       # solo IRIS
    "ANOS":    ("exact", "AÑOS"),                        # edad a la DESCARGA
    "ANOS_AT": [("subs", ["AÑOS", "ATENCION"]), ("subs", ["ANOS", "ATENCION"])],  # edad a la ATENCIÓN (REM)
}


def cargar_canonico(entrada, ancla, resolver):
    """Lee UNO o VARIOS .xlsx (los reportes acumulativos necesitan varios años) y
    arma el DataFrame canónico concatenado. `resolver(headers) -> {canon: columna}`.
    Devuelve (df, col_del_primero)."""
    import pandas as pd
    partes, col0 = [], None
    for e in (entrada if isinstance(entrada, (list, tuple)) else [entrada]):
        verificar_hoja_unica(e)          # rechaza exports modificados (datos en >1 hoja)
        hdr, filas = leer_xlsx(e, ancla=ancla)
        col = resolver(hdr)
        col0 = col0 or col
        idx = {c: i for i, c in enumerate(hdr)}
        partes.append(pd.DataFrame(
            {k: [f[idx[c]] if c is not None and idx[c] < len(f) else None for f in filas]
             for k, c in col.items()}))
    d = pd.concat(partes, ignore_index=True) if len(partes) > 1 else partes[0]
    return d, col0


def cargar_atenciones(entrada):
    """Export(s) de ATENCIONES (IRIS ó Monitoreo admin) -> DataFrame canónico +
    textos normalizados (act/diag/instr/tipo) + FECHA parseada. Ruta o lista."""
    import pandas as pd
    d, col = cargar_canonico(entrada, None, lambda h: resolver_columnas(h, MAPA_ATENCIONES))
    faltan = [k for k in ("RUN", "FECHA", "ACT", "DIAG", "INSTR", "TIPO") if not col[k]]
    if faltan:
        raise ValueError("No reconozco el export de atenciones (¿IRIS o Monitoreo "
                         "admin?); faltan columnas: " + ", ".join(faltan))
    # Monitoreo admin: estructura PADRE-HIJO — una atención se abre en varias filas
    # de actividad, con RUN y datos de cabecera SOLO en la 1ª. Se rellena la cabecera
    # a las filas HIJAS (RUN vacío) para atribuir cada actividad/diagnóstico a su
    # paciente. En IRIS (sin filas hijas) NO se toca nada (evita contaminar campos
    # vacíos legítimos entre pacientes distintos).
    cab = ["RUN", "ATENID", "FECHA", "INSTR", "PROF", "TIPO", "SEXO", "SECTOR", "NACION",
           "EMIG", "ALERTAS", "FORMCLIN", "PUEBLO", "FNAC", "NOMBRES", "APAT",
           "AMAT", "ANOS", "ANOS_AT"]
    child = d["RUN"].replace("", pd.NA).isna()
    if child.any():
        dd = d[cab].replace("", pd.NA)
        d[cab] = dd.where(~child, dd.ffill())
    d["FECHA"] = pd.to_datetime(d["FECHA"], errors="coerce", dayfirst=True)
    for k in ("ACT", "DIAG", "INSTR", "TIPO"):
        d[k + "_n"] = d[k].map(norm)
    return d


# ── Maestro de Actividades (catálogo RAYEN: actividad ↔ estamento ↔ casilla REM) ──
# Referencia transversal (no solo SM): mapea cada actividad a su NUM REM oficial. Una
# fila por (actividad × instrumento). Banner en fila 1 → ancla en la fila de headers.
# Se actualiza ~semestralmente. Usos: clasificar (¿tributa y a qué casilla?), validar
# la clasificación RAYEN vs la nuestra (el referente técnico arbitra), y chequear que
# cada estamento tenga las actividades que el exe reporta.
MAPA_MAESTRO = {
    "ACT":     ("exact", "ACTIVIDAD"),
    "INSTR":   [("subs", ["INSTRUMENTO", "ASOCIADO"]), ("exact", "INSTRUMENTO")],
    "NUMREM":  ("subs", ["NUM", "REM"]),
    "SECCION": ("subs", ["NUM", "SECCION"]),
    "REM":     ("exact", "REM"),
}


def cargar_maestro(entrada):
    """Lee el 'Maestro de Actividades' -> DataFrame (ACT/INSTR/NUMREM/SECCION/REM) +
    ACT_n y NUMREM_n normalizados. Una fila por (actividad × instrumento). Acepta el
    Maestro COMPLETO (.xlsx, con banner) o el SLIM comprimido (.csv.gz que shippea el
    repo, generado por tools/slim_maestro.py)."""
    import pandas as pd
    ent = str(entrada).lower()
    if ent.endswith((".csv", ".gz", ".csv.gz")):
        raw = pd.read_csv(entrada, dtype=str, keep_default_na=False)   # compresión inferida
        hdr = list(raw.columns)
        col = resolver_columnas(hdr, MAPA_MAESTRO)
        _guard_maestro(col)
        d = pd.DataFrame({k: (raw[c] if c is not None else "") for k, c in col.items()})
    else:
        hdr, filas = leer_xlsx(entrada, ancla=["ACTIVIDAD", "INSTRUMENTO ASOCIADO", "NUM REM"])
        col = resolver_columnas(hdr, MAPA_MAESTRO)
        _guard_maestro(col)
        idx = {c: i for i, c in enumerate(hdr)}
        d = pd.DataFrame({k: [f[idx[c]] if c is not None and idx[c] < len(f) else None
                              for f in filas] for k, c in col.items()})
    d = d[d["ACT"].map(lambda x: x not in (None, ""))].copy()
    d["ACT_n"] = d["ACT"].map(norm)
    d["NUMREM_n"] = d["NUMREM"].map(norm)
    return d


def _guard_maestro(col):
    if not col["ACT"] or not col["NUMREM"]:
        raise ValueError("No reconozco el Maestro de Actividades (faltan 'ACTIVIDAD' "
                         "y/o 'NUM REM'). ¿Es el archivo correcto?")


# ── Grilla de agregación edad×sexo (tablas REM copy-paste; SM · A23 · A03) ──
# Forma del template MINSAL: Ambos·Hombres·Mujeres + cada banda etaria × (H/M).
# Bandas inclusive; el último tramo (80,200) = '80 y más'. A04 separa <1 y 1-4;
# A06/A32/A03·D.3 agrupan 0-4. Compartido para no re-implementar en cada módulo.
BANDAS_A04 = [(0, 0), (1, 4), (5, 9), (10, 14), (15, 19), (20, 24), (25, 29),
              (30, 34), (35, 39), (40, 44), (45, 49), (50, 54), (55, 59),
              (60, 64), (65, 69), (70, 74), (75, 79), (80, 200)]
LBL_A04 = ["<1", "1-4", "5-9", "10-14", "15-19", "20-24", "25-29", "30-34",
           "35-39", "40-44", "45-49", "50-54", "55-59", "60-64", "65-69",
           "70-74", "75-79", "80+"]
BANDAS_A06 = [(0, 4)] + BANDAS_A04[2:]
LBL_A06 = ["0-4"] + LBL_A04[2:]


def _mujer(s):
    n = norm(s); return ("FEMENIN" in n) or ("MUJER" in n)


def _hombre(s):
    n = norm(s); return ("MASCULIN" in n) or ("HOMBRE" in n)


def _band_idx(edad, bandas):
    import pandas as pd
    if pd.isna(edad):
        return None
    e = int(edad)
    return next((i for i, (lo, hi) in enumerate(bandas) if lo <= e <= hi), None)


def _isum(s):
    """Suma robusta a Series vacías (evita el '' de una object-Series vacía)."""
    return int(s.sum()) if len(s) else 0


def grid(sub, bandas, lbls, con_sexo=True):
    """Fila de conteos en el ORDEN del template: Ambos·Hombres·Mujeres y luego cada
    banda × sexo (o solo total por banda si con_sexo=False). `sub` = DataFrame con
    columnas 'sexo' y 'edad'."""
    hom = sub["sexo"].map(_hombre)
    muj = sub["sexo"].map(_mujer)
    bi = sub["edad"].map(lambda x: _band_idx(x, bandas))
    if con_sexo:
        out = {"Ambos": len(sub), "Hombres": _isum(hom), "Mujeres": _isum(muj)}
        for i, l in enumerate(lbls):
            m = bi.eq(i)
            out[f"{l} H"] = _isum(m & hom)
            out[f"{l} M"] = _isum(m & muj)
    else:
        out = {"Total": len(sub)}
        for i, l in enumerate(lbls):
            out[l] = _isum(bi.eq(i))
    return out


def _rango_mes(mes):
    """(inicio, fin) Timestamp del mes de reporte. mes=(año,mes) o None -> mes
    anterior. Compartido por los módulos pandas (A23, SM Actividades, Trabajo
    Perdido); los cálculos van hacia atrás desde el mes reportado, NO TODAY()."""
    import calendar
    from datetime import date
    import pandas as pd
    if mes is None:
        hoy = date.today()
        y, m = (hoy.year, hoy.month - 1) if hoy.month > 1 else (hoy.year - 1, 12)
    else:
        y, m = mes
    return pd.Timestamp(y, m, 1), pd.Timestamp(y, m, calendar.monthrange(y, m)[1])


def mes_de_celda(v):
    """(año, mes) desde una celda de fecha RAYEN, o None si no se puede parsear.
    Acepta datetime/date (lo que openpyxl entrega en celdas con formato fecha) y
    texto en las DOS formas que usa RAYEN: 'YYYY/MM/DD' (Admin/monitoreo) y
    'DD/MM/YYYY' (formularios clínicos), con '-', '/' o '.' y hora opcional.

    Distingue las dos formas por ESTRUCTURA (año de 4 dígitos adelante = ISO),
    NO por dayfirst a ciegas: 'YYYY/MM/DD' con dayyfirst se lee como el mes
    equivocado en silencio (justo lo que no queremos). Cualquier otra cosa -> None
    (se excluye y se avisa; nunca se adivina un mes plausible pero errado)."""
    if v is None:
        return None
    from datetime import datetime as _dt, date as _date
    if isinstance(v, (_dt, _date)):
        return (v.year, v.month)
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.]\d{1,2}", s)      # YYYY/MM/DD
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
    else:
        m = re.match(r"^\d{1,2}[-/.](\d{1,2})[-/.](\d{4})", s)  # DD/MM/YYYY
        if not m:
            return None
        mo, y = int(m.group(1)), int(m.group(2))
    return (y, mo) if 1 <= mo <= 12 else None


def maestro_rem_map(dfm):
    """{ACT_n -> NUM REM normalizado} desde el Maestro (dedup por actividad; una
    actividad tiene el mismo NUM REM en todas sus filas de instrumento)."""
    return dict(zip(dfm["ACT_n"], dfm["NUMREM_n"]))


# ── Demografía por atención (columnas del REM: SENAME, migrante, pueblos…) ──
# Fuente = ADA IRIS (ALERTAS ADMINISTRATIVAS / ES IMIGRANTE / PUEBLO / DIAG / ACT).
# El grupal y el Monitoreo admin NO traen estos campos -> todos los flags quedan
# en False (limitación conocida). Reutilizable por cualquier módulo (A23, SM…).
_PUEBLO_VACIO = {"NINGUNO", "NO SABE", "NO CONTESTA", "NO INFORMADO", ""}


def marcar_demografia(d):
    """Agrega columnas booleanas `dem_*` por atención. Devuelve el mismo df.
    dem_migrante · dem_originario · dem_sename · dem_mejorninez · dem_cuidador ·
    dem_demencia · dem_campana. (Gestante es a nivel RUN -> `gestante_runs`.)"""
    cols = ("dem_migrante", "dem_originario", "dem_sename", "dem_mejorninez",
            "dem_cuidador", "dem_campana", "dem_demencia")
    if len(d) == 0:                       # df vacío: columnas booleanas vacías
        for c in cols:
            d[c] = False
        return d

    def alerta(*subs):   # ALERTAS ADMIN contiene ALGUNA subcadena
        s = d["ALERTAS"].map(norm) if "ALERTAS" in d else None
        if s is None:
            return d.get("RUN", d.index).map(lambda _: False)
        m = None
        for x in subs:
            c = s.str.contains(norm(x), regex=False, na=False)
            m = c if m is None else (m | c)
        return m.fillna(False)

    emig = d["EMIG"].map(norm) if "EMIG" in d else None
    pueblo = d["PUEBLO"].map(norm) if "PUEBLO" in d else None
    d["dem_migrante"] = ((emig == "SI") if emig is not None else False) | alerta("MIGRANTE")
    d["dem_originario"] = (~pueblo.isin(_PUEBLO_VACIO)) if pueblo is not None else False
    d["dem_sename"] = alerta("SENAME")
    d["dem_mejorninez"] = alerta("MEJOR NINEZ", "SPE EX MEJOR")
    d["dem_cuidador"] = alerta("CUIDADOR")
    d["dem_campana"] = d["ACT_n"].str.contains(norm("campaña de invierno"), regex=False, na=False)
    d["dem_demencia"] = d["DIAG_n"].str.contains(norm("demencia"), regex=False, na=False)
    return d


def trans_map(entrada):
    """Lee el 'Informe Inscritos y Adscritos' -> dict RUN -> 'M' | 'F' | 'X' para
    personas TRANS. RAYEN ahora permite seleccionar TRANS directamente en GÉNERO
    ('Transgénero Masculino/Femenina', 'Femenino Trans'), así que se usa ESE campo
    (la vieja heurística género≠sexo quedó obsoleta y era ruidosa). La dirección
    (M/F) alimenta el split TRANS Masculino/Femenina del template. Archivo ENORME
    (toda la población del CESFAM) → se carga solo si el usuario lo aporta."""
    hdr, filas = leer_xlsx(entrada)
    hn = [norm(h) for h in hdr]

    def ci(*subs):
        return next((i for i, n in enumerate(hn) if all(norm(s) in n for s in subs)), None)

    i_run = ci("NUMERO", "IDENTIFICACION")
    if i_run is None:
        i_run = ci("RUN")
    i_gen = ci("GENERO")
    if i_run is None or i_gen is None:   # archivo modificado o reporte equivocado
        falta = " y ".join(c for c, i in [("RUN", i_run), ("GÉNERO", i_gen)] if i is None)
        raise ValueError(f"el 'Informe Inscritos' no trae la(s) columna(s) {falta}. "
                         "¿Está modificado o es otro reporte? Descárgalo de nuevo SIN tocar.")
    out = {}
    for f in filas:
        g = norm(f[i_gen]) if i_gen < len(f) else ""
        if "TRANS" in g:
            out[str(f[i_run]).strip()] = "M" if "MASCULIN" in g else "F" if "FEMENIN" in g else "X"
    return out


def atenid_multiprofesional(entrada):
    """Set de ATEN ID MULTIPROFESIONALES (2+ profesionales), del reporte 'Monitoreo
    Multiprofesional': la columna 'Multiprofesional-1' no vacía = tuvo ≥1 profesional
    adicional. Sirve para marcar composición (Un Profesional vs Dos o Más). Opcional:
    sin este reporte, todo se asume mono-profesional."""
    hdr, filas = leer_xlsx(entrada)
    hn = [norm(h) for h in hdr]

    def ci(*subs):
        return next((i for i, n in enumerate(hn) if all(norm(s) in n for s in subs)), None)

    i_aten, i_m1 = ci("ATEN", "ID"), ci("MULTIPROFESIONAL", "1")
    if i_aten is None or i_m1 is None:
        raise ValueError("el 'Monitoreo Multiprofesional' no trae ATEN ID / "
                         "Multiprofesional-1. ¿Modificado o reporte equivocado?")
    return {str(f[i_aten]).strip() for f in filas if i_m1 < len(f) and norm(f[i_m1])}


def gestante_runs(d, ini, fin):
    """Set de RUNs GESTANTES (patrón PowerBI): atención de MATRONA con 'control
    prenatal' (o formulario clínico 'gestante') en la ventana [ini, fin] (3 meses
    terminando en el mes reportado). Requiere columnas del ADA IRIS."""
    w = d[(d["FECHA"] >= ini) & (d["FECHA"] <= fin)]
    if w.empty:
        return set()
    mat = w["INSTR_n"].str.contains(norm("matron"), regex=False, na=False)
    prenatal = w["ACT_n"].str.contains(norm("control prenatal"), regex=False, na=False)
    form = (w["FORMCLIN"].map(norm).str.contains(norm("gestante"), regex=False, na=False)
            if "FORMCLIN" in w else prenatal & False)
    return set(w.loc[mat & (prenatal | form), "RUN"])


def contiene_todos(serie, *subs):
    """Máscara booleana pandas: la Serie (ya normalizada) contiene TODAS las
    subcadenas (case/acento-insensible vía norm). Para clasificar reportes."""
    m = None
    for x in subs:
        c = serie.str.contains(norm(x), regex=False, na=False)
        m = c if m is None else (m & c)
    return m


def contiene_alguno(serie, subs):
    """Máscara booleana pandas: la Serie contiene ALGUNA de las subcadenas."""
    m = None
    for x in subs:
        c = serie.str.contains(norm(x), regex=False, na=False)
        m = c if m is None else (m | c)
    return m


# ── Utilidad de SO ────────────────────────────────────────────────────
def abrir_carpeta(carpeta: Path):
    """Abre el explorador de archivos en `carpeta` (Windows / macOS / Linux)."""
    import subprocess
    try:
        if sys.platform.startswith("win"):
            import os
            os.startfile(str(carpeta))   # noqa
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(carpeta)])
        else:
            subprocess.Popen(["xdg-open", str(carpeta)])
    except Exception:
        pass
