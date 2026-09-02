#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# ==========================================================================
"""
Pruebas automáticas de autoREM. Datos 100% SINTÉTICOS (sin PII).

Correr:
    python tests/test_autorem.py        # runner propio, imprime PASS/FAIL
    pytest tests/                        # si tienes pytest instalado

Cubre: equivalencia de egresos vs la versión validada (v1.2), edad_anios,
ingresos, perfiles IRIS/Administrativo (detección, columnas, demografía
ausente), validación cruzada de formato, y el dispatcher multi-hoja.
"""

import importlib.util
import sys
import tempfile
from pathlib import Path

import openpyxl

# ── Acceso al código del proyecto (carpeta padre de tests/) ──
REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

import programas.rem_saludmental as sm       # noqa: E402
import modulos.rem_a05_o_egresos as egresos    # noqa: E402
import modulos.rem_a05_n_ingresos as ingresos  # noqa: E402
import autorem                               # noqa: E402

_TMP = Path(tempfile.mkdtemp(prefix="autorem_tests_"))


# ── Fixtures sintéticas ───────────────────────────────────────────────
def _iris_fixture():
    """Export estilo IRIS: 1 egreso (Alta+subtipo) y 1 ingreso."""
    p = _TMP / "iris.xlsx"
    if p.exists():
        return p
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Servicio de Salud Metropolitano Central", None, None])
    ws.append(["Filtros: bla", None, None])
    ws.append(["NUMERO TIPO IDENTIFICACION", "AÑO APLICACIÓN FORMULARIO", "SEXO",
               "1.- ¿Usted es Madre de Hijo menor de 5 años?", "PUEBLO ORIGINARIO",
               "ALERTAS ADMINISTRATIVAS", "GÉNERO", "18.- ¿ TIENE DEPRESIÓN ?",
               "18.- ESTADO", "20.- TIPO DE DEPRESIÓN"])
    ws.append(["11111111-1", 45, "Mujer", "SI", "", "MIGRANTE; PRAIS", "Femenino",
               "SI", "EGRESO ALTA", "Depresión Moderada"])
    ws.append(["22222222-2", 30, "Hombre", "", "Mapuche", "", "Trans Masculino",
               "SI", "INGRESO", "Depresión Severa"])
    wb.save(p)
    return p


def _admin_fixture():
    """Export estilo Administrativo (nombres/columnas reales): banner + header en
    fila 9, RUT pelado, edad en texto; sin columnas ALERTAS/PUEBLO/GÉNERO."""
    p = _TMP / "admin.xlsx"
    if p.exists():
        return p
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Servicio de Salud", "S.S. Metropolitano Central"])
    ws.append(["Comuna", "Maipu"])
    ws.append(["Establecimiento", "[CESFAM] Dr. Luis Ferrada"])
    ws.append(["Año", "2026"]); ws.append(["Mes", "JULIO"])
    ws.append([None, None, "Reporte Formularios RAYEN"])
    ws.append([None, None, "Formulario: Control de Salud Mental"])
    ws.append([])  # fila 8 vacía
    ws.append(["RUT", "Numero de Fichas", "Paciente", "Edad de registro formulario",
               "Sexo", "Sector", "Dirección", "Telefono 1", "Telefono 2", "Prevision",
               "Convenio", "Fecha Formulario", "Funcionario",
               "1.- ¿Usted es Madre de Hijo menor de 5 años?",
               "18.- ¿ Tiene Depresión ?", "19.- Estado", "20.- Tipo de depresión"])
    ws.append(["11111111-1", "F1", "", "45 años 3 meses 2 días", "Mujer", "A", "", "", "",
               "Fonasa", "Fonasa D", "2026/07/06", "Func", "SI", "SI", "Ingreso",
               "Depresión Moderada"])
    ws.append(["22222222-2", "F2", "", "8 meses 10 días", "Hombre", "B", "", "", "",
               "Fonasa", "Fonasa D", "2026/07/06", "Func", "", "SI", "Egreso Alta", ""])
    wb.save(p)
    return p


def _iris_fecha_fixture():
    """Export IRIS con columna FECHA FORMULARIO: 2 egresos en 07/2026, 1 en 08/2026.
    Fecha en texto DD/MM/YYYY (formato ambiguo -> prueba el dayfirst)."""
    p = _TMP / "iris_fecha.xlsx"
    if p.exists():
        return p
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Servicio de Salud", None]); ws.append(["Filtros: bla", None])
    ws.append(["NUMERO TIPO IDENTIFICACION", "AÑO APLICACIÓN FORMULARIO", "SEXO",
               "FECHA FORMULARIO", "18.- ¿ TIENE DEPRESIÓN ?", "18.- ESTADO",
               "20.- TIPO DE DEPRESIÓN"])
    ws.append(["11111111-1", 45, "Mujer", "06/07/2026", "SI", "EGRESO ALTA", "Depresión Moderada"])
    ws.append(["22222222-2", 30, "Hombre", "20/07/2026", "SI", "EGRESO ALTA", "Depresión Severa"])
    ws.append(["33333333-3", 60, "Mujer", "05/08/2026", "SI", "EGRESO ALTA", "Depresión Leve"])
    wb.save(p)
    return p


def _dump(path, hoja):
    ws = openpyxl.load_workbook(path)[hoja]
    return [tuple(c.value for c in row) for row in ws.iter_rows()]


def _quiet(*_args, **_kwargs):
    pass


# ── Pruebas ───────────────────────────────────────────────────────────
def test_iris_equivalencia_v12():
    """La salida de egresos (perfil IRIS) es idéntica al monolito validado v1.2."""
    monolito = REPO / "legacy" / "rem_marcar_egresos 1.2.py"
    if not monolito.exists():
        print("    SKIP: no está 'legacy/rem_marcar_egresos 1.2.py' (referencia v1.2)")
        return
    spec = importlib.util.spec_from_file_location("viejo_v12", monolito)
    viejo = importlib.util.module_from_spec(spec); spec.loader.exec_module(viejo)

    iris = _iris_fixture()
    out_ref = _TMP / "ref.xlsx"; viejo.procesar(iris, out_ref, log=_quiet)
    # v1.2 usa nombres de patología CRUDOS; acá comparamos el refactor, no el
    # renombre (que es un cambio intencional). Forzamos crudo solo para el test.
    prev = sm.LIMPIAR_NOMBRE_PATOLOGIA
    sm.LIMPIAR_NOMBRE_PATOLOGIA = False
    try:
        out_new = _TMP / "new.xlsx"; egresos.procesar(iris, out_new, log=_quiet)
    finally:
        sm.LIMPIAR_NOMBRE_PATOLOGIA = prev
    assert _dump(out_new, "A05_Egresos") == _dump(out_ref, "A05_Egresos")


def test_nombre_patologia_y_exclusion():
    """LIMPIAR_NOMBRE_PATOLOGIA=True: nombre canónico + epilepsia (75) excluida."""
    p = _TMP / "nombres.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Servicio", None]); ws.append(["Filtros", None])
    ws.append(["NUMERO TIPO IDENTIFICACION", "AÑO APLICACIÓN FORMULARIO", "SEXO",
               "18.- ¿ TIENE DEPRESIÓN ?", "18.- ESTADO", "20.- TIPO DE DEPRESIÓN",
               "75.- ¿Paciente Presenta Epilepsia?", "76.- Estado"])
    ws.append(["11111111-1", 50, "Hombre", "SI", "EGRESO ALTA", "Depresión Moderada",
               "SI", "EGRESO ALTA"])          # 1 egreso depresión + 1 epilepsia (se excluye)
    wb.save(p)
    out = _TMP / "nombres_out.xlsx"
    egresos.procesar(p, out, log=_quiet)      # LIMPIAR=True por defecto
    ws2 = openpyxl.load_workbook(out)["A05_Egresos"]
    head = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    pi = head.index("Patologia") + 1
    pats = [ws2.cell(row=r, column=pi).value for r in range(2, ws2.max_row + 1)]
    assert pats == ["Depresión"]              # epilepsia excluida; nombre canónico


def test_subtipo_ansiedad_override():
    """Ansiedad Q43: nombres cortos + los DOS pánicos se juntan en 'Pánico'."""
    h = "43.- TIPO DE TRASTORNO DE ANSIEDAD"
    assert sm.limpiar_subtipo("Trastorno de Pánico", h) == "Pánico"
    assert sm.limpiar_subtipo("Trastorno de Pánico sin Agorofobia", h) == "Pánico"   # merge
    assert sm.limpiar_subtipo("Fobias Sociales", h) == "Fobia social"
    assert sm.limpiar_subtipo("Trastornos de Ansiedad Generalizada", h) == "Generalizada"
    assert sm.limpiar_subtipo("Trastorno de Estrés Post Traumático", h) == "TEPT"
    assert sm.limpiar_subtipo("Otros Trastornos de Ansiedad", h) == "Otros"
    # los otros subtipos siguen por recorte del header:
    assert sm.limpiar_subtipo("Depresión Moderada", "20.- TIPO DE DEPRESIÓN") == "Moderada"
    assert sm.limpiar_subtipo("Moderado", "45.- ETAPA") == "Moderado"


def test_madre_menor5_filtra_por_sexo_no_por_genero():
    """Pregunta 1 marcada en un hombre -> se ANULA (por definición no cuenta).
    Sexo femenino con género transmasculino -> SÍ cuenta (el filtro es por SEXO)."""
    p = _TMP / "madre.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Servicio", None]); ws.append(["Filtros", None])
    ws.append(["NUMERO TIPO IDENTIFICACION", "AÑO APLICACIÓN FORMULARIO", "SEXO",
               "1.- ¿Usted es Madre de Hijo menor de 5 años?", "GÉNERO",
               "18.- ¿ TIENE DEPRESIÓN ?", "18.- ESTADO"])
    ws.append(["11111111-1", 30, "Mujer",  "SI", "Femenino",       "SI", "EGRESO ALTA"])
    ws.append(["22222222-2", 40, "Hombre", "SI", "Masculino",      "SI", "EGRESO ALTA"])
    ws.append(["33333333-3", 25, "Mujer",  "SI", "Trans Masculino", "SI", "EGRESO ALTA"])
    wb.save(p)
    out = _TMP / "madre_out.xlsx"
    avisos = []
    egresos.procesar(p, out, log=avisos.append)
    ws2 = openpyxl.load_workbook(out)["A05_Egresos"]
    head = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    ci = head.index("Madre_menor5") + 1
    ruts = [ws2.cell(row=r, column=1).value for r in range(2, ws2.max_row + 1)]
    madre = {ws2.cell(row=r, column=1).value: (ws2.cell(row=r, column=ci).value or "")
             for r in range(2, ws2.max_row + 1)}   # celda vacía vuelve como None
    assert len(ruts) == 3                      # las 3 filas siguen siendo egresos
    assert madre["11111111-1"] == "SI"         # mujer cis: cuenta
    assert madre["22222222-2"] == ""           # hombre mal marcado: ANULADO
    assert madre["33333333-3"] == "SI"         # transmasculino de sexo femenino: cuenta
    assert any("Madre_menor5" in a and "NO " in a for a in avisos)   # avisa, no calla


def test_edad_anios():
    assert sm.edad_anios("45 años 3 meses 2 días") == 45
    assert sm.edad_anios("8 meses 10 días") == 0        # menor de 1 año
    assert sm.edad_anios(45) == 45                        # IRIS: número pelado
    assert sm.edad_anios(None) is None


def test_ingresos_iris():
    out = _TMP / "ing_iris.xlsx"
    ingresos.procesar(_iris_fixture(), out, log=_quiet)
    rows = _dump(out, "A05_Ingresos")
    assert rows[0][3] == "Tipo_Ingreso"
    ruts = {r[0] for r in rows[1:]}
    assert ruts == {"22222222-2"}           # solo el ingreso; ignora el egreso
    assert rows[1][5] == "Severa"           # subtipo recortado


def test_admin_egresos():
    out = _TMP / "adm_eg.xlsx"
    egresos.procesar(_admin_fixture(), out, perfil=sm.PERFIL_ADMIN, log=_quiet)
    rows = _dump(out, "A05_Egresos")
    assert len(rows) == 2                    # 1 egreso (Egreso Alta)
    r = rows[1]
    assert r[0] == "22222222-2"              # RUT desde columna 'RUT'
    assert r[1] == 0                         # edad <1 año -> 0
    assert r[6] == "SI"                      # Falta_Subtipo (Depresión sin subtipo)
    # demográficas que no existen en el Administrativo -> vacías
    assert all(r[i] in (None, "") for i in (8, 9, 10, 11, 12))


def test_admin_ingresos():
    out = _TMP / "adm_in.xlsx"
    ingresos.procesar(_admin_fixture(), out, perfil=sm.PERFIL_ADMIN, log=_quiet)
    rows = _dump(out, "A05_Ingresos")
    assert len(rows) == 2
    r = rows[1]
    assert r[0] == "11111111-1"
    assert r[1] == 45                        # '45 años 3 meses...' -> 45
    assert r[5] == "Moderada"


def test_validacion_cruzada():
    """El perfil equivocado levanta ArchivoInvalido con la categoría correcta."""
    def categoria(entrada, perfil):
        try:
            sm.abrir_validado(entrada, perfil)
            return "OK"
        except sm.ArchivoInvalido as e:
            return e.categoria
    assert categoria(_admin_fixture(), sm.PERFIL_IRIS) == "administrativo"
    assert categoria(_iris_fixture(), sm.PERFIL_ADMIN) == "iris"


def test_mes_de_celda():
    """Parseo de fecha: datetime directo, texto DD/MM/YYYY (dayfirst), y basura."""
    from datetime import datetime
    from programas.rem_utils import mes_de_celda
    assert mes_de_celda(datetime(2026, 7, 6)) == (2026, 7)
    assert mes_de_celda("06/07/2026") == (2026, 7)     # dayfirst: 6 de julio
    assert mes_de_celda("2026/07/06") == (2026, 7)     # ISO-ish
    assert mes_de_celda("") is None
    assert mes_de_celda(None) is None
    assert mes_de_celda("no es fecha") is None


def test_filtro_mes():
    """mes=(año,mes) filtra por FECHA FORMULARIO; mes ausente -> ArchivoInvalido."""
    fx = _iris_fecha_fixture()
    # Julio: 2 egresos
    out = _TMP / "mes_jul.xlsx"
    egresos.procesar(fx, out, log=_quiet, mes=(2026, 7))
    ruts = {r[0] for r in _dump(out, "A05_Egresos")[1:]}
    assert ruts == {"11111111-1", "22222222-2"}
    # Agosto: 1 egreso
    out = _TMP / "mes_ago.xlsx"
    egresos.procesar(fx, out, log=_quiet, mes=(2026, 8))
    ruts = {r[0] for r in _dump(out, "A05_Egresos")[1:]}
    assert ruts == {"33333333-3"}
    # Archivo completo: los 3
    out = _TMP / "mes_todo.xlsx"
    egresos.procesar(fx, out, log=_quiet, mes=None)
    assert len(_dump(out, "A05_Egresos")) == 4         # header + 3
    # Mes sin formularios -> error ruidoso, NO archivo con 0 filas
    try:
        egresos.procesar(fx, _TMP / "mes_vacio.xlsx", log=_quiet, mes=(2026, 1))
        assert False, "debió levantar ArchivoInvalido por mes vacío"
    except sm.ArchivoInvalido as e:
        assert e.categoria == "mes_vacio"


def test_dispatcher_multisheet():
    """Correr egresos+ingresos juntos -> un archivo con ambas hojas."""
    perfil = sm.perfil_por_id("iris")
    resultados, salida = autorem._correr_tareas(
        autorem.TAREAS, _iris_fixture(), perfil, log=_quiet)
    wb = openpyxl.load_workbook(salida)
    assert {"A05_Egresos", "A05_Ingresos"} <= set(wb.sheetnames)
    assert len(resultados) == 2


# ── Runner propio (sin depender de pytest) ────────────────────────────
def _main():
    pruebas = [v for k, v in sorted(globals().items())
               if k.startswith("test_") and callable(v)]
    fallos = 0
    for fn in pruebas:
        try:
            fn()
            print(f"PASS  {fn.__name__}")
        except AssertionError as e:
            fallos += 1
            print(f"FAIL  {fn.__name__}  -> {e or 'assert'}")
        except Exception as e:  # noqa: BLE001
            fallos += 1
            print(f"ERROR {fn.__name__}  -> {type(e).__name__}: {e}")
    print("-" * 50)
    total = len(pruebas)
    print(f"{total - fallos}/{total} OK" + (f"  ({fallos} con problemas)" if fallos else ""))
    return 1 if fallos else 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    sys.exit(_main())
