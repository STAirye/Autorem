#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# SPDX-License-Identifier: GPL-3.0-or-later
# ==========================================================================
"""Pruebas del módulo REM SM Trabajo Perdido + guarda multi-hoja. Datos SINTÉTICOS.
    python tests/test_trabajo_perdido.py"""

import sys
import tempfile
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

import modulos.rem_sm_trabajo_perdido as tp     # noqa: E402
from programas.rem_utils import ArchivoInvalido, verificar_hoja_unica   # noqa: E402

_TMP = Path(tempfile.mkdtemp(prefix="autorem_tp_"))

_ADA_HDR = ["NUMERO TIPO IDENTIFICACION", "ATEN ID", "FECHA ATENCION", "ACTIVIDADES",
            "DIAGNOSTICOS", "INSTRUMENTO", "PROFESIONAL ATENCION", "TIPO ATENCION",
            "SEXO", "AÑOS ATENCION"]
_ADA_K = {"run": 0, "id": 1, "fecha": 2, "act": 3, "dg": 4, "instr": 5,
          "prof": 6, "tipo": 7, "sexo": 8, "edad": 9}

_MAESTRO_HDR = ["ACTIVIDAD", "INSTRUMENTO ASOCIADO", "NUM REM", "NUM SECCION", "REM"]


def _mk_ada(rows, nombre="ada.xlsx"):
    p = _TMP / nombre
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(_ADA_HDR)
    for r in rows:
        line = [""] * len(_ADA_HDR)
        for k, v in r.items():
            line[_ADA_K[k]] = v
        ws.append(line)
    wb.save(p)
    return p


def _mk_maestro(pares, nombre="maestro.xlsx"):
    """pares = [(actividad, num_rem), ...]. Banner en fila 1 (como el real)."""
    p = _TMP / nombre
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Maestro de Actividades"])          # banner
    ws.append(_MAESTRO_HDR)
    for act, numrem in pares:
        ws.append([act, "Psicólogo(a)", numrem, "", "desc"])
    wb.save(p)
    return p


def _a(act, prof, run="1-1", instr="Psicólogo(a)", fecha="10/07/2026"):
    return dict(act=act, prof=prof, run=run, instr=instr, fecha=fecha, id=f"AT{run}{act[:3]}",
                dg="x", tipo="Espontánea", sexo="Femenino", edad="30 años")


def _quiet(*_a, **_k): pass


_MAESTRO = [
    ("Consulta De Salud Mental", "REM-A04"),
    ("Controles Salud Mental", "REM-A06"),
    ("AG_Alta programa salud mental", "REM-Gestion"),
    ("AG_Atencion de salud mental adolescente", "REM-Gestion"),
    ("Confirmacion diagnostica salud mental", "REM-A03"),      # otro REM real
    ("Curacion simple", "REM-A28"),                            # no SM
]


def _run(ada_rows, con_maestro=True, mes=(2026, 7)):
    ada = _mk_ada(ada_rows)
    maestro = _mk_maestro(_MAESTRO) if con_maestro else None
    E = tp.procesar(ada, maestro=maestro, mes=mes, log=_quiet)
    return E


# -- Tests --------------------------------------------------------------

def test_gestion_es_perdida():
    E = _run([_a("Consulta De Salud Mental", "ANA"),          # A04 -> tributa
              _a("AG_Alta programa salud mental", "JUAN"),    # Gestion -> PERDIDA
              _a("Controles Salud Mental", "ANA")])           # A06 -> tributa
    assert len(E) == 1, f"esperaba 1 perdida, hubo {len(E)}"
    assert E.iloc[0]["num_rem"].upper() == "REM-GESTION", E.iloc[0]["num_rem"]
    assert E.iloc[0]["profesional"] == "JUAN"


def test_otro_rem_real_tambien_es_perdida():
    # Definición elegida (referente): todo lo no-SM-estadístico es perdido, incluso A03.
    E = _run([_a("Confirmacion diagnostica salud mental", "MARIA")])
    assert len(E) == 1
    assert E.iloc[0]["num_rem"].upper() == "REM-A03"


def test_no_sm_se_ignora():
    E = _run([_a("Curacion simple", "PEDRO")])       # sin mental/demencia -> ni se mira
    assert len(E) == 0


def test_actividad_nueva_no_en_maestro():
    # 'Taller mental raro' no está en el Maestro y no matchea heurística -> PERDIDA.
    # 'Controles Salud Mental por chat' no está en el Maestro pero SÍ matchea heur -> tributa.
    E = _run([_a("Taller mental raro inventado", "LUCIA"),
              _a("Controles Salud Mental por chat", "LUCIA")])
    assert len(E) == 1, f"esperaba 1, hubo {len(E)}"
    assert "NO EN MAESTRO" in E.iloc[0]["num_rem"].upper()


def test_por_funcionario_rankea():
    E = _run([_a("AG_Alta programa salud mental", "JUAN"),
              _a("AG_Atencion de salud mental adolescente", "JUAN"),
              _a("AG_Alta programa salud mental", "ANA")])
    pf = E.attrs["tablas"]["Por_Funcionario"]
    assert pf.iloc[0]["Funcionario"] == "JUAN"
    assert int(pf.iloc[0]["N a saco vacío"]) == 2


def test_sin_maestro_usa_heuristica():
    # Sin Maestro: 'AG_Alta programa salud mental' es SM-ish y no matchea heurística -> PERDIDA.
    E = _run([_a("AG_Alta programa salud mental", "JUAN"),
              _a("Consulta De Salud Mental", "ANA")], con_maestro=False)
    assert len(E) == 1
    assert E.iloc[0]["num_rem"] == "(sin maestro)"


def test_solo_del_mes():
    E = _run([_a("AG_Alta programa salud mental", "JUAN", fecha="10/07/2026"),
              _a("AG_Alta programa salud mental", "JUAN", fecha="10/06/2026")])
    assert len(E) == 1, "solo la atención de julio cuenta"


def test_guarda_multihoja_rechaza():
    # ADA con datos en 2 hojas = modificado -> el loader debe rechazar.
    p = _TMP / "modificado.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(_ADA_HDR); ws.append(["1-1", "AT1", "10/07/2026", "Consulta De Salud Mental",
                                    "x", "Médico(a)", "ANA", "Esp", "F", "30 años"])
    wb.create_sheet("TablaDinamica").append(["pivote", "aqui"])
    wb.save(p)
    try:
        tp.procesar(p, mes=(2026, 7), log=_quiet)
        assert False, "debió rechazar el archivo modificado"
    except ArchivoInvalido as e:
        assert e.categoria == "modificado"


def test_guarda_una_hoja_pasa():
    p = _mk_ada([_a("Consulta De Salud Mental", "ANA")])   # 1 hoja con datos -> pasa
    verificar_hoja_unica(p)   # no debe levantar


def _main():
    pruebas = [v for k, v in sorted(globals().items())
               if k.startswith("test_") and callable(v)]
    fallos = 0
    for fn in pruebas:
        try:
            fn(); print(f"PASS  {fn.__name__}")
        except AssertionError as e:
            fallos += 1; print(f"FAIL  {fn.__name__} -> {e or 'assert'}")
        except Exception as e:  # noqa: BLE001
            import traceback
            fallos += 1; print(f"ERROR {fn.__name__} -> {type(e).__name__}: {e}")
            traceback.print_exc()
    print("-" * 50)
    print(f"{len(pruebas) - fallos}/{len(pruebas)} OK" + (f" ({fallos} problemas)" if fallos else ""))
    return 1 if fallos else 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    sys.exit(_main())
