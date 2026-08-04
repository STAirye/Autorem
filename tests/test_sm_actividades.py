#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# SPDX-License-Identifier: GPL-3.0-or-later
# ==========================================================================
"""Pruebas del módulo REM SM Actividades. Datos SINTÉTICOS. Correr:
    python tests/test_sm_actividades.py"""

import sys
import tempfile
from datetime import date
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

import modulos.rem_sm_actividades as sm   # noqa: E402

_TMP = Path(tempfile.mkdtemp(prefix="autorem_sm_"))

_ADA_HDR = ["NUMERO TIPO IDENTIFICACION", "ATEN ID", "FECHA ATENCION", "ACTIVIDADES",
            "DIAGNOSTICOS", "INSTRUMENTO", "TIPO ATENCION", "SEXO", "AÑOS ATENCION",
            "ALERTAS ADMINISTRATIVAS", "ES IMIGRANTE", "PUEBLO ORIGINARIO", "FORMULARIOS CLINICOS"]
_GRP_HDR = ["NUMERO TIPO IDENTIFICACION", "FECHA ATENCION", "ACTIVIDADES",
            "ASISTE (SI/NO)", "SEXO", "EDAD", "INSTRUMENTO", "FUNCIONARIO PRESTADOR"]

# alias cortos -> nombre real de columna
_ADA_K = {"run": 0, "id": 1, "fecha": 2, "act": 3, "dg": 4, "instr": 5, "tipo": 6, "sexo": 7, "edad": 8,
          "alertas": 9, "emig": 10, "pueblo": 11, "formclin": 12}
_GRP_K = {"run": 0, "fecha": 1, "act": 2, "asiste": 3, "sexo": 4, "edad": 5, "instr": 6, "prest": 7}


def _mk(hdr, keymap, rows, nombre):
    p = _TMP / nombre
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(hdr)
    for r in rows:
        line = [""] * len(hdr)
        for k, v in r.items():
            line[keymap[k]] = v
        ws.append(line)
    wb.save(p)
    return p


def _mk_ada(rows):
    return _mk(_ADA_HDR, _ADA_K, rows, "ada.xlsx")


def _mk_grupal(rows):
    return _mk(_GRP_HDR, _GRP_K, rows, "grupal.xlsx")


_INS_HDR = ["NUMERO TIPO IDENTIFICACION", "SEXO", "GENERO"]


def _mk_inscritos(rows):
    p = _TMP / "inscritos.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.append(_INS_HDR)
    for r in rows:
        ws.append([r.get(h, "") for h in _INS_HDR])
    wb.save(p)
    return p


def _mk_multi(rows):
    """Monitoreo Multiprofesional sintético: [ATEN ID, Multiprofesional-1]."""
    p = _TMP / "multi.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.append(["ATEN ID", "Multiprofesional-1"])
    for r in rows:
        ws.append([r.get("aten", ""), r.get("m1", "")])
    wb.save(p)
    return p


def _quiet(*_a, **_k): pass


def _run(ada_rows, grupal_rows=None, mes=(2026, 7)):
    ada = _mk_ada(ada_rows)
    grupal = _mk_grupal(grupal_rows) if grupal_rows is not None else None
    E = sm.procesar(ada, grupal=grupal, mes=mes, log=_quiet)
    return E, E.attrs["tablas"]


def _n(E, casilla, sub=None):
    m = E["casilla"] == casilla
    if sub is not None:
        m = m & (E["sub"] == sub)
    return int(m.sum())


def _cell(tabla, col_key, col_val, dato):
    """Valor de `dato` en la fila donde tabla[col_key]==col_val."""
    fila = tabla[tabla[col_key] == col_val].iloc[0]
    return fila[dato]


def _a06_tot(tabla, col):
    """Suma `col` en las filas de estamento de A06 (ya no hay fila TOTAL; el grupal
    'Intervención Psicosocial Grupal' se excluye)."""
    ests = tabla[tabla["Profesional"] != "Intervención Psicosocial Grupal"]
    return int(ests[col].sum())


# ── A04: consultas médicas SM (solo médico) ──
def test_a04_solo_medico():
    E, t = _run([
        {"run": "A", "id": "1", "fecha": date(2026, 7, 3), "act": "Consulta De Salud Mental  ;",
         "instr": "Médico", "sexo": "Mujer", "edad": 30},
        {"run": "B", "id": "2", "fecha": date(2026, 7, 4), "act": "Consulta De Salud Mental  ;",
         "instr": "Psicólogo(a)", "sexo": "Hombre", "edad": 20},   # NO médico -> fuera de A04
    ])
    assert _n(E, "A04") == 1
    assert _cell(t["A04_Consultas_Medicas"], "Consulta", "Salud Mental", "Ambos") == 1
    assert _cell(t["A04_Consultas_Medicas"], "Consulta", "Salud Mental", "Mujeres") == 1


# ── A06: controles por estamento + SENAME excluido ──
def test_a06_controles_estamento_y_sename():
    E, t = _run([
        {"run": "A", "id": "1", "fecha": date(2026, 7, 3), "act": "Controles Salud Mental  ;",
         "instr": "Médico", "sexo": "Hombre", "edad": 40},
        {"run": "B", "id": "2", "fecha": date(2026, 7, 4), "act": "Controles Salud Mental  ;",
         "instr": "Psicólogo(a)", "sexo": "Mujer", "edad": 8},
        {"run": "C", "id": "3", "fecha": date(2026, 7, 5), "act": "Control Salud Mental a Paciente SENAME  ;",
         "instr": "Psicólogo(a)", "sexo": "Hombre", "edad": 12},   # SENAME -> string aparte, excluido
    ])
    assert _n(E, "A06") == 2
    a06 = t["A06_Controles"]
    assert _cell(a06, "Profesional", "Médico/a", "Ambos") == 1
    assert _cell(a06, "Profesional", "Psicólogo/a", "Ambos") == 1
    assert _a06_tot(a06, "Ambos") == 2
    assert _cell(a06, "Profesional", "Psicólogo/a", "5-9 M") == 1   # edad 8, mujer


# ── ADA cuenta por ATEN ID (distinct), no por fila ──
def test_ada_conteo_por_atenid():
    E, _ = _run([
        {"run": "A", "id": "1", "fecha": date(2026, 7, 3), "act": "Controles Salud Mental  ;", "instr": "Médico", "edad": 30},
        {"run": "A", "id": "1", "fecha": date(2026, 7, 3), "act": "Controles Salud Mental  ;", "instr": "Médico", "edad": 30},  # MISMO ATEN ID
        {"run": "A", "id": "2", "fecha": date(2026, 7, 9), "act": "Controles Salud Mental  ;", "instr": "Médico", "edad": 30},  # otra atención
    ])
    assert _n(E, "A06") == 2   # dos ATEN ID distintos (la fila repetida NO suma)


# ── Grupal cuenta por ASISTENCIA (sin dedup) + filtro Asiste=SI ──
def test_grupal_por_asistencia():
    E, t = _run([], grupal_rows=[
        # EDAD en TEXTO ('30 años…') como viene del export crudo del grupal
        {"run": "P", "fecha": date(2026, 7, 5), "act": "Intervencion psicosocial grupal.", "asiste": "SI", "sexo": "Mujer", "edad": "30 años 2 meses 1 día"},
        {"run": "P", "fecha": date(2026, 7, 5), "act": "Intervencion psicosocial grupal.", "asiste": "SI", "sexo": "Mujer", "edad": "31 años"},  # mismo día, 2º taller -> cuenta 2
        {"run": "Q", "fecha": date(2026, 7, 6), "act": "Intervencion psicosocial grupal.", "asiste": "NO", "sexo": "Hombre", "edad": "40 años"},  # no asiste -> fuera
    ])
    assert _n(E, "A06PG") == 2
    pg = _cell_row(t["A06_Controles"], "Profesional", "Intervención Psicosocial Grupal")
    assert pg["Ambos"] == 2 and pg["Mujeres"] == 2
    assert pg["30-34 M"] == 2      # ambas mujeres 30/31 caen en la banda 30-34 (EDAD texto parseada)


# ── Ventana de mes (por FECHA ATENCIÓN) ──
def test_ventana_de_mes():
    E, _ = _run(
        [{"run": "A", "id": "1", "fecha": date(2026, 6, 30), "act": "Controles Salud Mental  ;", "instr": "Médico", "edad": 30},
         {"run": "B", "id": "2", "fecha": date(2026, 7, 1), "act": "Controles Salud Mental  ;", "instr": "Médico", "edad": 30}],
        grupal_rows=[{"run": "P", "fecha": date(2026, 8, 1), "act": "Intervencion psicosocial grupal.", "asiste": "SI", "edad": 30}],
    )
    assert _n(E, "A06") == 1        # solo la de julio
    assert _n(E, "A06PG") == 0      # la grupal es de agosto


# ── A19a: ADA (individual) + grupal, y el guion evita comerse las VDI de A26 ──
def test_a19a_ada_mas_grupal_sin_vdi():
    E, t = _run(
        [{"run": "A", "id": "1", "fecha": date(2026, 7, 3), "instr": "Psicólogo(a)",
          "act": "Consejerías familiares - Temas Prioridad - Con integrante con problema de salud mental (Ind)  ;", "edad": 40},
         {"run": "B", "id": "2", "fecha": date(2026, 7, 4), "instr": "Médico",
          "act": "Visita domiciliaria integral familia con integrante con problema de salud mental - Primera visita  ;", "edad": 50}],
        grupal_rows=[{"run": "C", "fecha": date(2026, 7, 5), "asiste": "SI", "edad": 35,
                      "act": "Consejerías familiares - Temas Prioridad - Con integrante con problema de salud mental (Grp)"}],
    )
    assert _n(E, "A19a", "97") == 2          # 1 ADA + 1 grupal (la VDI NO cuenta acá)
    assert _n(E, "A26") == 1                 # la VDI va a A26
    a19 = t["A19a_Consejerias_Fam"]
    fila = a19[a19["Tema prioridad (familiar)"] == "Con integrante con problema de salud mental"].iloc[0]
    assert fila["Total Actividades"] == 2 and fila["  · desde ADA (individual)"] == 1 and fila["  · desde Grupal"] == 1


# ── A26: split etario 5-9 (A.31) excluye de A.30, + secuencia de visita ──
def test_a26_split_5a9_y_visita():
    E, t = _run([
        {"run": "K", "id": "1", "fecha": date(2026, 7, 3), "instr": "Médico", "edad": 7,
         "act": "Visita domiciliaria integral familia con integrante con problema de salud mental - Primera visita  ;"},
        {"run": "L", "id": "2", "fecha": date(2026, 7, 4), "instr": "Médico", "edad": 45,
         "act": "Visita domiciliaria integral familia con integrante con problema de salud mental - Tercera o más visitas de seguimiento  ;"},
    ])
    a26 = t["A26_VDI_SM"]
    r30 = a26[a26["Concepto"].str.startswith("A.30")].iloc[0]
    r31 = a26[a26["Concepto"].str.startswith("A.31")].iloc[0]
    assert r30["Total"] == 1 and r30["Tercera o Más"] == 1     # adulto 45
    assert r31["Total"] == 1 and r31["Primera Visita"] == 1     # niño 7
    assert r30["Un Profesional"] == 1                          # default mono-profesional


def test_a26_multiprofesional():
    """Con el Monitoreo Multiprofesional, las VDI cuya ATEN ID está en el reporte
    (Multiprofesional-1 no vacío) pasan a 'Dos o Más Prof.'; el resto queda mono."""
    ada = _mk_ada([
        {"run": "K", "id": "1", "fecha": date(2026, 7, 3), "instr": "Médico", "edad": 45,
         "act": "Visita domiciliaria integral familia con integrante con problema de salud mental - Primera visita  ;"},
        {"run": "L", "id": "2", "fecha": date(2026, 7, 4), "instr": "Médico", "edad": 50,
         "act": "Visita domiciliaria integral familia con integrante con problema de salud mental - Primera visita  ;"},
    ])
    mp = _mk_multi([{"aten": "1", "m1": "Enfermero(a)"}, {"aten": "2", "m1": ""}])  # solo la 1 es multi
    E = sm.procesar(ada, multiprofesional=str(mp), mes=(2026, 7), log=_quiet)
    r30 = _cell_row(E.attrs["tablas"]["A26_VDI_SM"], "Concepto",
                    "A.30 Familia con integrante con problema de salud mental")
    assert r30["Total"] == 2 and r30["Dos o Más Prof."] == 1 and r30["Un Profesional"] == 1


# ── A32 F1: desagregado llamada / videollamada / mensaje (video ≠ llamada) ──
def test_a32f1_desagregado():
    E, t = _run([
        {"run": "A", "id": "1", "fecha": date(2026, 7, 3), "instr": "Psicólogo(a)", "edad": 30,
         "act": "Acciones remotas de salud mental - Llamadas telefónicas  ;"},
        {"run": "B", "id": "2", "fecha": date(2026, 7, 4), "instr": "Psicólogo(a)", "edad": 20,
         "act": "Acciones remotas de salud mental - Videollamadas  ;"},
        {"run": "C", "id": "3", "fecha": date(2026, 7, 5), "instr": "Psicólogo(a)", "edad": 40,
         "act": "Acciones remotas de salud mental - Mensajería de texto  ;"},
    ])
    f1 = t["A32_F1_Acciones_Remotas"]
    assert _cell(f1, "Vía", "Llamadas Telefónicas", "Total") == 1
    assert _cell(f1, "Vía", "Videollamadas", "Total") == 1
    assert _cell(f1, "Vía", "Mensajería de Texto", "Total") == 1
    assert _n(E, "A32F1") == 3


# ── A27: A = asistentes (usuarios), B = sesiones (por prestador/fecha/actividad) ──
def test_a27_asistentes_y_sesiones():
    E, t = _run([], grupal_rows=[
        {"run": "X", "fecha": date(2026, 7, 10), "asiste": "SI", "edad": 30, "prest": "Dra A",
         "act": "Educación en grupo - Prevención de salud mental - Prevención trastorno mental"},
        {"run": "Y", "fecha": date(2026, 7, 10), "asiste": "SI", "edad": 40, "prest": "Dra A",
         "act": "Educación en grupo - Prevención de salud mental - Prevención trastorno mental"},  # misma sesión
    ])
    a27 = t["A27_Educacion_Prev"]
    fila = a27[a27["Área temática"] == "Prevención trastorno mental"].iloc[0]
    assert fila["A · Asistentes (usuarios)"] == 2
    assert fila["B · Sesiones (actividades)"] == 1


# ── Demografía: SENAME / Mejor Niñez / migrante / pueblo / demencia (Beneficiarios=todos) ──
def test_demografia_flags():
    E, t = _run([
        {"run": "A", "id": "1", "fecha": date(2026, 7, 3), "act": "Controles Salud Mental  ;",
         "instr": "Médico", "edad": 40, "emig": "SI", "pueblo": "Mapuche"},
        {"run": "B", "id": "2", "fecha": date(2026, 7, 4), "act": "Controles Salud Mental  ;",
         "instr": "Psicólogo(a)", "edad": 30, "alertas": "Programa SENAME - Ambulatorio",
         "dg": "F03.X-Demencia, No Especificada"},
        {"run": "C", "id": "3", "fecha": date(2026, 7, 5), "act": "Controles Salud Mental  ;",
         "instr": "Médico", "edad": 10, "alertas": "SPE ex Mejor Niñez- Ambulatorio", "pueblo": "Ninguno"},
    ])
    a06 = t["A06_Controles"]
    assert _a06_tot(a06, "Beneficiarios") == 3        # todos (Fonasa)
    assert _a06_tot(a06, "Migrantes") == 1            # A
    assert _a06_tot(a06, "Pueblos Originarios") == 1  # A (Mapuche); C=Ninguno no cuenta
    assert _a06_tot(a06, "SENAME") == 1               # B
    assert _a06_tot(a06, "Prot. Especializada") == 1  # C
    assert _a06_tot(a06, "Demencia") == 1             # B (norm: 'demencia' vs DIAG en MAYÚSCULA)
    assert _a06_tot(a06, "TRANS Masculino") == 0 and _a06_tot(a06, "TRANS Femenina") == 0   # sin inscritos


# ── Gestante: matrona + control prenatal en la ventana → flag en el evento SM ──
def test_gestante_flag():
    E, _ = _run([
        {"run": "G", "id": "1", "fecha": date(2026, 7, 2), "act": "Control Prenatal  ;",
         "instr": "Matron(a)", "edad": 25},
        {"run": "G", "id": "2", "fecha": date(2026, 7, 10), "act": "Controles Salud Mental  ;",
         "instr": "Psicólogo(a)", "edad": 25},
        {"run": "H", "id": "3", "fecha": date(2026, 7, 11), "act": "Controles Salud Mental  ;",
         "instr": "Médico", "edad": 40},
    ])
    a06 = E[E["casilla"] == "A06"]
    assert bool(a06.loc[a06["run"] == "G", "dem_gestante"].iloc[0]) is True
    assert bool(a06.loc[a06["run"] == "H", "dem_gestante"].iloc[0]) is False


def test_trans_inscritos_modificado():
    """Inscritos SIN columna GÉNERO (archivo modificado/otro reporte): trans_map
    levanta ValueError y procesar NO crashea (TRANS queda en 0 con aviso)."""
    from programas.rem_utils import trans_map
    p = _TMP / "inscritos_malo.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["NUMERO TIPO IDENTIFICACION", "SEXO"]); ws.append(["T", "Mujer"])  # sin GÉNERO
    wb.save(p)
    try:
        trans_map(p); raised = False
    except ValueError:
        raised = True
    assert raised
    ada = _mk_ada([{"run": "T", "id": "1", "fecha": date(2026, 7, 3),
                    "act": "Controles Salud Mental  ;", "instr": "Médico", "edad": 30}])
    E = sm.procesar(ada, inscritos=str(p), mes=(2026, 7), log=_quiet)   # no debe crashear
    a06 = E.attrs["tablas"]["A06_Controles"]
    assert _a06_tot(a06, "TRANS Masculino") == 0 and _a06_tot(a06, "TRANS Femenina") == 0


def _cell_row(tabla, col_key, col_val):
    return tabla[tabla[col_key] == col_val].iloc[0]


# ── TRANS: selección explícita en GÉNERO del padrón de Inscritos, split M/F ──
def test_trans_flag():
    ins = _mk_inscritos([
        {"NUMERO TIPO IDENTIFICACION": "T", "SEXO": "Mujer", "GENERO": "Transgénero Masculino"},
        {"NUMERO TIPO IDENTIFICACION": "U", "SEXO": "Hombre", "GENERO": "Masculino"},       # cis
        {"NUMERO TIPO IDENTIFICACION": "V", "SEXO": "Hombre", "GENERO": "Femenino Trans"},
    ])
    ada = _mk_ada([
        {"run": "T", "id": "1", "fecha": date(2026, 7, 3), "act": "Controles Salud Mental  ;", "instr": "Médico", "edad": 30},
        {"run": "U", "id": "2", "fecha": date(2026, 7, 4), "act": "Controles Salud Mental  ;", "instr": "Médico", "edad": 30},
        {"run": "V", "id": "3", "fecha": date(2026, 7, 5), "act": "Controles Salud Mental  ;", "instr": "Psicólogo(a)", "edad": 30},
    ])
    E = sm.procesar(ada, inscritos=ins, mes=(2026, 7), log=_quiet)
    a06 = E.attrs["tablas"]["A06_Controles"]
    assert _a06_tot(a06, "TRANS Masculino") == 1   # T (Transgénero Masculino)
    assert _a06_tot(a06, "TRANS Femenina") == 1    # V (Femenino Trans)
    assert not bool(E.loc[E["run"] == "U", "dem_trans_m"].iloc[0])   # U cis
    assert not bool(E.loc[E["run"] == "U", "dem_trans_f"].iloc[0])


def _main():
    pruebas = [v for k, v in sorted(globals().items())
               if k.startswith("test_") and callable(v)]
    fallos = 0
    for fn in pruebas:
        try:
            fn(); print(f"PASS  {fn.__name__}")
        except AssertionError as e:
            fallos += 1; print(f"FAIL  {fn.__name__} -> {e or 'assert'}")
        except Exception as e:  # noqa: BLE001
            import traceback
            fallos += 1; print(f"ERROR {fn.__name__} -> {type(e).__name__}: {e}")
            traceback.print_exc()
    print("-" * 50)
    print(f"{len(pruebas) - fallos}/{len(pruebas)} OK" + (f" ({fallos} problemas)" if fallos else ""))
    return 1 if fallos else 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    sys.exit(_main())
