#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# ==========================================================================
"""
Pruebas del módulo de screening A03 D.3 (PSC/PSC-Y/GHQ-12). Datos SINTÉTICOS.

Correr:  python tests/test_screening.py   (o  pytest tests/)
Cubre: clasificadores, detección de instrumento por contenido, ambos formatos
(IRIS/Admin), los dos resultados (automático vs DISAM) + discrepancia, estamento,
edad en texto, y el rechazo de un export que no es de instrumentos.
"""

import sys
import tempfile
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

import modulos.rem_a03_d3_instrumentos as scr   # noqa: E402

_TMP = Path(tempfile.mkdtemp(prefix="autorem_scr_"))


# -- Fixtures ----------------------------------------------------------
def _iris(fn, formulario, filas):
    """Export IRIS de instrumento: encabezado en fila 1 + filas de datos.
    `filas` = lista de (rut, edad, sexo, funcionario, estamento, estado, puntaje, resultado)."""
    p = _TMP / fn
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["NUMERO TIPO IDENTIFICACION", "AÑO APLICACIÓN FORMULARIO", "SEXO",
               "FUNCIONARIO", "INSTRUMENTO", "FORMULARIO", "1.- ESTADO",
               "14.- PUNTAJE", "15.- RESULTADO"])
    for rut, edad, sexo, func, estam, estado, punt, resu in filas:
        ws.append([rut, edad, sexo, func, estam, formulario, estado, punt, resu])
    wb.save(p)
    return p


def _admin(fn, formulario, filas):
    """Export Administrativo: banner + encabezado fila 9. Sin columna estamento.
    `filas` = lista de (rut, edad_txt, sexo, funcionario, estado, puntaje, resultado)."""
    p = _TMP / fn
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Servicio de Salud", "S.S. Metropolitano Central"])
    ws.append(["Comuna", "Maipu"]); ws.append(["Establecimiento", "[CESFAM] Dr. Luis Ferrada"])
    ws.append(["Año", "2026"]); ws.append(["Mes", "ENERO - JULIO"])
    ws.append([None, None, "Reporte Formularios RAYEN"])
    ws.append([None, None, f"Formulario: {formulario}"])
    ws.append([])
    ws.append(["RUT", "Numero de Fichas", "Paciente", "Edad de registro formulario",
               "Sexo", "Funcionario", "1.- Estado", "40.- Puntaje", "41.- Resultado"])
    for rut, edad, sexo, func, estado, punt, resu in filas:
        ws.append([rut, "F", "", edad, sexo, func, estado, punt, resu])
    wb.save(p)
    return p


def _dump(path):
    # Lectura por índice de columna (robusto: las celdas vacías vuelven como None
    # y no desalinean la fila como sí puede pasar con iter_rows).
    ws = openpyxl.load_workbook(path)[scr.NOMBRE_HOJA_SALIDA]
    ncol = ws.max_column
    head = [ws.cell(row=1, column=c).value for c in range(1, ncol + 1)]
    rows = [{head[c - 1]: ws.cell(row=r, column=c).value for c in range(1, ncol + 1)}
            for r in range(2, ws.max_row + 1)]
    return head, rows


def _quiet(*_a, **_k):
    pass


# -- Pruebas -----------------------------------------------------------
def test_clasificadores():
    assert scr.clasificar_psc(None) is None          # sin puntaje -> None
    assert scr.clasificar_psc(0) == "Sin riesgo"     # <33: bajo el corte (lo más común)
    assert scr.clasificar_psc(32) == "Sin riesgo"
    assert scr.clasificar_psc(33) == "Bajo"
    assert scr.clasificar_psc(63) == "Bajo"
    assert scr.clasificar_psc(64) == "Medio"
    assert scr.clasificar_psc(70) == "Alto"
    assert scr.clasificar_ghq12(4) == "Bajo"         # GHQ no tiene 'sin riesgo': 0-4 ya es Bajo
    assert scr.clasificar_ghq12(6) == "Medio"
    assert scr.clasificar_ghq12(12) == "Alto"
    assert scr.clasificar_ghq12(None) is None
    assert scr.clasificar_ghq12(13) is None          # >12: fuera del scoring binario 0-12
    assert scr.clasificar_ghq12(-1) is None


def test_psc_bajo_el_corte_se_etiqueta():
    """Un PSC <33 (el resultado MÁS común) no se pierde: sale 'Sin riesgo',
    se cuenta en el total y aparece en el desglose por_resultado."""
    p = _iris("psc_bajo.xlsx", "Cuestionario para padres PSC", [
        ("55555555-5", 8, "Hombre", "F", "Médico", "Ingreso", 20, ""),   # RAYEN va BLANCO bajo el corte
    ])
    out = _TMP / "psc_bajo_out.xlsx"
    res = scr.procesar(p, out, log=_quiet)
    assert res["total"] == 1                          # NO se descarta
    assert res["por_resultado"].get("Sin riesgo") == 1
    _, filas = _dump(out)
    assert filas[0]["Resultado_DISAM"] == "Sin riesgo"
    assert filas[0]["Resultado_RAYEN"] in (None, "")  # RAYEN venía en blanco
    assert filas[0]["Discrepancia"] in (None, "")     # blanco + 'Sin riesgo' -> concuerdan


def test_canon_resultado_wording():
    """Canoniza la redacción REAL de RAYEN (Goldberg = frases; PSC/PSC-Y = bandas)."""
    c = scr.canon_resultado
    assert c("Ausencia de psicopatología") == "Bajo"
    assert c("Sospecha de psicopatología subumbral") == "Medio"
    assert c("Indicativos de presencia de psicopatología") == "Alto"
    assert c("Alto") == "Alto"
    assert c("Bajo") == "Bajo"
    assert c("Medio") == "Medio"
    assert c("") is None
    assert c(None) is None


def test_iris_goldberg_dos_resultados():
    # GHQ-12 REAL: RAYEN usa frases clínicas, NO 'Bajo/Medio/Alto'. Hay que
    # canonizar antes de comparar; si no, TODO Goldberg saldría discrepante (falso).
    p = _iris("gold.xlsx", "Cuestionario de Salud de Goldberg", [
        ("11111111-1", 30, "Hombre", "Func Uno", "Psicólogo(a)", "Ingreso", 8,
         "Indicativos de presencia de psicopatología"),   # 8->Alto ; canon Alto -> coinciden
        ("22222222-2", 40, "Mujer",  "Func Dos", "Médico",       "Egreso",  8,
         "Ausencia de psicopatología"),                   # RAYEN 'Ausencia'(Bajo) pero 8->Alto: DISCREPA
    ])
    out = _TMP / "gold_out.xlsx"
    res = scr.procesar(p, out, log=_quiet)
    assert res["instrumento"] == "GHQ-12" and res["formato"] == "iris" and res["total"] == 2
    _, filas = _dump(out)
    by = {f["RUT"]: f for f in filas}
    a = by["11111111-1"]
    assert a["Momento"] == "Ingreso" and a["Puntaje"] == 8
    assert a["Resultado_RAYEN"] == "Indicativos de presencia de psicopatología"
    assert a["Banda_RAYEN"] == "Alto" and a["Resultado_DISAM"] == "Alto"
    assert a["Discrepancia"] in (None, "")    # coinciden PESE a la redacción distinta
    assert a["Estamento"] == "Psicólogo(a)"
    b = by["22222222-2"]
    assert b["Banda_RAYEN"] == "Bajo" and b["Resultado_DISAM"] == "Alto"   # 8 -> Alto
    assert b["Discrepancia"] == "SI"          # banda RAYEN != banda DISAM
    assert res["discrepancias"] == 1


def test_admin_psc_edad_texto_sin_estamento():
    p = _admin("psc.xlsx", "Cuestionario para padres PSC", [
        ("33333333-3", "12 años 3 meses", "Mujer", "Func Tres", "Ingreso", 65, "Medio"),
    ])
    out = _TMP / "psc_out.xlsx"
    res = scr.procesar(p, out, log=_quiet)
    assert res["instrumento"] == "PSC"
    assert res["formato"] == "administrativo"
    _, filas = _dump(out)
    f = filas[0]
    assert f["Edad"] == 12                     # '12 años 3 meses' -> 12
    assert f["Resultado_DISAM"] == "Medio"     # PSC 65 -> 64-69
    assert f["Estamento"] in (None, "")        # admin no trae estamento
    assert f["Momento"] == "Ingreso"


def test_admin_estamento_desde_lookup():
    """Administrativo: sin columna estamento, se rellena por NOMBRE desde el lookup
    (reporte 'Utilización de Cupos'). Nombre no encontrado -> queda vacío."""
    from programas.rem_utils import norm
    p = _admin("psc_est.xlsx", "Cuestionario para padres PSC", [
        ("33333333-3", "12 años", "Mujer", "Catalina Andrea Mayorga Pino", "Ingreso", 65, "Medio"),
        ("44444444-4", "13 años", "Hombre", "Fulano No Listado",            "Ingreso", 68, "Medio"),
    ])
    out = _TMP / "psc_est_out.xlsx"
    tabla = {norm("Catalina Andrea Mayorga Pino"): "Psicólogo(a)"}
    res = scr.procesar(p, out, estamentos=tabla, log=_quiet)
    assert res["total"] == 2
    _, filas = _dump(out)
    by = {f["RUT"]: f for f in filas}
    assert by["33333333-3"]["Estamento"] == "Psicólogo(a)"   # rellenado por nombre
    assert by["44444444-4"]["Estamento"] in (None, "")       # sin match -> vacío


def test_admin_estamento_resolver_manual():
    """Failsafe: funcionario sin match se resuelve por callback (o se IGNORA)."""
    p = _admin("psc_res.xlsx", "Cuestionario para padres PSC", [
        ("33333333-3", "12 años", "Mujer",  "Ana Interna",  "Ingreso", 65, "Medio"),
        ("44444444-4", "13 años", "Hombre", "Beto Externo", "Ingreso", 40, ""),
    ])
    out = _TMP / "psc_res_out.xlsx"
    vistos = {}                                   # tabla vacía: ambos son "faltantes"
    def resolver(faltantes, opciones):
        assert "Médico" in opciones               # el selector ofrece los estamentos
        return {n: ("Médico" if "ANA" in n.upper() else None) for n in faltantes}
    res = scr.procesar(p, out, estamentos=vistos, resolver_estamento=resolver, log=_quiet)
    _, filas = _dump(out)
    by = {f["RUT"]: f for f in filas}
    assert by["33333333-3"]["Estamento"] == "Médico"     # resuelto a mano
    assert by["44444444-4"]["Estamento"] in (None, "")   # ignorado (externo) -> vacío
    assert res["estam_rellenados"] == 1


def test_deteccion_psc_y():
    p = _iris("pscy.xlsx", "Cuestionario para Adolescentes (PSC-Y) 10 a 14 años", [
        ("44444444-4", 12, "Hombre", "F", "Terapeuta Ocupacional", "Ingreso", 40, "Bajo"),
    ])
    out = _TMP / "pscy_out.xlsx"
    res = scr.procesar(p, out, log=_quiet)
    assert res["instrumento"] == "PSC-Y"       # no confundir con PSC


def test_rechaza_no_instrumento():
    """Un export sin PUNTAJE/RESULTADO (ej. el de Control de Salud Mental) se rechaza."""
    p = _TMP / "noinst.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["NUMERO TIPO IDENTIFICACION", "AÑO APLICACIÓN FORMULARIO", "SEXO", "18.- ESTADO"])
    ws.append(["1-9", 30, "Mujer", "EGRESO ALTA"])
    wb.save(p)
    try:
        scr.abrir_validado(p)
        assert False, "debió rechazar"
    except scr.ArchivoInvalido as e:
        assert e.categoria == "no_instrumento"


def test_tabla_d3_excluye_sin_riesgo():
    """La tabla A03·D.3 = 6 filas (ingreso/egreso × Bajo/Medio/Alto). 'Sin riesgo'
    (bajo el corte) NO cuenta en la tabla, aunque sí esté en el detalle."""
    import pandas as pd
    det = pd.DataFrame([
        dict(momento="Ingreso", resu_disam="Bajo", edad=7, sexo="Hombre"),
        dict(momento="Ingreso", resu_disam="Alto", edad=8, sexo="Mujer"),
        dict(momento="Ingreso", resu_disam="Sin riesgo", edad=6, sexo="Hombre"),  # fuera D.3
        dict(momento="Egreso", resu_disam="Medio", edad=40, sexo="Mujer"),
    ])
    t = scr._tabla_d3(det)
    assert len(t) == 6, f"esperaba 6 filas, hubo {len(t)}"

    def cel(mom, niv, col):
        f = t[t["Evaluación"].str.contains(mom) & (t["Resultado"] == niv)].iloc[0]
        return int(f[col])
    assert cel("ingreso", "Bajo", "Ambos") == 1 and cel("ingreso", "Bajo", "5-9 H") == 1
    assert cel("ingreso", "Alto", "5-9 M") == 1
    assert cel("egreso", "Medio", "40-44 M") == 1     # edad 40 -> banda 40-44
    assert int(t["Ambos"].sum()) == 3, "'Sin riesgo' no debe contar en el D.3"


def test_ghq12_fuera_de_rango_avisa():
    """Un puntaje GHQ-12 > 12 (imposible en el scoring binario 0-12) NO se descarta
    callado: se AVISA (posible cambio a scoring Likert 0-36) y no clasifica. El de
    dentro de rango sí clasifica y no gatilla el aviso (SIMP-2)."""
    p = _iris("ghq_fuera.xlsx", "Cuestionario de Salud de Goldberg", [
        ("11111111-1", 30, "Hombre", "F", "Psicólogo(a)", "Ingreso", 15, ""),   # 15 > 12: fuera
        ("22222222-2", 40, "Mujer",  "F", "Médico",       "Ingreso", 8,  ""),   # 8: Alto, normal
    ])
    out = _TMP / "ghq_fuera_out.xlsx"
    msgs = []
    scr.procesar(p, out, log=msgs.append)
    aviso = [m for m in msgs if "FUERA del rango" in m]
    assert aviso, "debió avisar el puntaje fuera de rango"
    assert "15" in aviso[0]                             # menciona el valor anómalo
    _, filas = _dump(out)
    by = {f["RUT"]: f for f in filas}
    assert by["11111111-1"]["Resultado_DISAM"] in (None, "")   # 15 no clasifica
    assert by["22222222-2"]["Resultado_DISAM"] == "Alto"       # 8 -> Alto


def test_ghq12_en_rango_no_avisa():
    """Un GHQ-12 dentro de 0-12 no gatilla el aviso de fuera de rango."""
    p = _iris("ghq_ok.xlsx", "Cuestionario de Salud de Goldberg", [
        ("33333333-3", 30, "Hombre", "F", "Médico", "Ingreso", 10, ""),
    ])
    out = _TMP / "ghq_ok_out.xlsx"
    msgs = []
    scr.procesar(p, out, log=msgs.append)
    assert not [m for m in msgs if "FUERA del rango" in m]


def test_procesar_unificado():
    """3 instrumentos -> 1 tabla D.3 + detalle auditable al final (con los 'Sin riesgo')."""
    ppsc = _iris("u_psc.xlsx", "Cuestionario para Padres PSC", [
        ("1-1", 7, "Hombre", "", "", "Ingreso", 40, "Bajo"),   # Bajo
        ("2-2", 8, "Mujer",  "", "", "Ingreso", 72, "Alto"),   # Alto
        ("3-3", 6, "Hombre", "", "", "Ingreso", 20, ""),       # <33 = Sin riesgo (fuera D.3)
    ])
    pghq = _iris("u_ghq.xlsx", "Cuestionario de Salud de Goldberg", [
        ("4-4", 40, "Mujer", "", "", "Egreso", 5, ""),         # GHQ 5 = Medio
    ])
    salida = _TMP / "u_out.xlsx"
    res = scr.procesar_unificado({"PSC": ppsc, "GHQ-12": pghq}, salida, log=_quiet)
    assert res["total"] == 4
    assert int(res["tabla"]["Ambos"].sum()) == 3        # Sin riesgo fuera de la tabla
    wb = openpyxl.load_workbook(salida)
    assert "A03_D3" in wb.sheetnames
    assert wb.sheetnames[-1] == "A03_D3_Detalle"        # auditable, al final
    assert wb["A03_D3_Detalle"].max_row - 1 == 4        # detalle incluye el Sin riesgo
    wb.close()


# -- Runner propio -----------------------------------------------------
def _main():
    pruebas = [v for k, v in sorted(globals().items())
               if k.startswith("test_") and callable(v)]
    fallos = 0
    for fn in pruebas:
        try:
            fn(); print(f"PASS  {fn.__name__}")
        except AssertionError as e:
            fallos += 1; print(f"FAIL  {fn.__name__} -> {e or 'assert'}")
        except Exception as e:  # noqa: BLE001
            fallos += 1; print(f"ERROR {fn.__name__} -> {type(e).__name__}: {e}")
    print("-" * 50)
    print(f"{len(pruebas) - fallos}/{len(pruebas)} OK" + (f" ({fallos} con problemas)" if fallos else ""))
    return 1 if fallos else 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    sys.exit(_main())
