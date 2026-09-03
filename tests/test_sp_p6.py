#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 5 (Anthropic).
# The human author reviewed, modified, and integrated the code.
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# SPDX-License-Identifier: GPL-3.0-or-later
# ==========================================================================
"""Pruebas de REM SP·P6 A.1 «Población en control PSM»: programas/poblacion.py
(port del DAX de 'Ferrada') + modulos/rem_sp_p6_poblacion.py (grilla P6·A.1).
Datos SINTÉTICOS. Ver docs/SP_P6_poblacion_plan.md y SP_P6_config_por_dx.md.
Correr:
    python tests/test_sp_p6.py"""

import sys
import tempfile
from datetime import date
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

import programas.poblacion as pob              # noqa: E402
import modulos.rem_sp_p6_poblacion as p6mod     # noqa: E402
from programas.rem_utils import ArchivoInvalido, dv_rut  # noqa: E402

_TMP = Path(tempfile.mkdtemp(prefix="autorem_sp_p6_"))


def _quiet(*_a, **_k):
    pass


# ======================================================================
# Fixtures — mismo layout que descarga RAYEN/IRIS (banner + encabezado real)
# ======================================================================
_Q = {
    1: "1.- ¿USTED ES MADRE DE HIJO MENOR DE 5 AÑOS?",
    4: "4.- ¿ ES VÍCTIMA O AGRESOR/A DE VIOLENCIA?", 5: "5.- ESTADO",
    6: "6.- TIPO DE VIOLENCIA", 7: "7.- EN LA VIOLENCIA ES",
    9: "9.- ¿ SUFRE DE ABUSO SEXUAL ?", 10: "10.- ESTADO",
    11: "11.- ¿EPISODIO DE SUICIDIO?", 12: "12.- TIPO DE SUICIDIO", 13: "13.- ESTADO",
    18: "18.- ¿ TIENE  DEPRESIÓN ?", 19: "19.- ESTADO", 20: "20.- TIPO DE DEPRESIÓN",
    21: "21.- ¿ TIENE DEPRESIÓN POST - PARTO ?", 22: "22.- ESTADO",
    23: "23.- ¿ TIENE TRANSTORNO BIPOLAR ?", 24: "24.- ESTADO",
    31: "31.- CONSUMO PERJUDICIAL DE ALCOHOL", 32: "32.- ESTADO",
    33: "33.- CONSUMO DEPENDIENTE DEL ALCOHOL", 34: "34.- ESTADO",
    35: "35.- CONSUMO PERJUDICIAL DE DROGAS", 36: "36.- ESTADO",
    37: "37.- CONSUMO DEPENDIENTE DE DROGAS", 38: "38.- ESTADO",
    39: "39.- CONSUMO DE DROGAS Y ALCOHOL", 40: "40.- ESTADO",
    41: "41.- ¿ TIENE TRASTORNO DE ANSIEDAD ?", 42: "42.- ESTADO",
    43: "43.- TIPO DE TRASTORNO DE ANSIEDAD",
    44: "44.- ¿ TIENE ALZHEIMER Y/O OTRAS DEMENCIAS ?", 45: "45.- ETAPA", 46: "46.- ESTADO",
    49: "49.- ¿TIENE TRASTORNO ADAPTATIVO?", 50: "50.- ESTADO",
    51: "51.- ¿ TIENE ESQUIZOFRENIA ?", 52: "52.- ESTADO",
    55: "55.- ¿ TIENE TRASTORNO DE LA CONDUCTA ALIMENTARIA ?", 56: "56.- ESTADO",
    57: "57.- ¿ TIENE TRASTORNOS HIPERCINÉTICOS, DE LA ACTIVIDAD", 58: "58.- ESTADO",
    59: "59.- ¿ TIENE RETRASO MENTAL ?", 60: "60.- ESTADO",
    61: "61.- ¿ TIENE TRASTORNO DE PERSONALIDAD ?", 62: "62.- ESTADO",
    63: "63.- ¿ TIENE TRASTORNO GENERALIZADO DEL DESARROLLO ?", 64: "64.- ESTADO",
    65: "65.- OTRAS (TRASTORNOS NO INCLUIDOS EN SECCIÓN)", 66: "66.- ESTADO",
    69: "69.- ¿TIENE TRASTORNO DISOCIAL DESAFIANTE Y OPOSICIONIS", 70: "70.- ESTADO",
    71: "71.- ¿TIENE TRASTORNO DE ANSIEDAD DE SEPARACIÓN EN LA I", 72: "72.- ESTADO",
    73: "73.- ¿TIENE OTROS TRASTORNOS DEL COMPORTAMIENTO Y DE LA", 74: "74.- ESTADO",
    83: "83.- ¿TIENE AUTISMO?", 84: "84.- ESTADO",
    85: "85.- ¿TIENE ASPERGER?", 86: "86.- ESTADO",
    87: "87.- ¿TIENE SÍNDROME DE RETT?", 88: "88.- ESTADO",
    89: "89.- ¿TIENE TRASTORNO DESINTEGRATIVO DE LA INFANCIA?", 90: "90.- ESTADO",
    91: "91.- ¿TIENE TRASTORNO GENERALIZADO DEL DESARROLLO DE LA", 92: "92.- ESTADO",
}
_FORM_HDR = [
    "SERVICIO SALUD", "ESTABLECIMIENTO", "TIPO IDENTIFICACION", "NUMERO TIPO IDENTIFICACION",
    "CODIGO FAMILIA", "NUMERO DE FICHA RAYEN", "NUMERO DE FICHA CODIGO ANTIGUO", "PACIENTE",
    "FECHA DE NACIMIENTO", "EDAD PACIENTE", "AÑO APLICACIÓN FORMULARIO", "MES APLICACIÓN FORMULARIO",
    "DÍAS APLICACIÓN FORMULARIO", "PUEBLO ORIGINARIO", "ALERTAS ADMINISTRATIVAS", "NACIONALIDAD",
    "SEXO", "GENERO", "SECTOR INSCRIPCION", "SECTOR CITA", "DIRECCIÓN", "COMUNA", "TELEFONO 1",
    "TELEFONO 2", "PREVISION", "CONVENIO", "SITUACION", "ESTADO", "FUNCIONARIO PASIVADOR",
    "ATEN ID", "FECHA ATENCION", "FECHA FORMULARIO", "FUNCIONARIO", "INSTRUMENTO",
    "ESTABLECIMIENTO INSCRIPCION", "FORMULARIO", "FUNCIONARIOS FORMULARIO",
] + [_Q[n] for n in sorted(_Q)]


def _mk_formulario(filas):
    """Histórico 'Control de Salud Mental' (IRIS): banner de 16 filas + encabezado
    real en la 17 (mismo layout que consume rem_saludmental/A05). `filas` = lista de
    dicts con 'rut'/'fecha'/'instr' + respuestas (clave = texto de pregunta, _Q[n])."""
    p = _TMP / "formulario.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 17):
        ws.cell(row=r, column=1, value=f"banner {r}")
    ws.append(_FORM_HDR)
    # nunca 0 filas de datos: igual que el ADA, un histórico vacío hace que pandas
    # infiera columnas float64 y .str.contains() reviente más abajo (caso límite que
    # no ocurre en la práctica: el histórico siempre trae algo).
    filas = filas or [{"rut": "00000000-0", "fecha": date(2020, 1, 1)}]
    for f in filas:
        base = {"NUMERO TIPO IDENTIFICACION": f["rut"], "FECHA FORMULARIO": f["fecha"],
                "INSTRUMENTO": f.get("instr", "Medico"), "SEXO": "Mujer",
                "AÑO APLICACIÓN FORMULARIO": 30}
        vals = [f.get(h, base.get(h)) for h in _FORM_HDR]
        ws.append(vals)
    wb.save(p)
    return p


_INS_HDR = ["TIPO IDENTIFICACION", "NUMERO TIPO IDENTIFICACION", "SEXO", "GENERO",
           "FECHA DE NACIMIENTO", "EDAD AÑOS", "NACIONALIDAD", "PUEBLO INDIG",
           "ALERTAS ADMINISTRATIVAS", "SITUACION", "ESTADO", "SECTOR",
           "FECHA PASIVACION", "MOTIVO PASIVACION"]


def _mk_inscritos(filas):
    """'Informe Inscritos y Adscritos' (IRIS). `filas` = lista de dicts:
    rut/sexo/fnac/tipoid/nacionalidad/pueblo/alertas/sector/estado/situacion."""
    p = _TMP / "inscritos.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_INS_HDR)
    for f in filas:
        base = {"TIPO IDENTIFICACION": f.get("tipoid", "RUN"),
               "NUMERO TIPO IDENTIFICACION": f["rut"], "SEXO": f.get("sexo", "Mujer"),
               "GENERO": f.get("genero", "Femenina" if f.get("sexo", "Mujer") == "Mujer" else "Masculino"),
               "FECHA DE NACIMIENTO": f.get("fnac"), "EDAD AÑOS": f.get("edad", 30),
               "NACIONALIDAD": f.get("nacionalidad", "Chilena"), "PUEBLO INDIG": f.get("pueblo", "Ninguno"),
               "ALERTAS ADMINISTRATIVAS": f.get("alertas", ""), "SITUACION": f.get("situacion", "Inscrito"),
               "ESTADO": f.get("estado", "Activo"), "SECTOR": f.get("sector", "Norte"),
               "FECHA PASIVACION": f.get("fpasiv"), "MOTIVO PASIVACION": f.get("mpasiv", "")}
        ws.append([base.get(h, "") for h in _INS_HDR])
    wb.save(p)
    return p


_ADA_HDR = ["NUMERO TIPO IDENTIFICACION", "ATEN ID", "FECHA ATENCION", "ACTIVIDADES",
           "DIAGNOSTICOS", "INSTRUMENTO", "TIPO ATENCION", "SEXO", "SECTOR"]


def _mk_ada(filas):
    """ADA. `filas` = lista de dicts: rut/fecha/act/instr/id/diag/tipo/sexo/sector.
    Nunca queda VACÍO del todo (0 filas de datos): con solo el encabezado, pandas
    infiere columnas float64 y `.str.contains()` revienta más abajo — no es una
    situación real (el ADA siempre trae algo), así que se rellena con una fila
    irrelevante en vez de reproducir ese caso límite acá."""
    p = _TMP / "ada.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_ADA_HDR)
    filas = filas or [{"rut": "00000000-0", "fecha": date(2020, 1, 1), "act": "Consulta general"}]
    for f in filas:
        ws.append([f["rut"], f.get("id", "A"), f["fecha"].strftime("%d/%m/%Y"), f.get("act", ""),
                  f.get("diag", ""), f.get("instr", "Medico"), f.get("tipo", "Consulta"),
                  f.get("sexo", "Mujer"), f.get("sector", "Norte")])
    wb.save(p)
    return p


def _sm(rut, fecha=date(2026, 8, 5)):
    """Atención SM (una de las 7 actividades validadas) -> ¿Activo 12m?=SI."""
    return {"rut": rut, "fecha": fecha, "act": "Controles Salud Mental"}


def _poblacion(formulario_filas, inscritos_filas, ada_filas=None, mes=(2026, 8)):
    return pob.construir_poblacion(
        str(_mk_inscritos(inscritos_filas)), str(_mk_formulario(formulario_filas)),
        str(_mk_ada(ada_filas or [])), mes=mes, log=_quiet)


def _p6(P):
    return p6mod.construir_p6(P, log=_quiet)


def _fila_p6(grid, n):
    return grid[grid["Fila"] == n].iloc[0]


def _motivos(resultado):
    """Motivos de AMBAS hojas de revisión juntos (§5.5: Revisar_Administrativo +
    Revisar_Clinico) — a la mayoría de los tests no les importa en cuál cayó,
    solo que quedó trazado. Los que sí distinguen usan las claves directo."""
    partes = [resultado["revisar_administrativo"], resultado["revisar_clinico"]]
    return [m for r in partes if len(r) for m in r["Motivo"]]


def _poblacion_log(formulario_filas, inscritos_filas, ada_filas, mes=(2026, 8)):
    """Como `_poblacion`, pero devuelve (P, lineas_de_log) para probar avisos."""
    lineas = []
    P = pob.construir_poblacion(
        str(_mk_inscritos(inscritos_filas)), str(_mk_formulario(formulario_filas)),
        str(_mk_ada(ada_filas)), mes=mes, log=lineas.append)
    return P, lineas


# ======================================================================
# programas/poblacion.py — tabla «Ferrada»
# ======================================================================
def test_activo_simple_ingreso_sin_egreso():
    P = _poblacion(
        [{"rut": "11111111-1", "fecha": date(2026, 7, 10), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO"}}],
        [{"rut": "11111111-1"}])
    assert P.loc[P["Número"] == "11111111-1", "Depresión (form)"].iloc[0] == "Activo"


def test_egreso_por_diagnostico_no_el_bug_del_powerbi():
    """§4.3: RUN con Depresión Y Ansiedad activas; Depresión egresa en el mes. El port
    egresa SOLO Depresión (Ansiedad sigue Activo); el bug del PowerBI habría egresado
    las dos -> queda en Egreso_Divergencias, con el RUN."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 5, 1), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO"}},
        {"rut": "11111111-1", "fecha": date(2026, 5, 1), **{_Q[41]: "SI", _Q[42]: "INGRESO", _Q[43]: "Generalizada"}},
        {"rut": "11111111-1", "fecha": date(2026, 8, 20), **{_Q[18]: "SI", _Q[19]: "19.- EGRESO"}},
    ], [{"rut": "11111111-1"}])
    fila = P[P["Número"] == "11111111-1"].iloc[0]
    assert fila["Depresión (form)"] == "Egresado"
    assert fila["Ansiedad (form)"] == "Activo"
    div = P.attrs["egreso_divergencias"]
    assert len(div) == 1
    assert div.iloc[0]["RUN"] == "11111111-1" and div.iloc[0]["Diagnostico"] == "Ansiedad (form)"
    assert div.iloc[0]["Valor_port"] == "Activo" and div.iloc[0]["Valor_PowerBI"] == "Egresado"


def test_d2_factor_riesgo_sin_filtro_de_instrumento():
    """Violencia (factor de riesgo) registrada por Trabajador Social (NO médico)
    igual cuenta — al revés que los diagnósticos, que sí exigen INSTRUMENTOcontieneMEDIC."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 8, 1), "instr": "Trabajador Social",
         **{_Q[4]: "SI", _Q[5]: "INGRESO", _Q[6]: "Física", _Q[7]: "Víctima"}},
    ], [{"rut": "11111111-1"}])
    assert P.loc[P["Número"] == "11111111-1", "Violencia (form)"].iloc[0] == "Activo"


def test_diagnostico_exige_instrumento_medico():
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 8, 1), "instr": "Psicólogo(a)",
         **{_Q[18]: "SI", _Q[19]: "19.- INGRESO"}},
    ], [{"rut": "11111111-1"}])
    assert P.loc[P["Número"] == "11111111-1", "Depresión (form)"].iloc[0] == ""


def test_d1_fallback_tgd_pregunta_63():
    """La 91 (TGD no especificado) casi nunca tiene dato -> fallback a la 63, SOLO si
    ninguna TGD específica (83/85/87/89/91) está activa."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 4, 1), **{_Q[63]: "SI", _Q[64]: "INGRESO"}},
        # B tiene Autismo (83) activo Y también la 63 -> NO debe usar el fallback
        {"rut": "22222222-2", "fecha": date(2026, 4, 1), **{_Q[83]: "SI", _Q[84]: "INGRESO"}},
        {"rut": "22222222-2", "fecha": date(2026, 4, 1), **{_Q[63]: "SI", _Q[64]: "INGRESO"}},
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}])
    assert P.loc[P["Número"] == "11111111-1", "TGD (form)"].iloc[0] == "Activo"
    b = P[P["Número"] == "22222222-2"].iloc[0]
    assert b["Autismo (form)"] == "Activo" and b["TGD (form)"] == ""


def test_gestante_ventana_3_meses_matrona():
    P = _poblacion([], [{"rut": "11111111-1"}, {"rut": "22222222-2"}], ada_filas=[
        {"rut": "11111111-1", "fecha": date(2026, 7, 15), "act": "Control Prenatal", "instr": "Matrona"},
        {"rut": "22222222-2", "fecha": date(2026, 4, 1), "act": "Control Prenatal", "instr": "Matrona"},  # fuera de ventana
    ])
    assert P.loc[P["Número"] == "11111111-1", "¿Embarazada?"].iloc[0] == "SI"
    assert P.loc[P["Número"] == "22222222-2", "¿Embarazada?"].iloc[0] == "NO"


def test_activo12m_y_rescate_6m_13m_misma_lista_actividades():
    P = _poblacion([], [{"rut": "11111111-1"}], ada_filas=[_sm("11111111-1", date(2026, 2, 10))])   # mes-6 exacto
    fila = P[P["Número"] == "11111111-1"].iloc[0]
    assert fila["¿Activo 12m?"] == "SI"          # dentro de los 12 meses cerrados
    assert fila["¿Última atención hace 6m?"] == "Si"


def test_ingresado_si_cualquier_dx_activo():
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[57]: "SI", _Q[58]: "INGRESO"}},
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}])
    assert P.loc[P["Número"] == "11111111-1", "¿Ingresado?"].iloc[0] == "SI"
    assert P.loc[P["Número"] == "22222222-2", "¿Ingresado?"].iloc[0] == "NO"


def test_run_responsable_no_corrompe_a_la_madre():
    """§5.5.2, caso OBLIGATORIO del plan: un RN 'RUN Responsable' con el MISMO Número
    que su madre (listado DESPUÉS de ella, para que el viejo drop_duplicates(keep=
    'last') se hubiera quedado con la fila equivocada) no debe pisar la fila real."""
    P = _poblacion([
        {"rut": "99999999-9", "fecha": date(2026, 7, 5), **{_Q[21]: "SI", _Q[22]: "INGRESO"}},
    ], [
        {"rut": "99999999-9", "sexo": "Mujer", "fnac": date(1998, 4, 4)},
        {"rut": "99999999-9", "sexo": "Hombre", "fnac": date(2026, 8, 1), "tipoid": "RUN Responsable"},
    ])
    filas = P[P["Número"] == "99999999-9"]
    assert len(filas) == 1
    assert filas.iloc[0]["Sexo"] == "Mujer"
    assert filas.iloc[0]["Depresión Postparto (form)"] == "Activo"


def test_cobertura_avisa_si_ada_no_llega_al_mes_reportado():
    _, log = _poblacion_log(
        [{"rut": "11111111-1", "fecha": date(2026, 8, 1), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO"}}],
        [{"rut": "11111111-1"}],
        [{"rut": "11111111-1", "fecha": date(2025, 1, 5), "act": "Controles Salud Mental"},   # cubre 13m
         {"rut": "11111111-1", "fecha": date(2026, 6, 5), "act": "Controles Salud Mental"}],  # pero NO llega a agosto
        mes=(2026, 8))
    assert any("El ADA NO llega hasta el mes reportado" in m for m in log)
    assert not any("necesitan desde" in m and "13 meses cerrados" in m for m in log)   # sí cubre 13 meses


def test_identificador_no_rut_solo_se_reporta_si_habria_entrado_al_p6():
    """El check de identificador se calcula sobre el padron completo (guardarrail y
    unicidad lo necesitan), pero SOLO se REPORTA a quien habria entrado al P6. Un
    pasaporte sin nada de salud mental es ruido en la hoja, no un caso a revisar."""
    # padron de relleno con identificador valido: sin el, 2 pasaportes sobre 3
    # personas serian el 67% y saltaria el guardarrail del 5% (§5.5.1).
    relleno = [{"rut": f"1000{n:04d}-{dv_rut(f'1000{n:04d}')}"} for n in range(50)]
    P = _poblacion([
        # AB123456: pasaporte CON diagnostico activo -> habria entrado al P6 -> se reporta
        {"rut": "AB123456", "fecha": date(2026, 7, 1),
         **{_Q[18]: "SI", _Q[19]: "19.- INGRESO", _Q[20]: "Leve"}},
        {"rut": "11111111-1", "fecha": date(2026, 7, 1),
         **{_Q[18]: "SI", _Q[19]: "19.- INGRESO", _Q[20]: "Leve"}},
    ], [
        {"rut": "11111111-1"},
        {"rut": "AB123456", "tipoid": "Numero de Pasaporte"},
        # CD789012: pasaporte SIN nada de salud mental -> NO se reporta (puro ruido)
        {"rut": "CD789012", "tipoid": "Numero de Pasaporte"},
    ] + relleno, ada_filas=[_sm("11111111-1"), _sm("AB123456")])
    r = _p6(P)
    ruts = list(r["revisar_administrativo"]["RUN"])
    assert "AB123456" in ruts        # habria entrado al P6: accionable
    assert "CD789012" not in ruts    # nunca iba a tributar: se descarta callado


def test_cobertura_avisa_si_ada_no_cubre_13_meses():
    """El caso del autor: formulario con histórico largo (llega hasta el mes) pero
    ADA que arranca reciente -> no cubre los 13 meses que exigen Activo12m/rescate/
    gestante, aunque el ADA SÍ llegue hasta el mes reportado."""
    _, log = _poblacion_log(
        [{"rut": "11111111-1", "fecha": date(2022, 3, 1), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO"}},
         {"rut": "11111111-1", "fecha": date(2026, 8, 1), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO"}}],
        [{"rut": "11111111-1"}],
        [{"rut": "11111111-1", "fecha": date(2026, 7, 1), "act": "Controles Salud Mental"},
         {"rut": "11111111-1", "fecha": date(2026, 8, 5), "act": "Controles Salud Mental"}],   # arranca en julio
        mes=(2026, 8))
    assert any("necesitan desde" in m and "13 meses cerrados" in m for m in log)
    assert not any("El ADA NO llega hasta el mes reportado" in m for m in log)
    assert not any("El histórico del formulario SM NO llega" in m for m in log)


# ======================================================================
# modulos/rem_sp_p6_poblacion.py — grilla P6·A.1
# ======================================================================
def test_base_exige_activo_e_ingresado_y_activo12m():
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[57]: "SI", _Q[58]: "INGRESO"}},   # TDAH, con ADA
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[57]: "SI", _Q[58]: "INGRESO"}},   # TDAH, SIN ADA -> Activo12m NO
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}], ada_filas=[_sm("11111111-1")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 35)["Ambos"] == 1   # solo A


def test_guardarrail_techo_5_por_ciento():
    """Si el filtro de identificador descarta más del 5% del padrón, es el filtro el
    que está roto -> ArchivoInvalido explícito, nunca un P6_A1 en cero silencioso."""
    filas_ins = [{"rut": f"{i:07d}-{i % 10}"} for i in range(9)]
    filas_ins.append({"rut": "AB1234567", "tipoid": "Pasaporte"})   # 1/10 = 10% > techo
    P = _poblacion([], filas_ins)
    try:
        p6mod.construir_p6(P, log=_quiet)
        assert False, "debió levantar ArchivoInvalido"
    except ArchivoInvalido as e:
        assert e.categoria == "validador_rut"


def test_violencia_bucket_tipo_y_victima_agresor_mas_abuso_sexual():
    """Violencia/Suicidio son factores de riesgo (filas 15-23): se cuentan
    INDEPENDIENTES de los diagnósticos (25-58), nunca se mezclan con ellos. Se
    prueban explícitamente los DOS lados de 'Víctima o Agresor/a' (no solo inferirlo
    del default de abuso sexual): física-agresor (16), psicológica-víctima (19), y
    abuso sexual SIN sub-pregunta propia -> se reencamina a violencia sexual, víctima
    por defecto (17) — nunca a la fila 21 (que queda fija en 0, no capturable)."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1),
         **{_Q[4]: "SI", _Q[5]: "INGRESO", _Q[6]: "Física", _Q[7]: "Agresor/a"}},
        {"rut": "22222222-2", "fecha": date(2026, 7, 1),
         **{_Q[4]: "SI", _Q[5]: "INGRESO", _Q[6]: "Psicológica", _Q[7]: "Víctima"}},
        {"rut": "33333333-3", "fecha": date(2026, 7, 1), **{_Q[9]: "SI", _Q[10]: "INGRESO"}},   # abuso sexual
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}, {"rut": "33333333-3"}],
       ada_filas=[_sm("11111111-1"), _sm("22222222-2"), _sm("33333333-3")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 16)["Ambos"] == 1     # violencia física, agresor/a
    assert _fila_p6(r["grid"], 15)["Ambos"] == 0     # NO también en víctima
    assert _fila_p6(r["grid"], 19)["Ambos"] == 1     # violencia psicológica, víctima
    assert _fila_p6(r["grid"], 20)["Ambos"] == 0     # NO también en agresor/a
    assert _fila_p6(r["grid"], 17)["Ambos"] == 1     # abuso sexual -> violencia sexual, víctima por defecto
    assert _fila_p6(r["grid"], 18)["Ambos"] == 0     # nunca agresor/a por abuso sexual
    assert _fila_p6(r["grid"], 21)["Ambos"] == 0     # fila 21 (abuso sexual) siempre 0, no capturable
    # independientes de los diagnósticos: nadie de este fixture tiene dx 25-58 activo
    assert _fila_p6(r["grid"], 24)["Ambos"] == 0


def test_factor_de_riesgo_sin_diagnostico_se_reporta():
    """Un FR (violencia/suicidio, 15-23) sin ningún dx de trastorno mental (25-58) es
    registro incompleto -> se REPORTA en Revisar_Clinico, sin cambiar ningún número.
    El que además tiene dx NO se reporta."""
    P = _poblacion([
        # 11: solo factor de riesgo (violencia), sin diagnóstico -> se reporta
        {"rut": "11111111-1", "fecha": date(2026, 7, 1),
         **{_Q[4]: "SI", _Q[5]: "INGRESO", _Q[6]: "Física", _Q[7]: "Víctima"}},
        # 22: factor de riesgo + diagnóstico (depresión) -> NO se reporta
        {"rut": "22222222-2", "fecha": date(2026, 7, 1),
         **{_Q[11]: "SI", _Q[13]: "INGRESO", _Q[12]: "Intento"}},
        {"rut": "22222222-2", "fecha": date(2026, 7, 1),
         **{_Q[18]: "SI", _Q[19]: "19.- INGRESO", _Q[20]: "Leve"}},
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}],
       ada_filas=[_sm("11111111-1"), _sm("22222222-2")])
    r = _p6(P)
    rc = r["revisar_clinico"]
    fr = rc[rc["Motivo"] == "Factor de riesgo SIN diagnóstico"]
    assert list(fr["RUN"]) == ["11111111-1"]          # solo el que no tiene dx
    assert "22222222-2" not in list(fr["RUN"])         # el que tiene dx no se reporta
    assert _fila_p6(r["grid"], 15)["Ambos"] == 1       # el FR igual cuenta (no cambia números)
    assert _fila_p6(r["grid"], 24)["Ambos"] == 1       # solo 22 tiene dx


def test_suicidio_bucket_ideacion_intento():
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[11]: "SI", _Q[13]: "INGRESO", _Q[12]: "Ideación"}},
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[11]: "SI", _Q[13]: "INGRESO", _Q[12]: "Intento"}},
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}], ada_filas=[_sm("11111111-1"), _sm("22222222-2")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 22)["Ambos"] == 1
    assert _fila_p6(r["grid"], 23)["Ambos"] == 1


def test_depresion_bucket_y_d5_subtipo_vacio_no_tributa():
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO", _Q[20]: "Depresión Grave"}},
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO"}},  # sin subtipo -> D5
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}], ada_filas=[_sm("11111111-1"), _sm("22222222-2")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 27)["Ambos"] == 1        # solo A (grave)
    assert "Dx activo sin subtipo registrado/reconocible" in _motivos(r)


def test_ansiedad_reusa_override_subtipo_panico_no_fobia():
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1),
         **{_Q[41]: "SI", _Q[42]: "INGRESO", _Q[43]: "Pánico sin agorafobia"}},
    ], [{"rut": "11111111-1"}], ada_filas=[_sm("11111111-1")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 40)["Ambos"] == 1   # Pánico
    assert _fila_p6(r["grid"], 41)["Ambos"] == 0   # NO Fobia social


def test_demencia_solo_con_etapa():
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[44]: "SI", _Q[46]: "INGRESO", _Q[45]: "Moderado"}},
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[44]: "SI", _Q[46]: "INGRESO"}},   # sin etapa -> no tributa
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}], ada_filas=[_sm("11111111-1"), _sm("22222222-2")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 45)["Ambos"] == 1
    assert sum(_fila_p6(r["grid"], f)["Ambos"] for f in (44, 45, 46)) == 1


def test_edad_se_pliega_no_se_descarta_fila37_y_fila38():
    """37 Ansiedad separación: 0-14 (>=15 pliega a 10-14). 38 Otros infancia: 0-24 en
    la plantilla REAL (>=25 pliega a 20-24 — corregido; el plan original decía 0-19)."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[71]: "SI", _Q[72]: "INGRESO"}},   # 20 años
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[73]: "SI", _Q[74]: "INGRESO"}},   # 30 años
    ], [
        {"rut": "11111111-1", "fnac": date(2006, 1, 1)},
        {"rut": "22222222-2", "fnac": date(1996, 1, 1)},
    ], ada_filas=[_sm("11111111-1"), _sm("22222222-2")])
    r = _p6(P)
    f37 = _fila_p6(r["grid"], 37)
    assert f37["Ambos"] == 1 and f37["10-14 M"] == 1     # 20 años plegado a 10-14
    f38 = _fila_p6(r["grid"], 38)
    assert f38["Ambos"] == 1 and f38["20-24 M"] == 1     # 30 años plegado a 20-24
    assert any(m.startswith("Edad plegada") for m in _motivos(r))


def test_dato_demografico_bloqueado_va_a_revisar_y_no_se_escribe():
    """TDAH (fila 35): AO (madre<5) está bloqueado en la plantilla -> no se escribe,
    aunque la persona tenga el flag, y queda trazado en P6_Revisar."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[1]: "SI", _Q[57]: "SI", _Q[58]: "INGRESO"}},
    ], [{"rut": "11111111-1", "sexo": "Mujer"}], ada_filas=[_sm("11111111-1")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 35)["AO"] == 0
    assert any(m == "Dato demográfico no aplica en esta fila" for m in _motivos(r))


def test_sin_exclusion_comodin_cajon_de_sastre_cuenta_todo():
    """§5.1 ELIMINADA (sep-2026): las filas cajón de sastre (38/43/48) tributan SIN
    restringir por comorbilidad. El adaptativo con 2 comorbilidades ya NO se excluye,
    y no aparece el motivo 'Excluido por comorbilidad' en ninguna hoja de revisión."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[49]: "SI", _Q[50]: "INGRESO"}},   # adaptativo solo
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[49]: "SI", _Q[50]: "INGRESO"}},
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO", _Q[20]: "Leve"}},
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[23]: "SI", _Q[24]: "INGRESO"}},   # 2 comorbilidades
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}], ada_filas=[_sm("11111111-1"), _sm("22222222-2")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 48)["Ambos"] == 2     # AMBOS cuentan: ya no se filtra
    motivos = list(r["revisar_clinico"]["Motivo"]) + list(r["revisar_administrativo"]["Motivo"])
    assert not any("comorbilidad" in str(m).lower() for m in motivos)


def test_av_pic_regla_wip_ges_y_total_filas():
    """AV = total de la fila en GES/factor de riesgo, 0 en el resto; fila24 = suma de
    AV(25:58); fila13 = suma de AV(15:24)."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1),
         **{_Q[18]: "SI", _Q[19]: "19.- INGRESO", _Q[20]: "Grave"}},   # GES depresión
        {"rut": "22222222-2", "fecha": date(2026, 7, 1), **{_Q[49]: "SI", _Q[50]: "INGRESO"}},  # NO es GES-PIC
    ], [{"rut": "11111111-1"}, {"rut": "22222222-2"}], ada_filas=[_sm("11111111-1"), _sm("22222222-2")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 27)["AV"] == 1     # depresión grave: GES
    assert _fila_p6(r["grid"], 48)["AV"] == 0     # adaptativo: no está en la regla PIC
    assert _fila_p6(r["grid"], 24)["AV"] == 1     # suma AV(25:58)
    assert _fila_p6(r["grid"], 13)["AV"] == 1     # suma AV(15:24)


def test_fila13_suma_literal_fila24_distinct_count():
    """13 = suma literal 15-24 (dobla los factores de riesgo). 24 = personas distintas
    con >=1 dx 25-58 (una persona con 2 dx cuenta 1 sola vez). El doble conteo de FR es
    entre categorías INDEPENDIENTES (Violencia + Suicidio son columnas Ferrada
    separadas) — dentro de 'Violencia Tipo (form)' solo sobrevive el último valor
    enviado (1 fila por RUN), así que física+psicológica en la MISMA persona no se
    pueden dar simultáneamente ahí."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1),
         **{_Q[4]: "SI", _Q[5]: "INGRESO", _Q[6]: "Física", _Q[7]: "Víctima"}},
        {"rut": "11111111-1", "fecha": date(2026, 7, 1),
         **{_Q[11]: "SI", _Q[13]: "INGRESO", _Q[12]: "Ideación"}},
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[57]: "SI", _Q[58]: "INGRESO"}},   # TDAH
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[69]: "SI", _Q[70]: "INGRESO"}},   # disocial
    ], [{"rut": "11111111-1"}], ada_filas=[_sm("11111111-1")])
    r = _p6(P)
    assert _fila_p6(r["grid"], 15)["Ambos"] == 1     # violencia física, víctima
    assert _fila_p6(r["grid"], 22)["Ambos"] == 1     # suicidio, ideación
    assert _fila_p6(r["grid"], 24)["Ambos"] == 1     # cuenta UNA vez (TDAH + disocial)
    assert _fila_p6(r["grid"], 13)["Ambos"] == 3     # 15 + 22 + 24 (doble conteo de FR heredado)


def test_sexo_no_binario_no_se_ubica_y_va_a_revisar():
    """No hay columna H/M donde ubicarlo: 'Hombres'/'Mujeres' quedan en 0 (que es lo
    que de verdad se pega — la plantilla real calcula C=SUM(D:E), no un headcount
    aparte). Nuestro 'Ambos' informativo SÍ sigue contándolo (mismo criterio que el
    resto del proyecto: 'Ambos sexos' = total, no la suma H+M) — no es lo relevante acá."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[57]: "SI", _Q[58]: "INGRESO"}},
    ], [{"rut": "11111111-1", "sexo": "Intersex"}], ada_filas=[_sm("11111111-1")])
    r = _p6(P)
    fila = _fila_p6(r["grid"], 35)
    assert fila["Hombres"] == 0 and fila["Mujeres"] == 0
    # va a Revisar_ADMINISTRATIVO (problema de registro, se corrige en la ficha)
    assert (r["revisar_administrativo"]["Motivo"] == "Sexo/género no binario").any()
    assert "Sexo/género no binario" not in list(r["revisar_clinico"]["Motivo"])


def test_revisar_clinico_siempre_trae_fila13_vs_distinct():
    """'Fila 13 vs distinct' (§5.2) es una MAGNITUD, no un caso por RUN -> siempre
    aparece en Revisar_Clinico (aunque sea con diferencia 0), nunca en la hoja
    administrativa."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 7, 1), **{_Q[57]: "SI", _Q[58]: "INGRESO"}},
    ], [{"rut": "11111111-1"}], ada_filas=[_sm("11111111-1")])
    r = _p6(P)
    assert (r["revisar_clinico"]["Motivo"] == "Fila 13 vs distinct").any()
    assert "Fila 13 vs distinct" not in list(r["revisar_administrativo"]["Motivo"])


def test_egreso_multidx_divergente_llega_a_revisar_clinico():
    """El §4.3 (egreso multi-dx divergente, calculado en programas.poblacion) se
    refleja también en Revisar_Clinico, no solo en la hoja Egreso_Divergencias."""
    P = _poblacion([
        {"rut": "11111111-1", "fecha": date(2026, 5, 1), **{_Q[18]: "SI", _Q[19]: "19.- INGRESO"}},
        {"rut": "11111111-1", "fecha": date(2026, 5, 1), **{_Q[41]: "SI", _Q[42]: "INGRESO", _Q[43]: "Generalizada"}},
        {"rut": "11111111-1", "fecha": date(2026, 8, 20), **{_Q[18]: "SI", _Q[19]: "19.- EGRESO"}},
    ], [{"rut": "11111111-1"}], ada_filas=[_sm("11111111-1")])
    r = _p6(P)
    clin = r["revisar_clinico"]
    fila = clin[clin["Motivo"] == "Egreso multi-dx divergente"]
    assert len(fila) == 1
    assert fila.iloc[0]["RUN"] == "11111111-1"
    assert fila.iloc[0]["Detalle"] == "Ansiedad (form)"


def test_bloques_pegables_fila28_es_bloque_propio():
    bloques = p6mod._bloques_pegables()
    fila28 = bloques[bloques["Filas"] == "28"]
    assert len(fila28) == 1
    assert fila28.iloc[0]["Pegar_desde"] == "J28"    # 10-14 años, mujeres = primera banda abierta


def _main():
    pruebas = [v for k, v in sorted(globals().items())
              if k.startswith("test_") and callable(v)]
    fallos = 0
    for fn in pruebas:
        try:
            fn(); print(f"PASS  {fn.__name__}")
        except AssertionError as e:
            fallos += 1; print(f"FAIL  {fn.__name__} -> {e or 'assert'}")
        except Exception as e:  # noqa: BLE001
            import traceback
            fallos += 1; print(f"ERROR {fn.__name__} -> {type(e).__name__}: {e}")
            traceback.print_exc()
    print("-" * 50)
    print(f"{len(pruebas) - fallos}/{len(pruebas)} OK" + (f" ({fallos} problemas)" if fallos else ""))
    return 1 if fallos else 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    sys.exit(_main())
