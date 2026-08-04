#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# SPDX-License-Identifier: GPL-3.0-or-later
# ==========================================================================
"""Pruebas del motor REM A23 (respiratorio). Datos SINTÉTICOS. Correr:
    python tests/test_a23.py"""

import sys
import tempfile
from datetime import date
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

import modulos.rem_a23_respiratorio as a23   # noqa: E402

_TMP = Path(tempfile.mkdtemp(prefix="autorem_a23_"))

_HDR = ["NUMERO TIPO IDENTIFICACION", "FECHA ATENCION", "ACTIVIDADES", "DIAGNOSTICOS",
        "INSTRUMENTO", "TIPO ATENCION", "SEXO", "SECTOR", "NACIONALIDAD",
        "PUEBLO ORIGINARIO", "FECHA DE NACIMIENTO", "NOMBRES", "APELLIDO PATERNO",
        "APELLIDO MATERNO", "AÑOS"]


def _mk(rows):
    """rows = lista de dicts parciales (claves = subset de _HDR)."""
    p = _TMP / "aten.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(_HDR)
    for r in rows:
        ws.append([r.get(h, "") for h in _HDR])
    wb.save(p)
    return p


def _quiet(*_a, **_k): pass


def _fer():
    rows = [
        # RUN A: bronquitis (J20) por médico + autocuidado/control sala por kine
        {"NUMERO TIPO IDENTIFICACION": "A", "FECHA ATENCION": date(2026, 7, 10),
         "DIAGNOSTICOS": "J20.9 Bronquitis aguda", "ACTIVIDADES": "Consulta Sala (IRA/ERA)",
         "INSTRUMENTO": "Médico", "FECHA DE NACIMIENTO": date(2000, 7, 15), "SEXO": "Hombre",
         "NACIONALIDAD": "Chilena", "PUEBLO ORIGINARIO": "Ninguno", "SECTOR": "Rojo"},
        {"NUMERO TIPO IDENTIFICACION": "A", "FECHA ATENCION": date(2026, 7, 12),
         "ACTIVIDADES": "Control Sala (IRA/ERA) - Autocuidado", "INSTRUMENTO": "Kinesiólogo(a)"},
        # RUN B: IRA alta (J0) + consulta de morbilidad por médico -> Morbi
        {"NUMERO TIPO IDENTIFICACION": "B", "FECHA ATENCION": date(2026, 7, 5),
         "DIAGNOSTICOS": "J06.9 IRA alta", "INSTRUMENTO": "Médico",
         "TIPO ATENCION": "Consulta de morbilidad", "NACIONALIDAD": "Venezolana"},
        # RUN C: junio (fuera de la ventana de julio)
        {"NUMERO TIPO IDENTIFICACION": "C", "FECHA ATENCION": date(2026, 6, 20),
         "DIAGNOSTICOS": "J06 IRA", "INSTRUMENTO": "Médico", "TIPO ATENCION": "morbilidad"},
        # RUN D: espirometría por enfermería
        {"NUMERO TIPO IDENTIFICACION": "D", "FECHA ATENCION": date(2026, 7, 8),
         "ACTIVIDADES": "Espirometría basal", "INSTRUMENTO": "Enfermero(a)"},
    ]
    fer = a23.procesar(_mk(rows), mes=(2026, 7), log=_quiet)
    return fer.set_index("RUN")


def test_indicadores_mes():
    f = _fer()
    assert set(f.index) == {"A", "B", "C", "D"}
    assert f.loc["A", "REMA23 Bronquitis Aguda"] == "SI"
    assert f.loc["A", "REMA23 Autocuidado"] == "SI"
    assert f.loc["A", "REMA23 Control SALA Kine (act)"] == "SI"
    assert f.loc["A", "REMA23 Ira Alta"] == "NO"          # J20 no contiene J0
    assert f.loc["B", "REMA23 Ira Alta"] == "SI"
    assert f.loc["B", "REMA23 Morbi respiratoria"] == "SI"   # médico+morbilidad+base resp
    assert f.loc["D", "REMA23 Espirometría (act)"] == "SI"


def test_ventana_de_mes():
    f = _fer()
    # C solo tiene atención en junio -> no atendido en julio, todo NO
    assert f.loc["C", "¿Atendido 1 mes?"] == "NO"
    assert f.loc["C", "REMA23 Ira Alta"] == "NO"
    assert f.loc["A", "¿Atendido 1 mes?"] == "SI"


def test_demografia():
    f = _fer()
    assert int(f.loc["A", "Edad"]) == 26                  # DOB 2000-07 al 2026-07
    assert f.loc["A", "¿Originario o Migrante?"] == "NO"   # chilena, pueblo Ninguno
    assert f.loc["B", "¿Originario o Migrante?"] == "Migrante"  # Venezolana


_OHDR = ["NUMERO TIPO IDENTIFICACION", "FECHA ATENCION", "INSTRUMENTO", "SEXO", "FECHA DE NACIMIENTO",
         "1.- ¿PADECE DE SÍNDROME BRONQUIAL OBSTRUCTIVO?", "2.- ¿ES RECURRENTE?", "3.- ESTADO", "4.- GRAVEDAD SBOR",
         "9.- ¿PADECE DE ASMA BRONQUIAL?", "10.- ESTADO", "11.- GRAVEDAD ASMA BRONQUIAL", "13.- ESTADO DE CONTROL ASMA",
         "12.- FECHA DEL PRÓXIMO CONTROL",
         "14.- ¿PADECE ENFERMEDAD PULMONAR CRONICA?", "16.- TIPO EPOC", "17.- ESTADO", "19.- ESTADO DE CONTROL EPOC",
         "99.- RESULTADO ENCUESTA CALIDAD DE VIDA"]
_OKEY = {"RUN": 0, "FECHA": 1, "INSTR": 2, "SEXO": 3, "FNAC": 4,
         "SBOR_p": 5, "SBOR_rec": 6, "SBOR_est": 7, "SBOR_grav": 8,
         "ASMA_p": 9, "ASMA_est": 10, "ASMA_grav": 11, "ASMA_ctrl": 12, "ASMA_prox": 13,
         "EPOC_p": 14, "EPOC_tipo": 15, "EPOC_est": 16, "EPOC_ctrl": 17, "CDV": 18}


def _mk_otros(rows):
    p = _TMP / "otros.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.append(_OHDR)
    for r in rows:
        line = [""] * len(_OHDR)
        for k, v in r.items():
            line[_OKEY[k]] = v
        ws.append(line)
    wb.save(p)
    return p


def test_sala_bajo_control():
    aten = _mk([
        {"NUMERO TIPO IDENTIFICACION": "A", "FECHA ATENCION": date(2026, 7, 10), "INSTRUMENTO": "Médico", "FECHA DE NACIMIENTO": date(1990, 1, 1)},
        {"NUMERO TIPO IDENTIFICACION": "S", "FECHA ATENCION": date(2026, 7, 11), "INSTRUMENTO": "Médico", "FECHA DE NACIMIENTO": date(2023, 1, 1)},
        {"NUMERO TIPO IDENTIFICACION": "E", "FECHA ATENCION": date(2026, 7, 12), "INSTRUMENTO": "Médico", "FECHA DE NACIMIENTO": date(1970, 1, 1)},
    ])
    otros = _mk_otros([
        {"RUN": "A", "FECHA": date(2026, 5, 1), "INSTR": "Médico", "ASMA_p": "Si", "ASMA_grav": "Moderado", "ASMA_ctrl": "Controlado", "ASMA_est": "Ingreso"},
        {"RUN": "S", "FECHA": date(2026, 5, 1), "INSTR": "Médico", "SBOR_p": "Si", "SBOR_rec": "Si", "SBOR_grav": "Leve", "SBOR_est": "Seguimiento"},
        {"RUN": "E", "FECHA": date(2026, 5, 1), "INSTR": "Médico", "EPOC_p": "Si", "EPOC_tipo": "Tipo A", "EPOC_ctrl": "Controlado", "EPOC_est": "Ingreso"},
    ])
    f = a23.procesar(aten, otros=otros, mes=(2026, 7), log=lambda *a: None).set_index("RUN")
    assert f.loc["A", "SALA ASMA"] == "SI" and f.loc["A", "SALA Ingresado"] == "SI"
    assert f.loc["A", "SALA ASMA Gravedad"] == "Moderado"
    assert f.loc["A", "SALA SBOR"] == "NO"                       # no <5
    assert f.loc["S", "SALA SBOR"] == "SI" and f.loc["S", "SALA SBOR Gravedad"] == "Leve"
    assert f.loc["E", "SALA EPOC"] == "SI" and f.loc["E", "SALA EPOC Tipo"] == "Tipo A"


def test_seccion_g_inasistentes():
    """Inasistente a control de crónico = padece+estado válido y última Fecha
    Próximo Control vencida > umbral(edad) al corte (último día del mes reportado)."""
    import pandas as pd
    otros = _mk_otros([
        # adulta asma, próximo control vencido hace ~2 años -> inasistente (umbral 11m29d)
        {"RUN": "AD", "FECHA": date(2024, 1, 1), "INSTR": "Médico", "SEXO": "Mujer", "FNAC": date(1980, 1, 1),
         "ASMA_p": "Si", "ASMA_est": "Seguimiento", "ASMA_grav": "Leve", "ASMA_ctrl": "Controlado", "ASMA_prox": date(2024, 6, 1)},
        # adulto asma al día (próximo control futuro) -> NO
        {"RUN": "OK", "FECHA": date(2026, 6, 1), "INSTR": "Médico", "SEXO": "Hombre", "FNAC": date(1990, 1, 1),
         "ASMA_p": "Si", "ASMA_est": "Ingreso", "ASMA_grav": "Leve", "ASMA_ctrl": "Controlado", "ASMA_prox": date(2026, 9, 1)},
    ])
    od, _ = a23.cargar_otros(otros)
    counts, flags = a23._seccion_g(od, pd.Timestamp(2026, 7, 31))
    assert counts["Asma"]["Total"] == 1                       # solo AD (vencida)
    assert counts["Asma"]["Mujer"] == 1 and counts["Asma"]["Hombre"] == 0
    assert "AD" in flags["ASMA"] and "OK" not in flags["ASMA"]


def test_carga_multiarchivo():
    """cargar_otros acepta una LISTA (histórico multi-año) y concatena."""
    a = _mk_otros([{"RUN": "X", "FECHA": date(2025, 1, 1), "INSTR": "Médico", "ASMA_p": "Si"}])
    import shutil
    b = _TMP / "otros_b.xlsx"; shutil.copy(a, b)
    od, _ = a23.cargar_otros([a, b])
    assert len(od) == 2


_NSP_HDR = ["INSTRUMENTO", "TIPO DE ATENCION", "FECHA HORA CITA",
            "NUMERO TIPO IDENTIFICACION", "AÑOS"]
_NKEY = {"instr": 0, "tipo": 1, "fecha": 2, "run": 3, "anos": 4}


def _mk_nsp(rows):
    p = _TMP / "nsp.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.append(_NSP_HDR)
    for r in rows:
        line = [""] * len(_NSP_HDR)
        for k, v in r.items():
            line[_NKEY[k]] = v
        ws.append(line)
    wb.save(p)
    return p


def test_seccion_h():
    """Sección H: citas Control/Ingreso IRA/ERA no asistidas, por estamento × tramo
    (<20 / >=20), del mes por FECHA HORA CITA. Excluye KTR y otros meses/tipos."""
    import pandas as pd
    nsp = _mk_nsp([
        {"instr": "Médico", "tipo": "Control IRA", "fecha": "10-07-2026 09:00:00", "run": "A", "anos": 40},
        {"instr": "Médico", "tipo": "Ingreso ERA", "fecha": "12-07-2026 09:00:00", "run": "B", "anos": 15},
        {"instr": "Kinesiólogo(a)", "tipo": "Control ERA", "fecha": "13-07-2026 09:00:00", "run": "C", "anos": 60},
        {"instr": "Kinesiólogo(a)", "tipo": "kinesioterapia respiratoria IRA", "fecha": "14-07-2026 09:00:00", "run": "D", "anos": 30},  # KTR -> fuera
        {"instr": "Médico", "tipo": "Control IRA", "fecha": "10-06-2026 09:00:00", "run": "E", "anos": 40},  # junio -> fuera
        {"instr": "Médico", "tipo": "Consulta SAC", "fecha": "11-07-2026 09:00:00", "run": "F", "anos": 40},  # no IRA/ERA -> fuera
    ])
    d = a23.cargar_inasistentes(nsp)
    h = a23._seccion_h(d, pd.Timestamp(2026, 7, 1), pd.Timestamp(2026, 7, 31)).set_index("Profesional")
    assert h.loc["Médico/a", "Total"] == 2 and h.loc["Médico/a", "Menor de 20"] == 1 and h.loc["Médico/a", "20 y más"] == 1
    assert h.loc["Kinesiólogo/a", "Total"] == 1        # Control ERA; KTR excluido
    assert h.loc["TOTAL", "Total"] == 3                # junio y Consulta SAC fuera


def _main():
    pruebas = [v for k, v in sorted(globals().items())
               if k.startswith("test_") and callable(v)]
    fallos = 0
    for fn in pruebas:
        try:
            fn(); print(f"PASS  {fn.__name__}")
        except AssertionError as e:
            fallos += 1; print(f"FAIL  {fn.__name__} -> {e or 'assert'}")
        except Exception as e:  # noqa: BLE001
            fallos += 1; print(f"ERROR {fn.__name__} -> {type(e).__name__}: {e}")
    print("-" * 50)
    print(f"{len(pruebas) - fallos}/{len(pruebas)} OK" + (f" ({fallos} problemas)" if fallos else ""))
    return 1 if fallos else 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    sys.exit(_main())
