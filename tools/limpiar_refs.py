#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# Version: 1.7.1
#
# Distributed WITHOUT ANY WARRANTY. GPL-3.0-or-later:
# <https://www.gnu.org/licenses/>.
# ==========================================================================
"""
limpiar_refs.py — recorta los EXPORTS de ejemplo de `refs tablas/` a SOLO EL HEADER.

Privacidad-by-design (CLAUDE.md §8): cualquier export de RAYEN/IRIS que se agregue a
la carpeta de referencia puede traer PII de paciente en sus filas. Este tool ubica la
fila de encabezado y **borra todo lo que hay debajo**, dejando solo la estructura
(columnas). Así el repo nunca ve valores de paciente aunque uno olvide anonimizar.

GARANTÍA: NO lee ni imprime valores de celda. Solo cuenta celdas no vacías para
localizar el header y reporta nº de filas antes/después. Nadie (ni humano ni modelo)
ve el dato.

NO toca (denylist): templates/planillas con fórmulas y specs, que SÍ necesitan sus
filas — `.xlsm/.xltx` (SA_26, SP_26), y por nombre: minimanual, REM comentado,
calculador, manual, arsenal. Todo lo demás (.xlsx export plano) se recorta.

USO:
    python tools/limpiar_refs.py            # DRY-RUN: reporta qué recortaría
    python tools/limpiar_refs.py --aplicar  # aplica el recorte
    python tools/limpiar_refs.py --aplicar archivo1.xlsx archivo2.xlsx  # solo esos
"""

import sys
from pathlib import Path

import openpyxl

# Windows: consola cp1252 revienta con ✔·→ (UnicodeEncodeError ⊂ ValueError). Ver §4.5.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# Carpeta de referencia (relativa a la raíz del repo = padre de tools/).
CARPETA = Path(__file__).resolve().parent.parent / "refs tablas"

# Se conservan intactos (tienen fórmulas/estructura que depende de sus filas).
DENY_EXT = {".xlsm", ".xltx", ".xltm", ".dotx", ".potx"}
# 'maestro' = Maestro de Actividades (catálogo completo RAYEN act↔profesional, 217k
# filas, sin PII de paciente). Es REFERENCIA de por vida → se mantiene INTACTO.
DENY_NOMBRE = ("minimanual", "comentado", "calculador", "manual", "arsenal", "maestro")

# Fila de header = primera fila con al menos esta cantidad de celdas no vacías.
MIN_CELDAS_HEADER = 5


def _es_referencia_export(p: Path) -> bool:
    """True si el archivo es un export de datos plano (candidato a recorte)."""
    if p.suffix.lower() not in (".xlsx",):
        return False
    n = p.name.lower()
    if p.suffix.lower() in DENY_EXT:
        return False
    return not any(k in n for k in DENY_NOMBRE)


def _fila_header(ws) -> int | None:
    """Primera fila (1-indexada) con >= MIN_CELDAS_HEADER celdas no vacías. NO
    devuelve ni mira los valores: solo cuenta no-vacíos."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50), start=1):
        no_vacias = sum(1 for c in row if c.value not in (None, ""))
        if no_vacias >= MIN_CELDAS_HEADER:
            return i
    return None


def recortar(p: Path, aplicar: bool) -> tuple[str, int, int]:
    """Recorta un .xlsx a header-only (todas las hojas). Devuelve (estado, filas_antes,
    filas_despues) SUMADAS sobre las hojas. Sin leer valores."""
    wb = openpyxl.load_workbook(p)
    antes = despues = 0
    cambio = False
    for ws in wb.worksheets:
        h = _fila_header(ws)
        antes += ws.max_row
        if h is None:                      # hoja sin header reconocible → no tocar
            despues += ws.max_row
            continue
        if ws.max_row > h:
            if aplicar:
                ws.delete_rows(h + 1, ws.max_row - h)
            cambio = True
        despues += h
    if aplicar and cambio:
        wb.save(p)
    wb.close()
    estado = "recortado" if (cambio and aplicar) else ("recortaría" if cambio else "ya limpio")
    return estado, antes, despues


def main(argv):
    aplicar = "--aplicar" in argv
    nombrados = [a for a in argv if not a.startswith("--")]
    if nombrados:
        objetivos = [Path(a) if Path(a).is_absolute() else CARPETA / a for a in nombrados]
    else:
        objetivos = sorted(CARPETA.glob("*.xlsx"))

    print(f"{'APLICANDO' if aplicar else 'DRY-RUN'} · carpeta: {CARPETA}")
    tocados = 0
    for p in objetivos:
        if not p.exists():
            print(f"  ✗ no existe: {p.name}")
            continue
        if not _es_referencia_export(p):
            print(f"  — omitido (denylist/no-export): {p.name}")
            continue
        estado, antes, despues = recortar(p, aplicar)
        marca = "·" if estado == "ya limpio" else "✔"
        print(f"  {marca} {estado}: {p.name}  ({antes} → {despues} filas)")
        if estado in ("recortado", "recortaría"):
            tocados += 1
    if not aplicar and tocados:
        print(f"\n{tocados} archivo(s) con filas de datos. Corré con --aplicar para recortarlos.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
