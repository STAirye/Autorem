#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 5 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# Version: 1.8.2
#
# This program is free software: you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by the
# Free Software Foundation, either version 3 of the License, or (at your
# option) any later version. Distributed WITHOUT ANY WARRANTY. See the GNU
# General Public License for more details: <https://www.gnu.org/licenses/>.
# ==========================================================================
"""
rem_sp_p6_poblacion.py — REM SP·P6 A.1 «Población en control PSM» (Fase 2).

Consume la tabla `PSM_Poblacion` de `programas.poblacion.construir_poblacion()`
y arma la grilla copy-paste al `SP_26_V1.1.xlsm` hoja P6, sección A.1 (filas
13-58), respetando la máscara de celdas protegidas de la plantilla real. Ver
`docs/SP_P6_poblacion_plan.md` §5.

La MÁSCARA (`MASCARA_BANDA`/`EXCLUYE_DEMO`) se extrajo DIRECTO de
`refs tablas/SP_26_V1.1.xlsm` (protección real de celdas), no de la prosa del
plan: se encontró y CORRIGIÓ una discrepancia ahí (fila 38 «Otros trastornos
del comportamiento…infancia»: el plan decía rango reportable 0-19, la
plantilla real abre hasta 20-24 — ver comentario en MASCARA_BANDA). Si MINSAL
cambia la plantilla, re-extraer con el mismo método (recorrer
`ws.cell(...).protection.locked` de la hoja P6) y actualizar acá.
"""

import re

import pandas as pd

from programas.rem_utils import (norm, BANDAS_A06 as BANDAS, LBL_A06 as LBL, _band_idx, grid,
                                 _hombre, _mujer, ArchivoInvalido)
from programas.rem_saludmental import OVERRIDE_SUBTIPO

# ======================================================================
# Máscara de la plantilla (SP_26_V1.1.xlsm, hoja P6, filas 13-58)
# ======================================================================
# fila -> (idx_min, idx_max) de banda etaria ABIERTA (0-based sobre las 17
# bandas de BANDAS_A06/LBL_A06: 0-4,5-9,...,80+). Filas sin recorte = (0,16).
MASCARA_BANDA = {
    13: (0, 16), 15: (0, 16), 16: (0, 16), 17: (0, 16), 18: (0, 16), 19: (0, 16),
    20: (0, 16), 21: (0, 16),
    22: (1, 16), 23: (1, 16),                 # Suicidio: sin 0-4 (pliega a 5-9)
    24: (0, 16),
    25: (0, 16), 26: (0, 16), 27: (0, 16),
    28: (2, 11),                              # Depresión post parto: 10-59 (pliega a 55-59 hacia arriba)
    29: (0, 16), 30: (0, 16), 31: (0, 16), 32: (0, 16), 33: (0, 16), 34: (0, 16),
    35: (0, 16), 36: (0, 16),                 # TDAH/disocial: SIN recorte etario en el SP (§5.0.1)
    37: (0, 2),                               # Ansiedad separación: 0-14 (pliega a 10-14)
    38: (0, 4),                               # Otros infancia: 0-24 (pliega a 20-24) — el plan decía
                                               # 0-19; la plantilla real abre hasta 20-24 (verificado
                                               # contra ws.cell().protection.locked, sep-2026).
    39: (0, 16), 40: (0, 16), 41: (0, 16), 42: (0, 16), 43: (0, 16),
    44: (0, 16), 45: (0, 16), 46: (0, 16),    # Demencias: sin recorte etario
    47: (0, 16), 48: (0, 16), 49: (0, 16), 50: (0, 16), 51: (0, 16),
    52: (0, 16), 53: (0, 16), 54: (0, 16), 55: (0, 16), 56: (0, 16),
    57: (0, 16), 58: (0, 16),
}

# fila -> etiquetas (concepto, subconcepto) — copiadas literal de las columnas
# A/B de la plantilla (fuente: refs tablas/SP_26_V1.1.xlsm).
ROW_LABELS = {
    13: ("Número de personas en control en el programa", None),
    15: ("Violencia física", "Víctima"), 16: (None, "Agresor/a"),
    17: ("Violencia sexual", "Víctima"), 18: (None, "Agresor/a"),
    19: ("Violencia psicológica", "Víctima"), 20: (None, "Agresor/a"),
    21: ("Abuso sexual", None),
    22: ("Suicidio", "Ideación"), 23: (None, "Intento"),
    24: ("PERSONAS CON DIAGNÓSTICOS DE TRASTORNOS MENTALES", None),
    25: ("Trastornos del humor (afectivos)", "Depresión leve"),
    26: (None, "Depresión moderada"), 27: (None, "Depresión grave"),
    28: (None, "Depresión post parto"),
    29: (None, "Trastorno bipolar"),
    30: ("Trastornos mentales y del comportamiento debido a consumo sustancias psicotrópicas",
         "Consumo perjudicial de alcohol"),
    31: (None, "Consumo dependiente del alcohol"),
    32: (None, "Consumo perjudicial de drogas"),
    33: (None, "Consumo dependiente de drogas"),
    34: (None, "Consumo de drogas y alcohol"),
    35: ("Trastornos del comportamiento y de las emociones de comienzo habitual en la infancia y adolescencia",
         "Trastorno hipercinético"),
    36: (None, "Trastorno disocial desafiante y oposicionista"),
    37: (None, "Trastorno de ansiedad de separación en la infancia"),
    38: (None, "Otros trastornos del comportamiento y de las emociones de comienzo habitual en la infancia y adolescencia"),
    39: ("Trastornos de ansiedad", "Trastorno de estrés post traumático"),
    40: (None, "Trastorno de pánico"), 41: (None, "Fobias sociales"),
    42: (None, "Trastornos de ansiedad generalizada"), 43: (None, "Otros trastornos de ansiedad"),
    44: ("Demencias (incluye Alzheimer)", "Leve"), 45: (None, "Moderado"), 46: (None, "Avanzado"),
    47: ("Esquizofrenia", None),
    48: ("Trastorno adaptativo", None),
    49: ("Trastornos de la conducta alimentaria", None),
    50: ("Retraso mental", None),
    51: ("Trastorno de personalidad", None),
    52: ("Trastorno generalizados del desarrollo", "Autismo"),
    53: (None, "Asperger"), 54: (None, "Síndrome de Rett"),
    55: (None, "Trastorno desintegrativo de la infancia"),
    56: (None, "Trastorno generalizado del desarrollo no específico"),
    57: ("Epilepsia", None),
    58: ("Otras", None),
}

# Columnas demográficas AN..AX del template, en orden, y a qué flag corresponden.
# AN/AO/AT/AU/AV son de valor único (no H/M); AP/AQ, AR/AS, AW/AX son pares H/M.
DEMO_TODAS = ["AN", "AO", "AP_H", "AQ_M", "AR_H", "AS_M", "AT", "AU", "AV", "AW", "AX"]
# fila -> columnas demográficas BLOQUEADAS (excepción; todo lo no listado abre).
EXCLUYE_DEMO = {
    28: {"AN", "AP_H", "AR_H", "AX"},                    # Depresión post parto: solo mujeres
    35: {"AO"}, 36: {"AO"},                              # TDAH/disocial: no aplica madre<5
    44: {"AN", "AO", "AT", "AU"}, 45: {"AN", "AO", "AT", "AU"}, 46: {"AN", "AO", "AT", "AU"},
}


def _mascara_demo(fila):
    return [c for c in DEMO_TODAS if c not in EXCLUYE_DEMO.get(fila, set())]


# ======================================================================
# Config por diagnóstico: fila P6 -> columna Ferrada (docs/SP_P6_config_por_dx.md)
# ======================================================================
# Filas 25-58 SIN subtipo -> 1:1 con una columna (form) de PSM_Poblacion.
FILA_SIMPLE = {
    29: "Bipolaridad (form)",
    30: "OH Perjudicial (form)", 31: "OH Dependiente (form)",
    32: "Drogas Perjudicial (form)", 33: "Drogas Dependiente (form)",
    34: "OH y Drogas (form)",
    35: "TDAH (form)", 36: "Oposicionista desafiante (form)",
    37: "Ansiedad separación (form)", 38: "Otras Infancia/Adolescencia (form)",
    47: "Esquizofrenia (form)", 48: "Adaptativo (form)",
    49: "Conducta Alimentaria (form)", 50: "Retraso Mental (form)", 51: "Personalidad (form)",
    52: "Autismo (form)", 53: "Asperger (form)", 54: "Rett (form)",
    55: "Desintegrativo niñez (form)", 56: "TGD (form)", 58: "Otros (form)",
}

# Todas las filas 25-58 -> su columna (form) de origen (incluye las bucketed).
# 57 no tiene columna: Epilepsia no se reporta.
FILA_COLUMNA_DX = dict(FILA_SIMPLE)
FILA_COLUMNA_DX.update({
    25: "Depresión (form)", 26: "Depresión (form)", 27: "Depresión (form)",
    28: "Depresión Postparto (form)",
    39: "Ansiedad (form)", 40: "Ansiedad (form)", 41: "Ansiedad (form)",
    42: "Ansiedad (form)", 43: "Ansiedad (form)",
    44: "Demencia (form)", 45: "Demencia (form)", 46: "Demencia (form)",
})

# Bucketing por subtipo (§5.1/Tabla C): fila objetivo, columna de estado
# origen, columna del subtipo crudo, reglas (keywords, fila) ordenadas.
DEPRESION_REGLAS = [(["LEVE"], 25), (["MODERAD"], 26), (["GRAVE"], 27)]
DEMENCIA_REGLAS = [(["LEVE"], 44), (["MODERAD"], 45), (["AVANZA"], 46)]
SUICIDIO_REGLAS = [(["IDEACION"], 22), (["INTENTO"], 23)]
# Ansiedad REUSA rem_saludmental.OVERRIDE_SUBTIPO[43] (misma regla que el A05,
# incl. el orden PÁNICO-antes-que-FOBIA por 'agoraFOBIA') en vez de duplicarla.
_ANSIEDAD_FILA = {"TEPT": 39, "Pánico": 40, "Fobia social": 41, "Generalizada": 42, "Otros": 43}
ANSIEDAD_REGLAS = [(keys, _ANSIEDAD_FILA[nombre]) for keys, nombre in OVERRIDE_SUBTIPO[43]]
VIOLENCIA_TIPO_REGLAS = [(["FISICA"], (15, 16)), (["SEXUAL"], (17, 18)), (["PSICOL"], (19, 20))]

# Exclusiones comodín (§5.1): ELIMINADAS (sep-2026). El autor las sacó por completo
# de su metodología manual del P6 -> el módulo tampoco las aplica. Las filas cajón de
# sastre (38, 43, 48) tributan sin restringir por comorbilidad. Motivo: el filtro con
# umbral 0 de la fila 43 sacaba ~199 personas que el REM manual sí cuenta (validado
# ago-2026, ver docs/SP_P6_poblacion_plan.md §5.1).

# Filas GES-obligatorias o factor de riesgo -> AV (PIC) = total de la fila (§5.4.2 WIP).
GES_FILAS_AV = set(range(15, 24)) | {25, 26, 27, 28, 44, 45, 46}


# ======================================================================
# Motor
# ======================================================================
_RUT_FORMATO = re.compile(r"^\d{6,9}-[\dkK]$")
_TECHO_NO_RUT = 0.05   # §5.5.1: si el filtro descarta más de esto, el roto es EL FILTRO
# §5.5.2: forma válida pero SIGNIFICADO incorrecto -> la forma sola no basta. Hoy solo
# "RUN Responsable" (colisión de clave con un tercero — programas.poblacion.
# cargar_inscritos ya lo saca ANTES de deduplicar; este chequeo es un segundo resguardo
# por si `P` llega armado de otra forma). NO cachear esta lista entre corridas: el
# conjunto rota mes a mes (Registro Civil ~5 días para emitir el RUT del RN).
_TIPOID_FORMA_OK_SIGNIFICADO_MAL = ["RUN RESPONSABLE"]


def _base_valida(P, log):
    """Filtro base §5.1 (Estado=Activo & Activo12m=SI & Ingresado=SI) + descarta
    identificador no-RUN (§5.5/§5.5.2). DOS chequeos, ninguno basta solo:
      1. FORMA — regex sobre 'Número' (^\\d{6,9}-[dkK]$): descarta FONASA/pasaportes
         y sobrevive a que RAYEN renombre la etiqueta (era "RUN", no "RUT" — la
         primera corrida real, exigiendo el literal "RUT", descartó al 100%).
      2. ETIQUETA (blacklist chica) — "RUN Responsable" tiene forma de RUT VÁLIDA
         pero es el RUN de un tercero (recién nacido <1 mes sin RUT propio): la forma
         sola no lo detecta.
    La etiqueta cruda queda como dato informativo en Revisar_Administrativo.

    Guardarraíl (§5.5.1): si el total descartado supera `_TECHO_NO_RUT` del padrón, el
    filtro es el que está roto, no los datos -> `ArchivoInvalido` explícito en vez de
    seguir y emitir un P6_A1 lleno de ceros en silencio. Y tras filtrar, 'Número' debe
    quedar ÚNICO — si sobrevive una colisión, `ArchivoInvalido` también.
    Devuelve (P_filtrada, filas_revisar)."""
    revisar = []

    numero_str = P["Número"].astype(str).str.strip()
    formato_ok = numero_str.str.match(_RUT_FORMATO)
    tipoid_n = P["Tipo de identificación"].map(norm)
    es_responsable = tipoid_n.isin(_TIPOID_FORMA_OK_SIGNIFICADO_MAL)
    no_rut = ~formato_ok | es_responsable

    n_no_rut, total = int(no_rut.sum()), len(P)
    if total and n_no_rut / total > _TECHO_NO_RUT:
        raise ArchivoInvalido(
            "validador_rut",
            f"El validador de identificador descartó {n_no_rut} de {total} personas "
            f"({n_no_rut / total:.0%}) — muy por encima del techo esperado "
            f"({_TECHO_NO_RUT:.0%}, §5.5.1 del plan). El roto es el validador, no los "
            "datos: revisa cómo viene 'Tipo de identificación'/'Número' en el Informe "
            "Inscritos y Adscritos (¿es el correcto, sin modificar?) antes de seguir.")
    for _, row in P.loc[no_rut].iterrows():
        motivo = ("RUN Responsable (colisión de clave con un tercero, probable "
                 "recién nacido <1 mes)" if norm(row["Tipo de identificación"]) in
                 _TIPOID_FORMA_OK_SIGNIFICADO_MAL else "Identificador no-RUN (formato inválido)")
        revisar.append({"RUN": row["Número"], "Motivo": motivo, "Fila_P6": "",
                        "Detalle": row["Tipo de identificación"], "Valor_crudo": row["Número"],
                        "Categoria": "Administrativo"})
    if n_no_rut:
        log(f"[sp_p6] {n_no_rut} persona(s) descartadas por identificador "
            "(formato no-RUN o RUN Responsable): van a Revisar_Administrativo.")
    P = P.loc[~no_rut].copy()

    dup = P["Número"].duplicated(keep=False)
    if dup.any():
        ruts = ", ".join(sorted(set(P.loc[dup, "Número"]))[:20])
        raise ArchivoInvalido(
            "colision_run",
            f"Quedaron {int(dup.sum())} 'Número' DUPLICADOS después de filtrar "
            f"(§5.5.2): {ruts}. Es una colisión de clave sin resolver (dos personas "
            "con el mismo RUN) — no se puede armar el P6 así.")

    f_estado = P["Estado"].map(norm) == "ACTIVO"
    f_a12m = P["¿Activo 12m?"] == "SI"
    f_ingr = P["¿Ingresado?"] == "SI"
    base = f_estado & f_a12m & f_ingr

    # CASCADA de filtros: N individual y acumulado por etapa. Sirve para diffear
    # paso a paso contra la tabla dinámica del PowerBI ('Población SM' -> filtros
    # globales estado + activo 12m + ingresado) y aislar en CUÁL filtro empieza la
    # divergencia, en vez de mirar solo el total final.
    log(f"[sp_p6] --- cascada de filtros (padron con identificador valido: {len(P)}) ---")
    acum = pd.Series(True, index=P.index)
    for nombre, f in (("Estado=Activo", f_estado), ("Activo 12m=SI", f_a12m),
                      ("Ingresado=SI", f_ingr)):
        acum &= f
        log(f"[sp_p6]    {nombre:<16} solo={int(f.sum()):>6}   acumulado={int(acum.sum()):>6}")
    log(f"[sp_p6] base (los 3 filtros): {int(base.sum())} de {len(P)} personas del snapshot")

    sexo_raro = P.loc[base, "Sexo"].map(lambda s: not (_hombre(s) or _mujer(s)))
    for run, sx in zip(P.loc[base].loc[sexo_raro, "Número"], P.loc[base].loc[sexo_raro, "Sexo"]):
        motivo = "Sexo = No informado" if sx == "No informado" else "Sexo/género no binario"
        revisar.append({"RUN": run, "Motivo": motivo, "Fila_P6": "", "Detalle": "",
                        "Valor_crudo": sx, "Categoria": "Administrativo"})
    if sexo_raro.any():
        log(f"[sp_p6] {int(sexo_raro.sum())} persona(s) con sexo fuera de "
            "{{Hombre,Mujer}}: no se pueden ubicar en columna H/M del P6 (Revisar_Administrativo).")

    return P.loc[base].copy(), revisar


def _banda_efectiva(edad, lo, hi):
    """(edad_para_grid, plegada) — si la banda real cae fuera de [lo,hi], usa
    el límite de banda más cercano (edad representativa = límite inferior de
    esa banda) para que `grid()` la clasifique ahí. `edad_para_grid`=None si
    no hay edad."""
    idx = _band_idx(edad, BANDAS)
    if idx is None:
        return None, False
    folded = min(max(idx, lo), hi)
    return BANDAS[folded][0], folded != idx


def _grid_y_detalle(sub, fila, detalle_rows, revisar):
    """Arma la fila del grid (bandas F..AM vía `grid()` + demografía AN..AX)
    para un DataFrame `sub` de personas que tributan a `fila`, y acumula
    filas de detalle/revisión (edad plegada)."""
    lo, hi = MASCARA_BANDA[fila]
    sub = sub.copy()
    efectiva, plegada = [], []
    for edad in sub["Edad"]:
        e, p = _banda_efectiva(edad, lo, hi)
        efectiva.append(e)
        plegada.append(p)
    sub["_edad_grid"] = efectiva
    sub["_plegada"] = plegada
    grid_sub = sub.rename(columns={"Sexo": "sexo"}).assign(edad=sub["_edad_grid"])
    fila_grid = grid(grid_sub, BANDAS, LBL)

    demo_abiertas = set(_mascara_demo(fila))
    for col, flag in DEMO_COLS_MAP.items():
        activo_col = col in demo_abiertas
        marcados = sub[sub[flag]] if flag in sub.columns else sub.iloc[0:0]
        if activo_col:
            fila_grid[col] = len(marcados)
        else:
            fila_grid[col] = 0
            if len(marcados):
                for run in marcados["Número"]:
                    revisar.append({"RUN": run, "Motivo": "Dato demográfico no aplica en esta fila",
                                    "Fila_P6": fila, "Detalle": col, "Valor_crudo": "SI",
                                    "Categoria": "Administrativo"})

    for _, r in sub.iterrows():
        if r["_plegada"]:
            revisar.append({"RUN": r["Número"], "Motivo": "Edad plegada al rango reportable",
                            "Fila_P6": fila, "Detalle": f"edad real {r['Edad']}",
                            "Valor_crudo": f"-> banda destino desde idx {lo}-{hi}",
                            "Categoria": "Clinico"})
        detalle_rows.append({
            "RUN": r["Número"], "Fila_P6": fila,
            "Concepto": ROW_LABELS.get(fila, ("", ""))[0], "Subconcepto": ROW_LABELS.get(fila, ("", ""))[1],
            "Sexo": r["Sexo"], "Edad": r["Edad"], "Plegada": "SI" if r["_plegada"] else "",
        })
    return fila_grid


# Flags de persona (columnas de PSM_Poblacion) -> columna demográfica del P6.
# Gestante/Madre<5 filtran por sexo registral femenino acá (§5.4.1) — NO en
# programas/poblacion.py, que las deja sin filtrar a propósito.
DEMO_COLS_MAP = {
    "AN": "_dem_gestante", "AO": "_dem_madre5",
    "AP_H": "_dem_pueblo_h", "AQ_M": "_dem_pueblo_m",
    "AR_H": "_dem_migrante_h", "AS_M": "_dem_migrante_m",
    "AT": "_dem_sename", "AU": "_dem_mejorninez",
    "AV": "_dem_pic",   # se sobreescribe aparte, ver §5.4.2
    "AW": "_dem_trans_masc", "AX": "_dem_trans_fem",
}


def _preparar_demografia(P, revisar, log=print):
    """Agrega al DataFrame las columnas booleanas `_dem_*` que consume
    `_grid_y_detalle`. Aplica el filtro por SEXO REGISTRAL de §5.4.1 a
    Gestante/Madre<5 (con aviso a Revisar_Administrativo si se descarta por sexo)."""
    P = P.copy()
    es_mujer = P["Sexo"].map(_mujer)
    es_hombre = P["Sexo"].map(_hombre)

    gestante_si = P["¿Embarazada?"] == "SI"
    P["_dem_gestante"] = gestante_si & es_mujer
    descartadas_g = gestante_si & ~es_mujer
    madre5_si = P["Madre <5 años"] == "SI"
    P["_dem_madre5"] = madre5_si & es_mujer
    descartadas_m = madre5_si & ~es_mujer
    for run in P.loc[descartadas_m, "Número"]:
        revisar.append({"RUN": run, "Motivo": "Hombre marcado «madre de hijo <5»",
                        "Fila_P6": "", "Detalle": "flag descartado, corregir en RAYEN",
                        "Valor_crudo": "SI", "Categoria": "Administrativo"})
    n_descartadas = int(descartadas_g.sum()) + int(descartadas_m.sum())
    if n_descartadas:
        log(f"[sp_p6] {n_descartadas} flag(s) de Gestante/Madre<5 descartado(s) por "
            "sexo registral no-femenino (§5.4.1, Revisar_Administrativo).")

    origen = P["¿Originario o Migrante?"]
    P["_dem_pueblo_h"] = (origen == "Originario") & es_hombre
    P["_dem_pueblo_m"] = (origen == "Originario") & es_mujer
    P["_dem_migrante_h"] = (origen == "Migrante") & es_hombre
    P["_dem_migrante_m"] = (origen == "Migrante") & es_mujer
    P["_dem_sename"] = P["PROTECCION NIÑEZ"] == "SENAME"
    P["_dem_mejorninez"] = P["PROTECCION NIÑEZ"] == "Mejor Niñez"
    P["_dem_pic"] = False   # AV se computa aparte (§5.4.2), no por persona
    P["_dem_trans_masc"] = False
    P["_dem_trans_fem"] = False   # Fase 2 no trae el 'Informe Inscritos' (TRANS) — pendiente, ver notas
    return P


def _tributarios_violencia(P, revisar):
    """{fila: DataFrame} para 15-20 (Violencia física/sexual/psicológica x
    víctima/agresor), fusionando Abuso Sexual (§5.1: sin sub-pregunta propia
    de víctima/agresor -> se asume Víctima, fila 17, por defecto)."""
    out = {f: [] for f in (15, 16, 17, 18, 19, 20)}
    activo = P["Violencia (form)"] == "Activo"
    tipo_n = P["Violencia Tipo (form)"].map(norm)
    va_n = P["Violencia Victima o Agresor (form)"].map(norm)
    sin_subtipo = activo & ((tipo_n == "") | (va_n == ""))
    for run in P.loc[sin_subtipo, "Número"]:
        revisar.append({"RUN": run, "Motivo": "Dx activo sin subtipo registrado",
                        "Fila_P6": "15-20", "Detalle": "Violencia", "Valor_crudo": "",
                        "Categoria": "Clinico"})
    usable = activo & ~sin_subtipo
    for keys, (f_vic, f_agr) in VIOLENCIA_TIPO_REGLAS:
        m = usable & tipo_n.str.contains("|".join(norm(k) for k in keys), regex=True, na=False)
        es_agresor = va_n.str.contains("AGRESOR", na=False)
        out[f_vic].append(P.loc[m & ~es_agresor])
        out[f_agr].append(P.loc[m & es_agresor])

    abuso_activo = P["Abuso Sexual (form)"] == "Activo"
    out[17].append(P.loc[abuso_activo])   # sin víctima/agresor propio -> Víctima por defecto

    return {f: (pd.concat(dfs, ignore_index=True).drop_duplicates(subset="Número")
               if dfs else P.iloc[0:0]) for f, dfs in out.items()}


def _tributarios_suicidio(P, revisar):
    out = {22: [], 23: []}
    activo = P["Suicidio (form)"] == "Activo"
    tipo_n = P["Suicidio Tipo (form)"].map(norm)
    sin_subtipo = activo & (tipo_n == "")
    for run in P.loc[sin_subtipo, "Número"]:
        revisar.append({"RUN": run, "Motivo": "Dx activo sin subtipo registrado",
                        "Fila_P6": "22-23", "Detalle": "Suicidio", "Valor_crudo": "",
                        "Categoria": "Clinico"})
    usable = activo & ~sin_subtipo
    for keys, fila in SUICIDIO_REGLAS:
        m = usable & tipo_n.str.contains("|".join(norm(k) for k in keys), regex=True, na=False)
        out[fila] = P.loc[m]
    return out


def _tributarios_bucket(P, col_activo, col_subtipo, reglas, filas, revisar, etiqueta):
    """Genérico para Depresión (25-27), Ansiedad (39-43), Demencia (44-46):
    D5 — activo pero subtipo vacío/no reconocido -> no tributa, va a revisar."""
    out = {f: P.iloc[0:0] for f in filas}
    activo = P[col_activo] == "Activo"
    sub_n = P[col_subtipo].map(norm)
    asignada = pd.Series(False, index=P.index)
    for keys, fila in reglas:
        m = activo & ~asignada & sub_n.str.contains("|".join(norm(k) for k in keys), regex=True, na=False)
        out[fila] = P.loc[m]
        asignada |= m
    sin_reconocer = activo & ~asignada
    for run in P.loc[sin_reconocer, "Número"]:
        revisar.append({"RUN": run, "Motivo": "Dx activo sin subtipo registrado/reconocible",
                        "Fila_P6": "/".join(map(str, filas)), "Detalle": etiqueta,
                        "Valor_crudo": "", "Categoria": "Clinico"})
    return out


def construir_p6(P, log=print):
    """`P` = PSM_Poblacion (de `programas.poblacion.construir_poblacion()`).
    Devuelve dict: 'grid' (DataFrame, 1 fila por fila-P6), 'detalle' (DataFrame
    auditable), 'revisar_administrativo'/'revisar_clinico' (DataFrames — §5.5:
    identidad/registro vs criterio clínico, hojas separadas), 'bloques'
    (rectángulos pegables), 'mes' (año,mes)."""
    revisar = []
    mes = P.attrs.get("mes")
    Pv, revisar_base = _base_valida(P, log)
    Pv.attrs.clear()   # pandas intenta comparar .attrs al concatenar slices más abajo;
                       # 'egreso_divergencias' es un DataFrame y esa comparación revienta
                       # ("truth value of a DataFrame is ambiguous"). No se necesita en Pv:
                       # 'mes' ya se leyó de P arriba.
    revisar += revisar_base
    Pv = _preparar_demografia(Pv, revisar, log)

    tributarios = {}
    tributarios.update(_tributarios_violencia(Pv, revisar))
    tributarios.update(_tributarios_suicidio(Pv, revisar))
    tributarios.update(_tributarios_bucket(
        Pv, "Depresión (form)", "Depresión gravedad (form)", DEPRESION_REGLAS,
        (25, 26, 27), revisar, "Depresión"))
    tributarios[28] = Pv.loc[Pv["Depresión Postparto (form)"] == "Activo"]
    tributarios.update(_tributarios_bucket(
        Pv, "Ansiedad (form)", "Ansiedad (tipo)", ANSIEDAD_REGLAS,
        (39, 40, 41, 42, 43), revisar, "Ansiedad"))
    tributarios.update(_tributarios_bucket(
        Pv, "Demencia (form)", "Demencia gravedad (form)", DEMENCIA_REGLAS,
        (44, 45, 46), revisar, "Demencia"))
    for fila, col in FILA_SIMPLE.items():
        tributarios[fila] = Pv.loc[Pv[col] == "Activo"]
    tributarios[21] = Pv.iloc[0:0]   # Abuso sexual: no capturable, fila fija en 0
    tributarios[57] = Pv.iloc[0:0]   # Epilepsia: no es del programa SM, fila fija en 0

    # (exclusiones comodín §5.1 ELIMINADAS sep-2026: las filas cajón de sastre
    #  38/43/48 tributan sin restringir por comorbilidad. Ver módulo, arriba.)

    comorbidos_ges =(set(tributarios[25]["Número"]) | set(tributarios[26]["Número"]) |
                      set(tributarios[27]["Número"]) | set(tributarios[28]["Número"])) & \
                     (set(tributarios[44]["Número"]) | set(tributarios[45]["Número"]) |
                      set(tributarios[46]["Número"]))
    for run in comorbidos_ges:
        revisar.append({"RUN": run, "Motivo": "Comorbilidad GES depresión + demencia (AV24)",
                        "Fila_P6": 24, "Detalle": "el PIC se acumula 2 veces; decidir primacía a mano",
                        "Valor_crudo": "", "Categoria": "Clinico"})

    detalle_rows = []
    filas_grid = {}
    for fila in list(range(15, 24)) + list(range(25, 59)):
        if fila not in tributarios:
            continue
        filas_grid[fila] = _grid_y_detalle(tributarios[fila], fila, detalle_rows, revisar)

    # AV (Plan de Cuidado Integral) — regla WIP §5.4.2: total de fila si GES/FR, 0 en el resto.
    for fila, g in filas_grid.items():
        g["AV"] = g["Ambos"] if fila in GES_FILAS_AV else 0

    # Fila 24: DISTINCTCOUNT(RUN) con >=1 dx de las filas 25-58 (excluye 15-23).
    runs_24 = set()
    for f in range(25, 59):
        if f in tributarios:
            runs_24 |= set(tributarios[f]["Número"])
    sub24 = Pv[Pv["Número"].isin(runs_24)]
    filas_grid[24] = _grid_y_detalle(sub24, 24, detalle_rows, revisar)
    filas_grid[24]["AV"] = sum(filas_grid[f]["AV"] for f in range(25, 59) if f in filas_grid)

    # Factor de riesgo SIN diagnóstico: registro incompleto (no debería pasar). Un FR
    # (violencia/suicidio, filas 15-23) sin ningún dx de trastorno mental (25-58) es
    # población que cuenta en la fila 13 pero no en la 24 -> se REPORTA (no cambia
    # ningún número), para completar la ficha. Va a Revisar_Clinico.
    runs_fr = set()
    for f in range(15, 24):
        if f in tributarios:
            runs_fr |= set(tributarios[f]["Número"])
    fr_sin_dx = runs_fr - runs_24
    for run in fr_sin_dx:
        revisar.append({"RUN": run, "Motivo": "Factor de riesgo SIN diagnóstico",
                        "Fila_P6": "15-23", "Detalle": "tiene FR (violencia/suicidio) pero "
                        "ningún dx de trastorno mental; completar la ficha", "Valor_crudo": "",
                        "Categoria": "Clinico"})
    if fr_sin_dx:
        log(f"[sp_p6] {len(fr_sin_dx)} persona(s) con FACTOR DE RIESGO pero SIN "
            "diagnóstico de trastorno mental (registro incompleto) -> Revisar_Clinico.")

    # Fila 13: SUMA LITERAL de las filas 15-24 (hereda el doble conteo de los FR, §5.2).
    fila13 = {}
    claves = ["Ambos", "Hombres", "Mujeres"] + [f"{l} H" for l in LBL] + [f"{l} M" for l in LBL] + \
             ["AN", "AO", "AP_H", "AQ_M", "AR_H", "AS_M", "AT", "AU", "AV", "AW", "AX"]
    for k in claves:
        fila13[k] = sum(filas_grid[f].get(k, 0) for f in range(15, 25) if f in filas_grid)
    filas_grid[13] = fila13

    if not (1300 <= filas_grid[24]["Ambos"] <= 1500):
        log(f"[sp_p6] fila 24 = {filas_grid[24]['Ambos']} está fuera del rango histórico "
            "~1300-1500 (sanity check, no es un error automático).")
    suma_25_58 = sum(filas_grid[f]["Ambos"] for f in range(25, 59) if f in filas_grid)
    if suma_25_58 <= filas_grid[24]["Ambos"]:
        log(f"[sp_p6] suma(25..58)={suma_25_58} no es MAYOR que fila24={filas_grid[24]['Ambos']} "
            "(se esperaba > por comorbilidad; revisar si de verdad no hay nadie con 2+ dx).")

    # §5.2: fila 13 vs el DISTINCTCOUNT(RUN) real de "tiene FR o dx" (15-23 UNION 25-58,
    # que es justamente runs_24) -> la diferencia es la magnitud del doble conteo de FR.
    # Va TAMBIÉN a Revisar_Clinico (no solo al log): es una magnitud, no un RUN puntual.
    runs_fr = set().union(*[set(tributarios[f]["Número"]) for f in range(15, 24) if f in tributarios])
    distinct_fr_o_dx = len(runs_fr | runs_24)
    log(f"[sp_p6] fila 13 = {filas_grid[13]['Ambos']} (suma FR+dx, con doble conteo) vs "
        f"{distinct_fr_o_dx} personas DISTINTAS con algún factor de riesgo o dx (diferencia = "
        "doble conteo de factores de riesgo, esperado).")
    revisar.append({"RUN": "", "Motivo": "Fila 13 vs distinct", "Fila_P6": 13,
                    "Detalle": f"fila13={filas_grid[13]['Ambos']} vs {distinct_fr_o_dx} personas distintas",
                    "Valor_crudo": f"magnitud doble conteo FR = {filas_grid[13]['Ambos'] - distinct_fr_o_dx}",
                    "Categoria": "Clinico"})

    # §4.3: egreso multi-dx divergente (calculado en programas.poblacion) -> se refleja
    # también acá para que el revisor clínico vea TODO en un solo lugar; el detalle
    # completo (con estados) sigue viviendo en la hoja Egreso_Divergencias.
    div = P.attrs.get("egreso_divergencias")
    if div is not None and len(div):
        for _, d in div.iterrows():
            revisar.append({"RUN": d["RUN"], "Motivo": "Egreso multi-dx divergente",
                            "Fila_P6": "", "Detalle": d["Diagnostico"],
                            "Valor_crudo": f"port={d['Valor_port']} / PowerBI={d['Valor_PowerBI']}",
                            "Categoria": "Clinico"})

    grid_df = pd.DataFrame([{**{"Fila": f, "Concepto": ROW_LABELS.get(f, ("", ""))[0],
                                "Subconcepto": ROW_LABELS.get(f, ("", ""))[1]}, **g}
                           for f, g in sorted(filas_grid.items())])
    detalle_df = pd.DataFrame(detalle_rows)
    bloques_df = _bloques_pegables()

    cols_revisar = ["RUN", "Motivo", "Fila_P6", "Detalle", "Valor_crudo"]
    revisar_df = pd.DataFrame(revisar, columns=cols_revisar + ["Categoria"])
    revisar_admin = revisar_df[revisar_df["Categoria"] == "Administrativo"][cols_revisar].reset_index(drop=True)
    revisar_clin = revisar_df[revisar_df["Categoria"] == "Clinico"][cols_revisar].reset_index(drop=True)

    log(f"[sp_p6] P6·A.1 armado" + (f" (mes {mes[0]}-{mes[1]:02d})" if mes else ""))
    if len(revisar_admin):
        conteo = revisar_admin["Motivo"].value_counts()
        log("[sp_p6] Revisar_Administrativo: " + " · ".join(f"{m}={n}" for m, n in conteo.items()))
    if len(revisar_clin):
        conteo = revisar_clin["Motivo"].value_counts()
        log("[sp_p6] Revisar_Clinico: " + " · ".join(f"{m}={n}" for m, n in conteo.items()))

    return {"grid": grid_df, "detalle": detalle_df,
           "revisar_administrativo": revisar_admin, "revisar_clinico": revisar_clin,
           "bloques": bloques_df, "mes": mes}


# ======================================================================
# §5.6 — Bloques pegables (rectángulos maximales sin celdas bloqueadas)
# ======================================================================
def _firma_columnas(fila):
    """Set de columnas EFECTIVAMENTE abiertas para `fila` (bandas + demo),
    para agrupar filas consecutivas con la MISMA firma en un solo bloque."""
    lo, hi = MASCARA_BANDA[fila]
    bandas = frozenset(range(lo, hi + 1))
    demo = frozenset(_mascara_demo(fila))
    return (bandas, demo)


def _bloques_pegables():
    """Agrupa filas consecutivas (13, 15-58, saltando 14 y 21 -que no tiene
    nada que pegar-) con firma de columnas idéntica -> rectángulos maximales.
    Cada bloque queda rotulado con su rango destino en la plantilla real."""
    filas = [13] + list(range(15, 21)) + list(range(22, 59))   # 21 se omite (siempre 0, sin bloque)
    bloques = []
    inicio = filas[0]
    firma_actual = _firma_columnas(inicio)
    prev = inicio
    for f in filas[1:]:
        firma = _firma_columnas(f)
        contiguo = (f == prev + 1) or (prev == 20 and f == 22)  # el hueco de la fila 21 no rompe bloque
        if firma == firma_actual and contiguo:
            prev = f
            continue
        bloques.append((inicio, prev, firma_actual))
        inicio, firma_actual, prev = f, firma, f
    bloques.append((inicio, prev, firma_actual))

    rows = []
    for f_ini, f_fin, (bandas, demo) in bloques:
        lo, hi = min(bandas), max(bandas)
        col_ini = "F" if lo == 0 else _col_banda_letra(lo, hombre=True)
        col_fin = _col_banda_letra(hi, hombre=False)
        rows.append({
            "Filas": f"{f_ini}" if f_ini == f_fin else f"{f_ini}-{f_fin}",
            "Pegar_desde": f"{col_ini}{f_ini}",
            "Pegar_hasta_bandas": f"{col_fin}{f_fin}",
            "Demografia_abierta": ",".join(sorted(demo)) or "(ninguna)",
        })
    return pd.DataFrame(rows)


def _col_banda_letra(idx, hombre):
    from openpyxl.utils import get_column_letter
    return get_column_letter(6 + idx * 2 + (0 if hombre else 1))   # F=6


def escribir(P, resultado, salida):
    """Escribe PSM_Poblacion (+ Egreso_Divergencias, colapsable por diagnóstico) y
    las hojas del P6 (P6_A1, P6_Detalle, Revisar_Administrativo, Revisar_Clinico,
    P6_Bloques) en un solo .xlsx. Las dos de revisión se emiten SIEMPRE, aunque
    vengan vacías (§5.5: su ausencia no debe confundirse con "no había nada que
    revisar")."""
    from programas.poblacion import escribir_divergencias
    with pd.ExcelWriter(salida, engine="openpyxl") as xw:
        P.to_excel(xw, index=False, sheet_name="PSM_Poblacion")
        escribir_divergencias(xw.book, P.attrs.get("egreso_divergencias"))
        resultado["grid"].to_excel(xw, index=False, sheet_name="P6_A1")
        resultado["detalle"].to_excel(xw, index=False, sheet_name="P6_Detalle")
        resultado["revisar_administrativo"].to_excel(xw, index=False, sheet_name="Revisar_Administrativo")
        resultado["revisar_clinico"].to_excel(xw, index=False, sheet_name="Revisar_Clinico")
        resultado["bloques"].to_excel(xw, index=False, sheet_name="P6_Bloques")
    return str(salida)
