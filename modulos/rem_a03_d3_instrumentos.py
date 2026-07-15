#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# Version: 1.3.3
#
# This program is free software: you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by the
# Free Software Foundation, either version 3 of the License, or (at your
# option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTABILITY
# or FITNESS FOR A PARTICULAR PURPOSE. See the GNU General Public License
# for more details: <https://www.gnu.org/licenses/>.
# ==========================================================================
"""
Instrumentos de monitoreo del PSM (REM A03 sección D.3) — PSC / PSC-Y / GHQ-12.

Procesa el export RAYEN de un formulario de screening (aplicado al ingreso y
egreso del Programa de Salud Mental) y, por cada aplicación, reporta:
  - Instrumento (PSC / PSC-Y / GHQ-12) — DETECTADO por contenido (no por nombre
    de archivo: RAYEN baja todo como 'Formulario_Rayen.xlsx').
  - Momento (Ingreso / Egreso), del campo '1.- Estado'.
  - Puntaje (el que suma RAYEN).
  - Resultado AUTOMÁTICO (columna RESULTADO de RAYEN, tal cual).
  - Resultado CALCULADO (recalcula el puntaje con los cortes REM/DISAM, abajo).
  - Discrepancia (SI el automático != el calculado).
  - Estamento de quien aplicó (solo IRIS: la columna mal rotulada 'INSTRUMENTO').

Dos formatos (perfiles), igual que el A05: IRIS y Administrativo. Se autodetectan.

CORE: 1 fila por aplicación con ambos resultados + estamento (IRIS directo, o
Administrativo vía lookup 'Utilización de Cupos' → programas/estamentos.py).
PENDIENTE (v2): conteos agregados por rango etario para pegar en el SA.

Contexto completo: docs/CONTEXTO_REM_A03_D3_INSTRUMENTOS.md
"""

from programas.rem_utils import (
    OPENPYXL_OK, OPENPYXL_ERR, openpyxl,
    ArchivoInvalido,
    norm, solo_entero, buscar_col, num_pregunta,
    encontrar_fila_encabezado, edad_anios,
)
from programas.estamentos import (cargar_estamentos, buscar_estamento,
                                  faltantes, estamentos_conocidos, aplicar_resoluciones)

# ╔═══════════════════════════════════════════════════════════════════╗
# ║  CORTES (REM SA 2026 = los que pasó DISAM). Reclasifican el puntaje.║
# ╚═══════════════════════════════════════════════════════════════════╝
# Resultado para PSC/PSC-Y bajo el corte REM (<33): NO es una de las 3 bandas de
# riesgo del REM, pero es un resultado VÁLIDO y el MÁS común (la mayoría de los
# tamizajes salen negativos). Antes se perdía como None (celda vacía, sin contar,
# sin comparar) -> ahora se etiqueta, se cuenta y se compara. Renómbralo acá si
# DISAM/RAYEN usa otra palabra (ej. "Negativo", "Sin sospecha").
LABEL_SIN_RIESGO = "Sin riesgo"

def clasificar_psc(p):
    """PSC / PSC-Y (cortes idénticos). <33 = bajo el corte REM -> 'Sin riesgo'
    (resultado válido, no es banda de riesgo). None SOLO si no hay puntaje."""
    if p is None: return None
    if p < 33:  return LABEL_SIN_RIESGO
    if p <= 63: return "Bajo"
    if p <= 69: return "Medio"
    return "Alto"   # >= 70

def clasificar_ghq12(p):
    if p is None: return None
    if 0 <= p <= 4:  return "Bajo"
    if 5 <= p <= 6:  return "Medio"
    if 7 <= p <= 12: return "Alto"
    return None

# ── RESULTADO de RAYEN -> banda canónica ──────────────────────────────
# OJO: RAYEN NO usa un vocabulario único. PSC / PSC-Y dicen 'Bajo/Medio/Alto' (y
# blanco bajo el corte), pero GHQ-12 (Goldberg) usa frases clínicas:
#   'Ausencia de psicopatología'              -> Bajo   (0-4)
#   'Sospecha de psicopatología subumbral'    -> Medio  (5-6)
#   'Indicativos de presencia de psicopatología' -> Alto (7-12)
# Verificado contra exports reales (jul-2026). Sin canonizar, TODO Goldberg saldría
# como discrepancia FALSA (la redacción difiere aunque la banda coincida). Orden =
# keyword más específico primero. Extiende esto si aparece otra redacción.
_MAP_RESULTADO = [
    (["SUBUMBRAL"],                 "Medio"),        # GHQ: sospecha subumbral
    (["AUSENCIA"],                  "Bajo"),         # GHQ: ausencia de psicopatología
    (["INDICATIVOS", "PRESENCIA"],  "Alto"),         # GHQ: indicativos de presencia
    (["ALTO"],                      "Alto"),         # PSC / PSC-Y
    (["MEDIO"],                     "Medio"),
    (["BAJO"],                      "Bajo"),
    (["SIN RIESGO", "NEGATIVO"],    LABEL_SIN_RIESGO),
]

def canon_resultado(v):
    """Lleva el RESULTADO crudo de RAYEN a la banda canónica
    {Sin riesgo, Bajo, Medio, Alto}. None si viene en blanco (PSC/PSC-Y bajo el
    corte) o si la redacción no se reconoce (para no inventar discrepancias)."""
    n = norm(v)
    if not n:
        return None
    for keys, banda in _MAP_RESULTADO:
        if any(k in n for k in keys):
            return banda
    return None

# Registro de instrumentos. `patrones` = substrings (normalizados) para detectar
# el instrumento en el nombre del formulario. Orden de chequeo: PSC-Y antes que
# PSC (porque 'PSC' es substring), luego GHQ, luego PSC.
INSTRUMENTOS = {
    "PSC-Y":  {"nombre": "PSC-Y",  "clasificar": clasificar_psc,   "patrones": ["PSC-Y", "ADOLESCENTES"]},
    "GHQ-12": {"nombre": "GHQ-12", "clasificar": clasificar_ghq12, "patrones": ["GOLDBERG", "GHQ"]},
    # OJO: "PSC para padres" = el MISMO test que "PSC 5 a 9" (mismo instrumento,
    # distinto nombre según la fuente). En RAYEN baja como 'Cuestionario para
    # padres PSC'. NO es un instrumento aparte.
    "PSC":    {"nombre": "PSC",    "clasificar": clasificar_psc,   "patrones": ["PADRES PSC", "CUESTIONARIO PARA PADRES", "PADRES"]},
}
_ORDEN_DETECCION = ["PSC-Y", "GHQ-12", "PSC"]

# ── Firmas de formato (RAYEN genérico; igual que el A05) ──
ANCLA_IRIS   = ["AÑO", "APLICACION", "FORMULARIO"]
ANCLA_ADMIN  = ["EDAD", "REGISTRO", "FORMULARIO"]
ADMIN_BANNER = "SERVICIO DE SALUD"
ADMIN_MARKERS = ["NUMERO DE FICHAS", "EDAD DE REGISTRO FORMULARIO", "FECHA FORMULARIO"]
MAX_FILAS_HEADER = 60

# Detección de columnas (nombres de ambos formatos)
RUT_TOKENS_IRIS  = ["NUMERO", "IDENTIFICACION"]
RUT_EXACTO_ADMIN = "RUT"
EDAD_TOKENS_IRIS  = ["AÑO", "APLICACION", "FORMULARIO"]
EDAD_TOKENS_ADMIN = ["EDAD", "REGISTRO", "FORMULARIO"]

NOMBRE_HOJA_SALIDA = "A03_D3_Instrumentos"


# ── Detección de formato / encabezado / instrumento ───────────────────
def detectar_formato(ws):
    """'iris' | 'administrativo' | 'desconocido' (mismas firmas RAYEN que A05)."""
    tope = min(ws.max_row, MAX_FILAS_HEADER)
    ancla = [norm(t) for t in ANCLA_IRIS]
    for r in range(1, tope + 1):
        vals = [norm(c.value) for c in ws[r]]
        if all(any(tok in v for v in vals) for tok in ancla):
            return "iris"
    if norm(ws.cell(row=1, column=1).value) == ADMIN_BANNER:
        return "administrativo"
    for r in range(1, tope + 1):
        vals = [norm(c.value) for c in ws[r]]
        for m in ADMIN_MARKERS:
            if any(norm(m) in v for v in vals):
                return "administrativo"
    return "desconocido"


def _fila_encabezado(ws, formato):
    if formato == "administrativo":
        return encontrar_fila_encabezado(ws, ANCLA_ADMIN, usar_blanco_en_a=False,
                                         n_hardcode=8, max_filas=MAX_FILAS_HEADER)
    return encontrar_fila_encabezado(ws, ANCLA_IRIS, usar_blanco_en_a=True,
                                     n_hardcode=0, max_filas=MAX_FILAS_HEADER)


def detectar_instrumento(ws, formato, header_idx, headers_norm):
    """Detecta PSC / PSC-Y / GHQ-12 por CONTENIDO (nunca por nombre de archivo):
    en IRIS por la columna 'FORMULARIO'; en Administrativo por el banner
    'Formulario: ...'. Devuelve la clave o None si no se reconoce."""
    texto = ""
    if formato == "iris":
        col = buscar_col(headers_norm, exacto="FORMULARIO")
        if col:
            for r in range(header_idx + 1, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v not in (None, ""):
                    texto = norm(v); break
    else:
        for r in range(1, header_idx):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and "FORMULARIO" in norm(v) and ":" in str(v):
                    texto = norm(v); break
            if texto:
                break
    for key in _ORDEN_DETECCION:
        if any(norm(p) in texto for p in INSTRUMENTOS[key]["patrones"]):
            return key
    return None


# ── Apertura + validación ─────────────────────────────────────────────
def abrir_validado(entrada):
    """Carga el workbook y valida que sea un export de instrumentos RAYEN
    (formato reconocido + columnas PUNTAJE y RESULTADO). Devuelve
    (wb, ws, formato, header_idx). Levanta ImportError/ArchivoInvalido."""
    if not OPENPYXL_OK:
        raise ImportError(
            "Falta 'openpyxl'. Instálala con:  pip install openpyxl\n"
            f"(detalle: {OPENPYXL_ERR})")
    wb = openpyxl.load_workbook(entrada)
    ws = wb.active
    formato = detectar_formato(ws)
    if formato == "desconocido":
        raise ArchivoInvalido(
            "desconocido",
            "No reconozco este archivo como un export RAYEN (ni IRIS ni "
            "Administrativo). Descárgalo de nuevo sin modificarlo.")
    header_idx, _ = _fila_encabezado(ws, formato)
    headers_norm = [norm(ws.cell(row=header_idx, column=c).value)
                    for c in range(1, ws.max_column + 1)]
    if not (buscar_col(headers_norm, tokens=["PUNTAJE"]) and
            buscar_col(headers_norm, tokens=["RESULTADO"])):
        raise ArchivoInvalido(
            "no_instrumento",
            "Este export no parece un formulario de instrumento (PSC/PSC-Y/"
            "Goldberg): no encontré columnas PUNTAJE y RESULTADO.\n"
            "¿Quizás cargaste el formulario de 'Control de Salud Mental'? "
            "Ese va en el módulo de egresos/ingresos.")
    return wb, ws, formato, header_idx


# ── Núcleo ────────────────────────────────────────────────────────────
def procesar(entrada, salida, instrumento=None, estamentos=None,
             resolver_estamento=None, log=print):
    """Detecta formato + instrumento, arma la hoja de instrumentos y guarda.
    `instrumento` opcional fuerza el instrumento (si None, autodetecta).
    `estamentos` opcional (ruta al reporte 'Utilización de Cupos' o dict ya
    cargado): en Administrativo rellena el estamento por nombre de funcionario.
    `resolver_estamento` opcional: callback (faltantes, opciones) -> {nombre:
    estamento|None} para resolver a mano los que no matchean (None = IGNORAR).
    Devuelve un resumen."""
    wb, ws, formato, header_idx = abrir_validado(entrada)

    if instrumento is None:
        headers_norm0 = [norm(ws.cell(row=header_idx, column=c).value)
                         for c in range(1, ws.max_column + 1)]
        instrumento = detectar_instrumento(ws, formato, header_idx, headers_norm0)
    if instrumento not in INSTRUMENTOS:
        raise ArchivoInvalido(
            "instrumento_desconocido",
            "No pude detectar qué instrumento es (PSC / PSC-Y / GHQ-12). "
            "Especifícalo a mano al procesar.")
    inst = INSTRUMENTOS[instrumento]
    clasificar = inst["clasificar"]
    log(f"[corte] formato={formato} | encabezado fila {header_idx} | instrumento={inst['nombre']}")

    ncols = ws.max_column
    headers = [ws.cell(row=header_idx, column=c).value for c in range(1, ncols + 1)]
    hn = [norm(h) for h in headers]

    rut_col   = (buscar_col(hn, tokens=[norm(t) for t in RUT_TOKENS_IRIS])
                 or buscar_col(hn, exacto=RUT_EXACTO_ADMIN))
    edad_col  = (buscar_col(hn, tokens=[norm(t) for t in EDAD_TOKENS_IRIS])
                 or buscar_col(hn, tokens=[norm(t) for t in EDAD_TOKENS_ADMIN]))
    sexo_col  = buscar_col(hn, exacto="SEXO")
    punt_col  = buscar_col(hn, tokens=["PUNTAJE"])
    resu_col  = buscar_col(hn, tokens=["RESULTADO"])
    func_col  = buscar_col(hn, exacto="FUNCIONARIO")
    estam_col = buscar_col(hn, exacto="INSTRUMENTO")   # mal rotulada = estamento (solo IRIS)
    momento_col = next((i + 1 for i, h in enumerate(headers)
                        if num_pregunta(h) == 1 and norm(h).endswith("ESTADO")), None)
    log(f"[cols] RUT=col{rut_col} Edad=col{edad_col} Sexo=col{sexo_col} "
        f"Puntaje=col{punt_col} Resultado=col{resu_col} Momento=col{momento_col} "
        f"Estamento=col{estam_col}")
    tabla_estam = None
    if estamentos is not None:
        tabla_estam = (estamentos if isinstance(estamentos, dict)
                       else cargar_estamentos(estamentos, log=log)[0])
    if formato != "iris" and not estam_col and tabla_estam is None:
        log("[estamento] AUSENTE en Administrativo y sin tabla de estamentos -> "
            "columna vacía. (Carga el reporte 'Utilización de Cupos' para rellenarlo.)")

    filas = []
    no_reconocidos = {}   # RESULTADO de RAYEN con redacción no mapeada
    for r in range(header_idx + 1, ws.max_row + 1):
        fila = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        puntaje = solo_entero(fila[punt_col - 1]) if punt_col else None
        resu_rayen = fila[resu_col - 1] if resu_col else None
        # fila vacía / sin aplicación -> saltar
        if puntaje is None and (resu_rayen in (None, "")):
            continue
        rut  = fila[rut_col - 1]  if rut_col  else ""
        edad = edad_anios(fila[edad_col - 1]) if edad_col else None
        sexo = fila[sexo_col - 1] if sexo_col else ""
        momento = fila[momento_col - 1] if momento_col else ""
        funcionario = fila[func_col - 1] if func_col else ""
        estamento = fila[estam_col - 1] if estam_col else ""  # IRIS: en col; Admin: se rellena abajo
        resu_disam = clasificar(puntaje)
        banda_rayen = canon_resultado(resu_rayen)   # canoniza la redacción de RAYEN
        if resu_rayen not in (None, "") and banda_rayen is None:
            k = str(resu_rayen)
            no_reconocidos[k] = no_reconocidos.get(k, 0) + 1
        # Discrepancia = las BANDAS difieren (no la redacción). Solo si ambas existen.
        disc = ""
        if banda_rayen is not None and resu_disam is not None:
            disc = "SI" if banda_rayen != resu_disam else ""
        filas.append({
            "rut": rut, "edad": edad, "sexo": sexo, "instrumento": inst["nombre"],
            "momento": momento, "puntaje": puntaje,
            "resu_rayen": resu_rayen if resu_rayen is not None else "",
            "banda_rayen": banda_rayen if banda_rayen is not None else "",
            "resu_disam": resu_disam if resu_disam is not None else "",
            "disc": disc, "estamento": estamento, "funcionario": funcionario,
            "fila": r,
        })

    # ── Estamento (Administrativo): rellenar por nombre desde el lookup ──
    # Failsafe: los funcionarios sin match se resuelven a mano (o se IGNORAN, p.ej.
    # externos que prestan servicios transitorios) vía el callback resolver_estamento.
    n_estam = 0
    sin_estam = []
    if tabla_estam is not None:
        pend = [e["funcionario"] for e in filas if not e["estamento"] and e["funcionario"]]
        falt = faltantes(pend, tabla_estam)
        if falt:
            log(f"[estamento] sin match en la tabla: {len(falt)} "
                f"({', '.join(falt[:5])}{'…' if len(falt) > 5 else ''})")
            if resolver_estamento is not None:
                resol = resolver_estamento(falt, estamentos_conocidos(tabla_estam)) or {}
                if resol:
                    aplicar_resoluciones(tabla_estam, resol)
                    n_ig = sum(1 for v in resol.values() if not v)
                    log(f"[estamento] resueltos a mano: {sum(1 for v in resol.values() if v)}"
                        + (f" | ignorados: {n_ig}" if n_ig else ""))
        for e in filas:
            if not e["estamento"] and e["funcionario"]:
                v = buscar_estamento(e["funcionario"], tabla_estam)
                if v:
                    e["estamento"] = v; n_estam += 1
        sin_estam = sorted({e["funcionario"] for e in filas
                            if e["funcionario"] and not e["estamento"]})
        log(f"[estamento] rellenados: {n_estam}"
            + (f" | quedan sin estamento: {len(sin_estam)}" if sin_estam else ""))

    # ── hoja de salida ──
    if NOMBRE_HOJA_SALIDA in wb.sheetnames:
        del wb[NOMBRE_HOJA_SALIDA]
    ws2 = wb.create_sheet(NOMBRE_HOJA_SALIDA)
    cols = ["RUT", "Edad", "Sexo", "Instrumento", "Momento", "Puntaje",
            "Resultado_RAYEN", "Banda_RAYEN", "Resultado_DISAM", "Discrepancia",
            "Estamento", "Funcionario", "Fila_Origen"]
    ws2.append(cols)
    orden_mom = {"INGRESO": 0, "EGRESO": 1}
    filas.sort(key=lambda e: (orden_mom.get(norm(e["momento"]), 9),
                              str(e["resu_disam"]), str(e["rut"])))
    for e in filas:
        ws2.append([e["rut"], e["edad"], e["sexo"], e["instrumento"], e["momento"],
                    e["puntaje"], e["resu_rayen"], e["banda_rayen"], e["resu_disam"],
                    e["disc"], e["estamento"], e["funcionario"], e["fila"]])

    from openpyxl.styles import Font
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"
    if ws2.max_row >= 1:
        ws2.auto_filter.ref = ws2.dimensions
    from openpyxl.utils import get_column_letter
    anchos = [13, 6, 8, 12, 10, 8, 30, 12, 16, 12, 22, 22, 11]
    for i, w in enumerate(anchos, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(salida)
    log(f"[ok] guardado: {salida}  (hoja '{NOMBRE_HOJA_SALIDA}')")

    n_disc = sum(1 for e in filas if e["disc"] == "SI")
    por_momento = {}
    por_resultado = {}
    for e in filas:
        k = str(e["momento"]) or "(sin momento)"
        por_momento[k] = por_momento.get(k, 0) + 1
        rk = str(e["resu_disam"]) or "(sin puntaje)"
        por_resultado[rk] = por_resultado.get(rk, 0) + 1
    log(f"[resumen] aplicaciones: {len(filas)} | discrepancias RAYEN vs DISAM: {n_disc}")
    for k, v in por_momento.items():
        log(f"          {k}: {v}")
    log("          resultado DISAM -> " +
        " · ".join(f"{k}: {v}" for k, v in por_resultado.items()))
    if no_reconocidos:
        log("[aviso] RESULTADO de RAYEN con redacción NO reconocida (no se comparó "
            "con DISAM; agrégala a _MAP_RESULTADO): "
            + " · ".join(f"{k!r}: {v}" for k, v in no_reconocidos.items()))

    return {
        "salida": str(salida), "instrumento": inst["nombre"], "formato": formato,
        "total": len(filas), "discrepancias": n_disc, "por_momento": por_momento,
        "por_resultado": por_resultado, "hoja": NOMBRE_HOJA_SALIDA,
        "estam_rellenados": n_estam, "estam_faltan": len(sin_estam),
    }


# ── Descriptor para el registro de tareas (lo consumirá autorem.py) ──
# OJO: es un REPORTE distinto al 'Control de Salud Mental' (otros perfiles y
# autodetección propia). La integración a la GUI/dispatcher (con el paso de
# confirmar el instrumento detectado) es el próximo paso.
TAREA = {
    "id": "a03_d3_instrumentos",
    "nombre": "A03 D.3 · Instrumentos (PSC/PSC-Y/GHQ-12)",
    "correr": procesar,                  # (entrada, salida, instrumento=None, log) -> resumen
    "hoja": NOMBRE_HOJA_SALIDA,
}
