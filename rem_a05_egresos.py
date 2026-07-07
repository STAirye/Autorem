#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================================
# This code was generated with the assistance of Claude Opus 4.8 (Anthropic).
# The human author reviewed, modified, and integrated the code.
#
# Author: Simón Tobar — CESFAM Dr. Luis Ferrada Urzúa (APS, SSMC)
# Copyright (C) 2026 Simón Tobar
# SPDX-License-Identifier: GPL-3.0-or-later
# Version: 1.4
#
# This program is free software: you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by the
# Free Software Foundation, either version 3 of the License, or (at your
# option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTABILITY
# or FITNESS FOR A PARTICULAR PURPOSE. See the GNU General Public License
# for more details: <https://www.gnu.org/licenses/>.
# ==========================================================================
"""
Limpieza y marcado de egresos de Salud Mental (export RAYEN) para REM A05.

  1. Recorta filas de cabecera (banner + filtros).
  2. Convierte 'AÑO APLICACIÓN FORMULARIO' a número (= edad al llenado).
  3. Por cada egreso (Alta / Traslado / Otras Causas) agrega columnas con la
     patología y el subtipo. Reconoce el diagnóstico aunque el subtipo esté a
     la izq o der del ESTADO.
  4. Escribe todo en una hoja nueva 'A05_Egresos' (la original queda intacta).

Este módulo es HEADLESS (sin ventana): expone la lógica A05-egresos. La GUI y el
selector de tareas viven en autorem.py, que registra este módulo vía su
descriptor TAREA (al final del archivo).

USO como librería:
    from rem_a05_egresos import procesar
    procesar(entrada_xlsx, salida_xlsx, log=print)

NUEVO en 1.4:
  - GUI/CLI extraídas a autorem.py (dispatcher con registro de tareas). Este
    archivo queda solo con la lógica; expone TAREA/REPORTE para el registro.

De 1.3:
  - Modularización: las utilidades genéricas viven en rem_utils.py.
De 1.2:
  - Detector de "archivo equivocado": distingue el export ADMINISTRATIVO del
    de IRIS, y avisa si el archivo no tiene nada que ver.
"""

import re
from pathlib import Path

# Base común del proyecto (ver rem_utils.py). Debe estar junto a este archivo.
from rem_utils import (
    OPENPYXL_OK, OPENPYXL_ERR, openpyxl,
    ArchivoInvalido,
    norm, solo_entero, buscar_col, num_pregunta,
    encontrar_fila_encabezado,
)

# ╔═══════════════════════════════════════════════════════════════════╗
# ║  ZONA DE CONFIGURACIÓN CLÍNICA  — editar aquí si cambia el form     ║
# ╚═══════════════════════════════════════════════════════════════════╝

# Egresos a marcar. Alta y Traslado van a estadística; Otras Causas se flaggea
# para revisión MANUAL (decidir abandono vs clínica caso a caso).
BUSQUEDAS = {
    "Alta":        ["EGRESO", "ALTA"],
    "Traslado":    ["EGRESO", "TRASLADO"],
    "OtrasCausas": ["EGRESO", "OTRAS", "CAUSAS"],
}

# Diagnósticos QUE TIENEN SUBTIPO.
#   clave = N.- de la pregunta del diagnóstico
#   valor = N.- de la pregunta de su columna de subtipo
DIAGNOSTICOS_CON_SUBTIPO = {
    4:  6,    # Violencia   -> 6.- TIPO DE VIOLENCIA
    11: 12,   # Suicidio    -> 12.- TIPO DE SUICIDIO
    18: 20,   # Depresión   -> 20.- TIPO DE DEPRESIÓN
    41: 43,   # Ansiedad    -> 43.- TIPO DE TRASTORNO DE ANSIEDAD
    44: 45,   # Alzheimer   -> 45.- ETAPA
}

# Remapeo de bloques deprecated -> (patología, subtipo) forzados.
REMAP_DIAGNOSTICO = {
    9: ("Violencia", "Sexual"),   # Abuso Sexual (deprecated) -> Violencia > Víctima > Sexual
}

# Nombre de patología: False = header crudo (feo pero funciona).
# True (día de ocio) = usa limpiar_patologia() + OVERRIDE_PATOLOGIA.
LIMPIAR_NOMBRE_PATOLOGIA = False
OVERRIDE_PATOLOGIA = {
    # 65: "Otros trastornos del comportamiento",
}


# ── Columnas para armar la tabla A05 (se copian al final del archivo) ──
# OJO QUIRK RAYEN: el header 'AÑO APLICACIÓN FORMULARIO' NO trae el año; trae la
# EDAD a la fecha de LLENADO del formulario (lo que A05 necesita). En cambio
# 'EDAD PACIENTE' es la edad a la fecha de DESCARGA del reporte -> se ignora.
RUT_TOKENS  = ["NUMERO", "IDENTIFICACION"]   # 'NUMERO TIPO IDENTIFICACION' -> 11111111-1
EDAD_FORM_TOKENS = ["AÑO", "APLICACION", "FORMULARIO"]   # = edad al llenado
SEXO_HEADER = "SEXO"

# Salida en hoja nueva (la original queda intacta). Formato LARGO: 1 fila/egreso.
NOMBRE_HOJA_SALIDA = "A05_Egresos"
TIPO_LABEL = {"Alta": "Alta", "Traslado": "Traslado", "OtrasCausas": "Otras Causas"}

# ── Caracterización demográfica para A05 (devuelve "SI"/"" salvo Trans) ──
# Validado (jul-2026) contra los valores DISTINTOS reales de 'ALERTAS
# ADMINISTRATIVAS':  PRAIS · Fonasa Libre Elección · Jubilación de Vejez ·
# Atención Preferente (Mayor 60 / Cuidador / Discapacidad) ·
# SPE ex Mejor Niñez- Ambulatorio · Subsistema Seguridades y Oportunidades ·
# SUF · MIGRANTE · SENAME Justicia Juvenil.
# La celda trae VARIOS valores separados por ';' -> el match por substring sirve.
# NOTA: 'Gestante' NO existe en el export 'Control de Salud Mental' (jul-2026),
#   por eso no está aquí. Si algún día aparece la fuente, agregar la fila:
#     "Gestante": (["<header>"], ["GESTANTE", "EMBARAZ"]),
DEMOGRAFIA = {
    "Madre_menor5":        (["USTED", "MADRE"],       ["SI"]),                                   # pregunta 1 del form
    "Pueblos_Originarios": (["PUEBLO", "ORIGINARIO"], "_no_vacio"),
    "SENAME":              (["ALERTAS"],              ["SENAME", "REINSERCION"]),                # -> 'SENAME Justicia Juvenil'
    "Proteccion_Ninez":    (["ALERTAS"],              ["MEJOR NINEZ", "PROTECCION ESPECIAL"]),   # -> 'SPE ex Mejor Niñez'
    "Migrante":            (["ALERTAS"],              ["MIGRANTE"]),                             # alerta explícita (antes: NACIONALIDAD _no_chileno)
}
COL_GENERO_TOKENS = ["GENERO"]   # Trans: copia el valor de GÉNERO si contiene "TRANS"
NEGATIVOS_DEMO = {"", "NO", "NINGUNO", "NINGUNA", "NO APLICA", "SIN INFORMACION", "N/A"}

# Avisar egreso por Alta sin subtipo (solo en diagnósticos que SÍ tienen subtipo).
AVISAR_ALTA_SIN_SUBTIPO = True

# ╔═══════════════════════════════════════════════════════════════════╗
# ║  CONFIG TÉCNICA (rara vez se toca)                                  ║
# ╚═══════════════════════════════════════════════════════════════════╝
HOJA = None
ANCLA_ENCABEZADO = ["AÑO", "APLICACION", "FORMULARIO"]
USAR_BLANCO_EN_A = True
N_HARDCODE = 16
ANIO_TOKENS = ["AÑO", "APLICACION", "FORMULARIO"]
ANIO_COL_FALLBACK = 11
SEP_MULTI = " | "
MAX_FILAS_BUSQUEDA_HEADER = 60

SUBTIPO_NUMS = set(DIAGNOSTICOS_CON_SUBTIPO.values())

# ── Firma del export ADMINISTRATIVO (para rechazarlo con mensaje claro) ──
# El export administrativo trae banner 'Servicio de Salud / Comuna / ...' y un
# encabezado con estas columnas que el export de IRIS NO tiene.
ADMIN_MARKERS = ["NUMERO DE FICHAS", "EDAD DE REGISTRO FORMULARIO", "FECHA FORMULARIO"]
ADMIN_BANNER  = "SERVICIO DE SALUD"
# ───────────────────────────────────────────────────────────────────


def es_estado(h):
    return norm(h).endswith("ESTADO")


# ── LIMPIEZA DE NOMBRE DE PATOLOGÍA (pendiente, día de ocio) ──────────
def limpiar_patologia(header):
    """'18.- ¿ TIENE  DEPRESIÓN ?' -> 'Depresión'. Solo si LIMPIAR_NOMBRE_PATOLOGIA."""
    n = num_pregunta(header)
    if n in OVERRIDE_PATOLOGIA:
        return OVERRIDE_PATOLOGIA[n]
    s = re.sub(r"^\s*\d+\s*\.\-\s*", "", str(header))
    s = s.replace("¿", "").replace("?", "")
    s = re.sub(r"^\s*(TIENE|ES|SUFRE DE|PRESENTA|PADECE)\s+", "", s, flags=re.I)
    s = re.sub(r"^\s*PACIENTE\s+PRESENTA\s+", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:1].upper() + s[1:].lower() if s else s
# ─────────────────────────────────────────────────────────────────────


def limpiar_subtipo(valor, subtipo_header):
    """Recorta el sustantivo del header: valor 'Depresión Moderada' con header
    '20.- TIPO DE DEPRESIÓN' -> 'Moderada'. Si no hay 'TIPO DE', deja el valor."""
    if valor in (None, ""): return ""
    s = str(valor).strip()
    m = re.search(r"TIPO DE\s+(.+)$", norm(subtipo_header))
    if m:
        noun = m.group(1).strip()
        if norm(s).startswith(noun):
            rec = s[len(noun):].strip(" -,:")
            if rec: s = rec
    return s


def flag_demo(cell, regla):
    n = norm(cell)
    if regla == "_no_vacio":
        return "SI" if n not in NEGATIVOS_DEMO else ""
    if regla == "_no_chileno":
        return "SI" if (n and "CHILE" not in n) else ""
    if isinstance(regla, (list, tuple)):
        return "SI" if any(norm(k) in n for k in regla) else ""
    return ""


def encontrar_diagnostico(headers, c0):
    """Camina a la izquierda del ESTADO saltando subtipos/estados hasta dar
    con la pregunta del diagnóstico. Devuelve índice 0-based o c0 si no hay."""
    d = c0 - 1
    while d >= 0:
        h = headers[d]
        n = num_pregunta(h)
        if es_estado(h) or (n in SUBTIPO_NUMS) or (n is None):
            d -= 1
            continue
        return d
    return c0


# ── VALIDACIÓN DE FORMATO ─────────────────────────────────────────────
def validar_formato(ws):
    """Reconoce el export IRIS ('Control de Salud Mental'). Si no lo es, levanta
    ArchivoInvalido distinguiendo el export administrativo del desconocido.

    Firma IRIS  : una fila-encabezado con el ancla 'AÑO APLICACIÓN FORMULARIO'
                  Y una columna 'NUMERO ... IDENTIFICACION'.
    Firma admin : columnas 'Numero de Fichas' / 'Edad de registro formulario' /
                  'Fecha Formulario', o el banner 'Servicio de Salud' en A1.
    """
    tope = min(ws.max_row, MAX_FILAS_BUSQUEDA_HEADER)
    ancla = [norm(t) for t in ANCLA_ENCABEZADO]
    rut_tok = [norm(t) for t in RUT_TOKENS]

    iris_ancla = False
    iris_rut = False
    for r in range(1, tope + 1):
        vals = [norm(c.value) for c in ws[r]]
        if all(any(tok in v for v in vals) for tok in ancla):
            iris_ancla = True
        if any(all(t in v for t in rut_tok) for v in vals):
            iris_rut = True
    if iris_ancla and iris_rut:
        return  # es IRIS válido

    # No es IRIS. ¿Es el administrativo?
    a1 = norm(ws.cell(row=1, column=1).value)
    admin_hit = (a1 == ADMIN_BANNER)
    if not admin_hit:
        for r in range(1, tope + 1):
            vals = [norm(c.value) for c in ws[r]]
            for m in ADMIN_MARKERS:
                if any(norm(m) in v for v in vals):
                    admin_hit = True
                    break
            if admin_hit:
                break

    if admin_hit:
        raise ArchivoInvalido(
            "administrativo",
            "Este archivo es el REPORTE ADMINISTRATIVO, no el de IRIS.\n\n"
            "Tienen distinto formato y este script solo procesa el de IRIS.\n\n"
            "Cómo obtener el correcto:\n"
            "  IRIS  ->  Formularios RAYEN  ->  'Control de Salud Mental'\n"
            "  ->  descargar el export.\n\n"
            "El archivo correcto trae la columna 'AÑO APLICACIÓN FORMULARIO' "
            "y 'NÚMERO TIPO IDENTIFICACIÓN'; este trae 'Número de Fichas' y "
            "'Fecha Formulario', que son del reporte administrativo."
        )

    raise ArchivoInvalido(
        "desconocido",
        "No reconozco este archivo como un export de 'Control de Salud Mental'.\n\n"
        "No encontré la columna 'AÑO APLICACIÓN FORMULARIO' ni "
        "'NÚMERO TIPO IDENTIFICACIÓN'.\n\n"
        "Asegúrate de:\n"
        "  1. Descargarlo desde IRIS  ->  Formularios RAYEN  ->  'Control de Salud Mental'.\n"
        "  2. No haberlo modificado (no borres filas ni columnas del export)."
    )
# ─────────────────────────────────────────────────────────────────────


def procesar(entrada: Path, salida: Path, log=print):
    """Núcleo del procesamiento. `log` es un callback (por defecto print) para
    que la GUI pueda mostrar el avance en su propia área de texto."""
    if not OPENPYXL_OK:
        raise ImportError(
            "Falta la librería 'openpyxl'. Si corres el .py suelto, instálala con:\n"
            "    pip install openpyxl\n"
            f"(detalle: {OPENPYXL_ERR})"
        )

    wb = openpyxl.load_workbook(entrada)
    ws = wb[HOJA] if HOJA else wb.active

    # Portón: rechazar archivos que no son el export IRIS correcto.
    validar_formato(ws)

    header_idx, modo = encontrar_fila_encabezado(
        ws, ANCLA_ENCABEZADO, USAR_BLANCO_EN_A, N_HARDCODE, MAX_FILAS_BUSQUEDA_HEADER)
    log(f"[corte] modo={modo} | encabezado en fila {header_idx} (la hoja original NO se modifica)")

    ncols = ws.max_column
    headers = [ws.cell(row=header_idx, column=c).value for c in range(1, ncols + 1)]
    headers_norm = [norm(h) for h in headers]
    num2col = {num_pregunta(h): i for i, h in enumerate(headers) if num_pregunta(h) is not None}

    edad_tok = [norm(t) for t in EDAD_FORM_TOKENS]
    edad_col = buscar_col(headers_norm, tokens=edad_tok)
    if edad_col is None or edad_col > ncols:
        edad_col = ANIO_COL_FALLBACK if ANIO_COL_FALLBACK <= ncols else None
    rut_col  = buscar_col(headers_norm, tokens=[norm(t) for t in RUT_TOKENS])
    sexo_col = buscar_col(headers_norm, exacto=SEXO_HEADER)
    log(f"[A05] RUT=col{rut_col} | Edad(llenado)=col{edad_col} | Sexo=col{sexo_col}")

    demo_cols = {flag: (buscar_col(headers_norm, tokens=[norm(t) for t in src]), regla)
                 for flag, (src, regla) in DEMOGRAFIA.items()}
    genero_col = buscar_col(headers_norm, tokens=[norm(t) for t in COL_GENERO_TOKENS])
    log("[demo] " + " ".join(f"{f}=col{c}" for f, (c, _) in demo_cols.items()) +
        f" Trans(Genero)=col{genero_col}")

    busq = {k: [norm(t) for t in v] for k, v in BUSQUEDAS.items()}
    eventos = []   # una fila por egreso

    for r in range(header_idx + 1, ws.max_row + 1):
        fila = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        fila_n = [norm(v) for v in fila]
        rut  = fila[rut_col - 1]  if rut_col  else ""
        edad = solo_entero(fila[edad_col - 1]) if edad_col else None
        sexo = fila[sexo_col - 1] if sexo_col else ""
        demo = {flag: (flag_demo(fila[c - 1], regla) if c else "")
                for flag, (c, regla) in demo_cols.items()}
        trans = ""
        if genero_col:
            g = fila[genero_col - 1]
            if "TRANS" in norm(g):
                trans = str(g)

        for k, toks in busq.items():
            for c0 in range(ncols):
                if not es_estado(headers[c0]):
                    continue
                if not all(t in fila_n[c0] for t in toks):
                    continue
                d = encontrar_diagnostico(headers, c0)
                diag_num = num_pregunta(headers[d])
                if diag_num in REMAP_DIAGNOSTICO:
                    pat, sub = REMAP_DIAGNOSTICO[diag_num]
                else:
                    pat = (limpiar_patologia(headers[d]) if LIMPIAR_NOMBRE_PATOLOGIA
                           else str(headers[d]))
                    sub_num = DIAGNOSTICOS_CON_SUBTIPO.get(diag_num)
                    sub_col0 = num2col.get(sub_num) if sub_num else None
                    sub_val = fila[sub_col0] if sub_col0 is not None else ""
                    sub = limpiar_subtipo(sub_val, headers[sub_col0] if sub_col0 is not None else "")
                falta_sub = ""
                if (AVISAR_ALTA_SIN_SUBTIPO and k == "Alta"
                        and diag_num in DIAGNOSTICOS_CON_SUBTIPO and not str(sub).strip()):
                    falta_sub = "SI"
                ev = {"rut": rut, "edad": edad, "sexo": sexo, "tipo": k,
                      "pat": pat, "sub": sub, "falta_sub": falta_sub,
                      "trans": trans, "fila": r}
                ev.update(demo)
                eventos.append(ev)

    # ── hoja nueva (la original intacta) ──
    if NOMBRE_HOJA_SALIDA in wb.sheetnames:
        del wb[NOMBRE_HOJA_SALIDA]
    ws2 = wb.create_sheet(NOMBRE_HOJA_SALIDA)
    demo_keys = list(DEMOGRAFIA.keys())
    cols = (["RUT", "Edad_Formulario", "Sexo", "Tipo_Egreso", "Patologia", "Subtipo",
             "Falta_Subtipo"] + demo_keys + ["Trans", "Fila_Origen"])
    ws2.append(cols)
    orden = {"Alta": 0, "Traslado": 1, "OtrasCausas": 2}
    eventos.sort(key=lambda e: (orden.get(e["tipo"], 9), str(e["pat"]), str(e["sub"])))
    for e in eventos:
        fila_out = [e["rut"], e["edad"], e["sexo"], TIPO_LABEL.get(e["tipo"], e["tipo"]),
                    e["pat"], e["sub"], e["falta_sub"]]
        fila_out += [e.get(fk, "") for fk in demo_keys]
        fila_out += [e["trans"], e["fila"]]
        ws2.append(fila_out)

    # presentación amigable
    from openpyxl.styles import Font
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"
    if ws2.max_row >= 1:
        ws2.auto_filter.ref = ws2.dimensions
    from openpyxl.utils import get_column_letter
    anchos = [14, 8, 8, 13, 44, 26, 13] + [11] * len(demo_keys) + [16, 11]
    for i, w in enumerate(anchos, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(salida)
    log(f"[ok] guardado: {salida}  (hoja '{NOMBRE_HOJA_SALIDA}')")

    n = {}
    for e in eventos:
        n[e["tipo"]] = n.get(e["tipo"], 0) + 1
    n_falta = sum(1 for e in eventos if e["falta_sub"] == "SI")
    log(f"[resumen] eventos de egreso: {len(eventos)}")
    if AVISAR_ALTA_SIN_SUBTIPO:
        log(f"          Altas sin subtipo (debiendo tenerlo): {n_falta}")
    for k in BUSQUEDAS:
        log(f"          {TIPO_LABEL.get(k,k)}: {n.get(k,0)}")

    # devolvemos un resumen para que la GUI arme el messagebox de confirmación
    return {
        "salida": str(salida),
        "total": len(eventos),
        "por_tipo": {TIPO_LABEL.get(k, k): n.get(k, 0) for k in BUSQUEDAS},
        "falta_subtipo": n_falta,
    }


# ╔═══════════════════════════════════════════════════════════════════╗
# ║  DESCRIPTORES PARA EL REGISTRO DE TAREAS (los consume autorem.py)   ║
# ╚═══════════════════════════════════════════════════════════════════╝
# El dispatcher (autorem.py) agrupa las tareas por 'reporte' de entrada: las que
# leen el mismo export se pueden correr juntas. Egresos e ingresos comparten el
# export IRIS 'Control de Salud Mental'.
REPORTE = {
    "id": "iris_salud_mental",
    "nombre": "IRIS · Control de Salud Mental",
}
TAREA = {
    "id": "a05_egresos",
    "nombre": "A05 · Egresos",
    "reporte": REPORTE["id"],
    "correr": procesar,          # (entrada, salida, log) -> resumen dict
    "hoja": NOMBRE_HOJA_SALIDA,
}
